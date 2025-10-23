#! python3

#LAO Functions Python3

import csv
from ctypes import windll, byref
import ctypes.wintypes as wintypes
import dicts
from os import environ
from os import system
from os import path
from os import remove
from sys import exit
import fun_text_date as td
from time import strftime
from time import sleep
from webs import getSeleniumDriver
from pprint import pprint

# Create dictionary of Area Code's State
def getAreaCodesDict():
	dAreaCodes = {}
	with open('{0}Area Codes by State.csv'.format(getPath('pyData')), 'r') as f:
		fout = csv.reader(f)
		# fout.next()
		next(fout)
		for row in fout:
			dAreaCodes[row[2]] = row[1]
	return dAreaCodes

# print dictionary key: value
def printDict(d):
	for i in d:
		print(i, d[i])

# Get Dictionary of agents & staff
def getAgentDict(dict_type='full', version='v2', skipFormerEmployees=True):
	# import csv
	if version == 'v2':
		# inFile = '{0}lao_staff_data_v2.csv'.format(getPath('pyData'))
		# inFile = 'F:/Research Department/Code/Research/data/lao_staff_data_v2.xlsx'
		inFile = 'F:/Research Department/Code/Databases/LAO_Staff_Db_v03.xlsx'
		dTemp = dicts.spreadsheet_to_dict(inFile)
		dAgents = {}
		for i in dTemp:
			if skipFormerEmployees:
				if dTemp[i]['LAO'] == 'No':
					continue
			name = dTemp[i]['Name']
			dAgents[name] = {}
			dAgents[name]['City'] = dTemp[i]['City']
			dAgents[name]['MC Audience'] = dTemp[i]['MC Audience']
			dAgents[name]['Email'] = dTemp[i]['Email']
			dAgents[name]['Id'] = dTemp[i]['Id']
			dAgents[name]['UserID'] = dTemp[i]['UserID']
			dAgents[name]['LAO'] = dTemp[i]['LAO']
			dAgents[name]['MC Aud ID'] = dTemp[i]['MC Aud ID']
			dAgents[name]['MC New Subscriber Market'] = dTemp[i]['MC New Subscriber Market']
			dAgents[name]['Markets'] = dTemp[i]['Markets']
			dAgents[name]['Office'] = dTemp[i]['Office']
			dAgents[name]['Phone'] =dTemp[i]['Phone']
			dAgents[name]['Phn Desk'] =dTemp[i]['Phn Desk']
			dAgents[name]['Phn Mobile'] =dTemp[i]['Phn Mobile']
			dAgents[name]['Phn Extention'] =dTemp[i]['Phn Extention']
			dAgents[name]['Roll'] = dTemp[i]['Roll']
			dAgents[name]['State'] = dTemp[i]['State']
			dAgents[name]['State2'] = dTemp[i]['State2']
			dAgents[name]['Zipcode'] = dTemp[i]['Zipcode']
			dAgents[name]['MktMailerMkt'] = dTemp[i]['MktMailerMkt']
			dAgents[name]['marketAbb'] = dTemp[i]['marketAbb']
			dAgents[name]['SF Password'] = dTemp[i]['SF Password']
			dAgents[name]['SF Token'] = dTemp[i]['SF Token']
			dAgents[name]['PC Name'] = dTemp[i]['PC Name']
	if dict_type == 'full':
		return dAgents
	# Return dict of MailChimp Report agent selection list
	elif dict_type == 'mc':
		lAgents = [] # Must create a list to sort before adding numbers
		for agent in dAgents:
			if dAgents[agent]['Roll'] == 'Agent' and dAgents[agent]['State2'] != 'SC':
				string = '{0}-{1:20} {2}'.format(dAgents[agent]['State2'], dAgents[agent]['MC Audience'], agent)
				# print(string)
				lAgents.append(string)
		lAgents.sort()
		d = {}
		i = 0
		for agent in lAgents:
			i += 1
			stri = str(i)
			d[stri] = agent
		return d
	elif dict_type == 'mcDict':
		dict = {}
		for agent in dAgents:
			d = dAgents[agent]
			# print(d)
			# td.uInput(dAgents[agent])
			dict[d['Email']] = d['MC Aud ID']
		return dict
	# Return dict of Terraforce Ids and agent name
	elif dict_type == 'tfid':
		d = {}
		for agent in dAgents:
			d[dAgents[agent]['Id']] = agent
		return d
	# Return dict of Email names and agent name
	elif dict_type == 'email':
		d = {}
		for agent in dAgents:
			d[dAgents[agent]['Email']] = agent
		return d
	# Return list of agent names
	elif dict_type == 'namesonly':
		l= []
		for agent in dAgents:
			if dAgents[agent]['Roll'] == 'Agent' and dAgents[agent]['LAO'] == 'Yes':
				l.append(agent)
		return l
	elif dict_type == 'modify':
		openFile(inFile)
	# elif dict_type == 'llr':
	# 	d = {}
	# 	#for agent in dAgents:
	elif dict_type == 'market':
		d = {}
		for agent in dAgents:
			d[dAgents[agent]['Email']] = dAgents[agent]['Markets']
		return d
	elif dict_type == 'researchers':
		d = {}
		for agent in dAgents:
			if dAgents[agent]['Roll'] == 'Research':
				d[agent] = dAgents[agent]['Id']
		return d
	elif dict_type == 'agentoffice':
		d = {}
		for agent in dAgents:
			if dAgents[agent]['Roll'] == 'Agent':
				d[agent] = dAgents[agent]['marketAbb']
		return d
	elif dict_type == 'marketlist':
		lMarkets = []
		for agent in dAgents:
			if not dAgents[agent]['marketAbb'] in lMarkets:
				lMarkets.append(dAgents[agent]['marketAbb'])
		return lMarkets
	# elif dict_type == 'officeaddress':
	# 	d = {}
	# 	for agent in dAgents:
	# 		r = dAgents[agent]
	# 		d[dAgents[agent]['Email']] = dAgents[agent]['Market']
	#	return d

# Get Agent TF ID (TFID=user TF 'UserID'=UserID)
def selectAgentReturnID(idtype='user'):
	dAgents = getAgentDict()
	lAgents = []
	# Create sorted list of agents
	for agent in dAgents:
		if dAgents[agent]['LAO'] == 'Yes':
			lAgents.append(agent)
			lAgents.sort()
	# Agent select list
	i = 1
	banner('Select Agent')
	for agent in lAgents:
		print(' {0:2}) {1}'.format(i, agent))
		i += 1
	ui = td.uInput('\n Select Agent > ')
	i = int(ui) - 1
	agentName = lAgents[i]
	agentId = dAgents[agentName][idtype]
	return agentId

# List of Classifications that are Lot Typ Raw Acerage
def rawAcerageClassification(classification):
	l = ['Agricultural', 'Church', 'Commercial', 'Data Center', 'Hospitality', 'Industrial', 'Medical', 'Mixed-Use', 'Office', 'Other', 'Public', 'Retail', 'Sand and Gravel', 'School', 'School Charter', 'Solar Wind', 'Speculative', 'Storage']
	for row in l:
		if classification == row:
			return True
	return False

#Format date to SalesForce format
def formatDateToSF(date): #formatSalesForceDate
	#In -> 09/20/2013
	#Return - > YYYY-MM-DDThh:mm:ss+hh:mm
	MO,DY,YR = date.split('/')
	if len(YR) == 2:	#change 2 digit year to 4 digit year
		YR = '20{0}'.format(YR)
	if len(MO) == 1:	#change 1 digit month to 2 digit month
		MO = '0{0}'.format(MO)
	if len(DY) == 1:	#change 1 digit day to 2 digit day
		DY = '0{0}'.format(DY)
	date = YR+'-'+MO+'-'+DY+'T00:00:00+00:00'
	return date

# Today's date in 'TF' (YYY-MM-DD), 'slash' (MM/DD/YYY) or 'ArcMap' (YYYY_MM_DD) formats
def todayDate(dateformat='TF', include_time=False):
	td.warningMsg('\n the lao.todayDate function has been replace with td.today_date')
	exit('\n Terminating program...')

def todayDateDifferenceDays(DATE):
	from datetime import date
	TODAY = date.today()
	lDATE = DATE.split('/')
	DATE = date(int(lDATE[2]), int(lDATE[0]), int(lDATE[1]))
	delta = TODAY - DATE
	return delta.days

# Returns current or last quarter date
def getDateQuarter(lastquarter=True, two_qrts_ago=False):
	from datetime import date
	d = date.today().isoformat()
	d = d.split('-')
	year = d[0]
	month = int(d[1])
	if month >= 1 and month <= 3:
		quarter = 1
	elif month >= 4 and month <= 6:
		quarter = 2
	elif month >= 7 and month <= 9:
		quarter = 3
	elif month >= 10 and month <= 12:
		quarter = 4
	if lastquarter:
		quarter -= 1
		if quarter == 0:
			year = int(year) - 1
			quarter = 4
	if two_qrts_ago:
		quarter -= 2
		if quarter == 0:
			year = int(year) - 1
			quarter = 4
		elif quarter == -1:
			year = int(year) - 1
			quarter = 3
	yearquarter = '{0}Q{1}'.format(year, quarter)
	return yearquarter

def OPR_Last_Update():
	from datetime import date
	today = date.today()
	# print(today.strftime('%m/%d/%Y'))
	while 1:
		inFile = '{0}OPR_Sent_Last_Date.csv'.format(getPath('comps'))
		dMarkets = dicts.spreadsheet_to_dict(inFile)
		banner('OPRs Last Updated')
		for row in dMarkets:
			daysDiff = todayDateDifferenceDays(dMarkets[row]['Last Update'])
			print('{0:>2}) {1:18}: {2:>3} days'.format(row, dMarkets[row]['Market'], daysDiff))

		ui = td.uInput("\n Select market to set with today's date or [Enter] to quit > ")

		if ui == '':
			return
		else:
			for row in dMarkets:
				if row == int(ui):
					dMarkets[row]['Last Update'] = today.strftime('%m/%d/%Y')
					break

		with open(inFile, 'w') as f:
			fout = csv.writer(f)
			fout.writerow(['Market', 'Last Update'])
			for row in dMarkets:
				lOut = [dMarkets[row]['Market'], dMarkets[row]['Last Update']]
				fout.writerow(lOut)

def todayAndMinusMonths(months, dateformat='TF'):
	from datetime import date
	from datetime import timedelta
	dnow = date.today().isoformat()
	b = 12
	months = months / float(b)
	dthen = (date.today() - timedelta(months * 365.25)).isoformat()
	if dateformat == 'TF':
		return dnow, dthen
	elif dateformat == 'slash':
		dnowlist = dnow.split('-')
		dthenlist = dthen.split('-')
		dnow = '{0}/{1}/{2}'.format(dnowlist[1], dnowlist[2], dnowlist[0])
		dthen = '{0}/{1}/{2}'.format(dthenlist[1], dthenlist[2], dthenlist[0])
		return dnow, dthen

# Clears the console and create a banner based on agrument
def banner(title):
	from colored import Fore, Back, Style
	system('cls')

	ban1 = '        '+('-' * (len(title)+4))
	ban2 = ' ------|'+(' ' * (len(title)+4))+'|------'
	ban3 = '  \\    |  '+title+'  |    /'
	ban4 = '   \\   |'+ (' ' * (len(title)+4)) + '|   /'
	ban5 = '   /    '+ ('-' * (len(title)+4)) + '    \\'
	ban6 = '  /      )'+ (' ' * (len(title))) + '(      \\'
	ban7 = ' --------'+ (' ' * (len(title))) + '  --------'
	ban3a = '  \\    |  '
	ban3b = '  |    /'

	# fore = '\033[94m'  # Cyan
	# print
	# print(f'{fore}{ban1}\033[00m')
	# print(f'{fore}{ban2}\033[00m')
	# print(f'{fore}{ban3a}\033[00m', end = '')
	# print(f'{title}', end = '')
	# print(f'{fore}{ban3b}\033[00m')
	# print(f'{fore}{ban4}\033[00m')
	# print(f'{fore}{ban5}\033[00m')
	# print(f'{fore}{ban6}\033[00m')
	# print(f'{fore}{ban7}\n\033[00m')
	# print
	# print
	# print('{0}{1}{2}'.format(Fore.dodger_blue_2, ban1, Style.reset))
	# print('{0}{1}{2}'.format(Fore.dodger_blue_2, ban2, Style.reset))
	# print('{0}{1}{2}'.format(Fore.dodger_blue_2, ban3a, Style.reset))
	# print('{0}'.format(title))
	# print('{0}{1}{2}'.format(Fore.dodger_blue_2, ban3b, Style.reset))
	# print('{0}{1}{2}'.format(Fore.dodger_blue_2, ban4, Style.reset))
	# print('{0}{1}{2}'.format(Fore.dodger_blue_2, ban5, Style.reset))
	# print('{0}{1}{2}'.format(Fore.dodger_blue_2, ban6, Style.reset))
	# print('{0}{1}{2}'.format(Fore.dodger_blue_2, ban7, Style.reset))
	print()
	print(f'{Fore.dodger_blue_2}{ban1}{Style.reset}')
	print(f'{Fore.dodger_blue_2}{ban2}{Style.reset}')
	print(f'{Fore.dodger_blue_2}{ban3a}{Style.reset}', end = '')
	print(f'{title}', end = '')
	print(f'{Fore.dodger_blue_2}{ban3b}{Style.reset}')
	print(f'{Fore.dodger_blue_2}{ban4}{Style.reset}')
	print(f'{Fore.dodger_blue_2}{ban5}{Style.reset}')
	print(f'{Fore.dodger_blue_2}{ban6}{Style.reset}')
	print(f'{Fore.dodger_blue_2}{ban7}\n{Style.reset}')
	print()

# User Input text
def uInput(text):
	# from colorama import Fore, Back, Style
	# import sty
	# from sty import fg, bg, ef, rs
	# try:
	# 	from colored import Fore, Back, Style
	# except ImportError:
	# 	from colored import fore
	# 	from colored import back as Back
	# 	from colored import style as Style
	# "\033[95m is cyan \033[00m is reset
	# ui = input(f'{Fore.LIGHTBLUE_EX}{text}{Style.RESET_ALL}')
	ui = input(f'\033[96m{text}\033[00m')
	ui = ui.strip()
	return ui

#Perform a mouse click at the specified x/y coordinates
def click(x='comwindow', y='None', mouseclick='Left'):
	try:
		from win32api import SetCursorPos, mouse_event
		from win32con import MOUSEEVENTF_LEFTDOWN, MOUSEEVENTF_LEFTUP, MOUSEEVENTF_RIGHTDOWN, MOUSEEVENTF_RIGHTUP
		if x == 'comwindow':
			x, y = 1200, 500
		sleep(1)
		SetCursorPos((x,y))
		if mouseclick == 'Left':
			mouse_event(MOUSEEVENTF_LEFTDOWN,x,y,0,0)
			mouse_event(MOUSEEVENTF_LEFTUP,x,y,0,0)
		elif mouseclick == 'Double':
			mouse_event(MOUSEEVENTF_LEFTDOWN,x,y,0,0)
			mouse_event(MOUSEEVENTF_LEFTUP,x,y,0,0)
			mouse_event(MOUSEEVENTF_LEFTDOWN,x,y,0,0)
			mouse_event(MOUSEEVENTF_LEFTUP,x,y,0,0)
		else:
			mouse_event(MOUSEEVENTF_RIGHTDOWN,x,y,0,0)
			mouse_event(MOUSEEVENTF_RIGHTUP,x,y,0,0)
	except:
		pass

# returns the type of company for quarterly sales review
def companyType(Company):
	Company = Company.upper()
	Company = Company.replace(',','')	#remove commas
	Company = Company.replace('.','')	#remove periods
	Company = Company.replace("'","")	#remove apostrophes 
	#Home Builders
	Names = [' HOME',' HOMES','A & B','AJF','ALTA VISTA COMMUNITIES ASNE','ASHTON WOODS','AT 92 LLC','AVATAR','BEAZER','BELLA CANYON','BELLAGO','BLANDFORD','BY TOWNE','CAVCO','COURTLAND','CUSTOM DREAM','CACHET','CAMELOT','CENTEX','CRESLEIGH','DEL WEBB','DVP','ELLIOT','EVERGREEN','FARNSWORTH','FHC','FREO ARIZONA','FULTON','GEHAN','GEMCOR','GOLD MOUNTAIN 22, LLC','HCZ','HEARTHSTONE','ARCADIA 1','HORTON','HOVNANIAN','JOSEPH CARL','KB ','KEYSTONE','LENNAR','LEXIN','LGI','MARACAY','MATTAMY','MERITAGE','MINT','MONARCH','MONTEREY','NEIDHART','PULTE','REGENCY','RICHMOND AMERICAN','ROBSON','ROSEWOOD','SHEA','SKR CONSTRUCTION','STANDARD PACIFIC','T3-CS','TAYLOR MORRISON','TESORO ESTATES LLC','TOLL BROTHERS','TOWNE DEVELOPMENT','TREND','TURNER','V82 HOLDINGS','VISTANCIA','VITALIA','WEEKLEY','WILLIAM LYON','WILLIAM RYAN','WILSON PARKER','WOODSIDE']
	if any(x in Company for x in Names):
		CoType = 'HOME BUILDER'
		return CoType
	#MultiFamily
	Names = ['APARTMENT','AVONDALE MHP','CAMDEN USA','CONDO','ECHO DEVELOPMENT','ELLSWORTH & US 60 LLC','GRE INDIGO','IS & 104','JLB CAMELBACK','MARK TAYLOR','MARK-TAYLOR','MCCW BUILDERS','NEW LEAF COMMUNITIES','OAKWOOD PLAZA','PB BELL','RANCH MIRAGE 98','ROOSEVELT PARK 52','SOUTH MOUNTAIN LOFTS','TEMPE 601 W RIO SALADO','VEDURA CHANDLER MANAGER','URBAN 30 II','VILLAGES AT SAN TAN','WALH']
	if any(x in Company for x in Names):
		CoType = 'MULTIFAMILY'
		return CoType
	#Land Banks
	Names = ['ACACIA SAN TAN','BB PROPERTY ','BRIMET II','COMMUNITY DEVELOPMENT CAPITAL GROUP','CMG 167','CMG 900','CORONA JOHNSON RANCH','CSW NS LAND 1','COPPER STAR BANK','CORONA JOHNSON RANCH','CSW NS LAND 1','DOA PROPERTIES','EN, LLC','ESTRELLA VISTA','GGW FUNDING','GSO ','HERITAGE INVESTMENT HOLDINGS','JEN ARIZONA','KLONDIKE LAND PORTFOLIO','MISSION LAND','MPC-VILLAGES AT WESTRIDGE PARK','NNP III','NSHE ','PPP DEVELOPMENT','RES-AZ','RHD 3','RRE I','SCC Canyon II','SEP TRIO','SILVER CANYON INVESTMENTS','SLV MAGMA HOLDINGS','TRIPLE G DEVELOPMENT','T W LEWIS-LAND HOLDING','TW LEWIS-LAND HOLDING','TW LEWIS-LAND HOLDINGS','T W LEWIS BLACKSTONE','TW LEWIS SEVILLE','T W LEWIS TORREMAR II','VISTANCIA SOUTH','WHGR']
	if any(x in Company for x in Names) and Company != 'BANKS':
		CoType = 'LAND BANK'
		return CoType
	#Banks
	Names = [' BK','BANK', 'SAVINGS','CHASE','CREDIT UNION','THRIFT','WELLS FARGO']
	if any(x in Company for x in Names):
		CoType = 'BANK'
		return CoType
	# Government
	Names = ['CITY OF ','GOVERNMENT','INDIAN COMMUNITY','INDIAN RESERVATION','MARICOPA COUNTY','PINAL COUNTY']
	if any(x in Company for x in Names):
		CoType = 'GOVNERNMENT'
		return CoType
	Names = ['ACADEMY','ARIZONA STATE UNIVERSITY','CHARTER AIM','COLLEGE','LEARNING CENTER','NEWCO GCC','SCHOOL','SCHOOLS','UNIVERSITY OF PHOENIX']
	if any(x in Company for x in Names):
		CoType = 'SCHOOL'
		return CoType
	Names = ['APS','ARIZONA PUBLIC SERVICE COMPANY','ARIZONA WATER COMPANY','COMMUNICATIONS','COX']
	if any(x in Company for x in Names):
		CoType = 'UTILITY'
		return CoType
	#Commercial
	Names = ['ARIZONA ENERGY','ARIZONA MACHINERY GROUP','ATV DOGFIGHTING','AUTO ',' AUTO','AVIATION','BAKERIES','BASHAS','CAR WASH',' CARS','CHICANOS POR LA CAUSA','CIRCLE K','CHEMICALS','CHICK-FIL-A','CN METAL','COMBUSTION ARTS','COPERNICUS ENERGY','CYRUS ONE','DIAMOND SHAMROCK','DRAINAGE','DISTRIBUTING','EMPIRE CATERPILLAR','EQUIPMENT','FAMILY DOLLAR','FINANCE','FOOD','GALVANIZING','HNM LLC','INSURANCE','INTERNATIONAL','LEAKY SHED STUDIO','MACHINE','MACYS','MAGMA DEVCO','MCDONALDS REAL ESTATE COMPANY','MINING','MR TANNER CONSTRUCTION','MURPHY OIL CORPORATION','NATIONAL OILWELL VARCO','NURSERY','NURSERIES','PARK N SWAP','PARK LUCERO','PARKING','QUIKTRIP','QUICKTRIP','RECURRENT ENERGY','RESIDENTIAL','RESOLUTION COPPER MINING','RAPID-SPAN','RESTAURANTS','RETAIL','SECTION','SELF STORAGE','SOLAR','SPORTING','SPORTS','STORES','SULLIVAN CONSTRUCTION','SUPPLIES','TIERRASCAPE IMPROVEMENT',' TIRE','TITLE SECURITY','TOWER CRANES',' TRADING','WARD NORTH AMERICAN','WARHOUSE','WESTERN STATES DECKING','WASSER & WINTERS','WHOLESALE','WILDLIFE WORLD ZOO']
	if any(x in Company for x in Names):
		CoType = 'COMMERCIAL'
		return CoType
	#Agricultural
	Names = ['AUZA RANCHES','CATTLE','DAIRY',' FARM','GRAIN','JOE AUZA SHEEP CO','NORTHSIDE HAY COMPANY','ORGANIC','RS CAVALLO','WESTERN MILLING']
	if any(x in Company for x in Names):
		CoType = 'AGRICULTURAL'
		return CoType
	#Medical
	Names = ['ACADEMIES',' ASSISTED LIVING','CANCER','HEALTH','HOSPICE','HOSPITAL','MEDICAL','SUNRISE LUXURY LIVING','WELLNESS CENTER']
	if any(x in Company for x in Names):
		CoType = 'MEDICAL'
		return CoType
	#NONPROFIT
	Names = ['CHURCH',' CHRIST','CHRIST ','CHRISTS ','CORPORATION OF THE PRESIDING BISHOP','FELLOWSHIP',' GOD','HABITAT FOR HUMANITY','KING OF KINGS PCA',' LDS','YOUTH CENTER']
	if any(x in Company for x in Names):
		CoType = 'NONPROFIT'
		return CoType
	#Investor
	Names = ['41ST 4800','ABCW','AGRO-IRON','ARCUS','AREAD','ASSOCIATES','BELL 115','BELLOC FAMILY','BERGER COMPANY','BHHS','BORGATA AT SAN TAN HEIGHTS','CAPITAL','CAHAVA SPRINGS CORPORATION','CDCG','CPGC','CHENG 90','COMMUNITIES SOUT','CR 26','DC LAND','DEL ESTE','DEVELOPMENT','DMB','ESTATE OF','EW GARDNER FAMILY','EQUITY','FALLS AT GILBERT','FINANCIAL','FOOTHILL CAPITAL','FOOTHILL MOUNTAIN','FOUNDATION','FUNDING','GROUP','HDP GOLF RIDGE LLC','HHB ','HOLDING','HOLDINGS','HOWARD HAWKS','INVESTING','INVESTMENT','INVESTOR','JD CAMPBELL','JOAN LEVINSON','KD DEVELOPMENT','KEMF WP','LITCHFIELD 5','LREP ARIZONA','MACDONALD','MBR LAND','MCRAE','MESQUITE MTN RANCH INVEST','METROPOLITAN LAND CO','MJFDRT REALTY','MTL 1 LLC','NEWLAND','PAR CAP','PARTNER','PARTNERS','PARTNERSHIP','PINNACLE PHX','PLAINS ROAD','PLHAZ','PORLIER 1945','PROPERTY','PROPERTIES','REIT','RJAL','RMG','ROMA QC','RRS & COMPANY','SKYBRIDGE QUAIL','SLV HOMESTEAD','TASILLO','TERRA WEST COMMUNITIES','TRES POINTS','TRUST','V AT ARIZONA','VALENCIA HEIGHTS','VDV RECOVERY','VENTURE','WALTON ARIZONA','WOLFSWINKEL']
	if any(x in Company for x in Names):
		CoType = 'INVESTOR'
		return CoType
	#Other Company
	Names = ['ADVISORS','ASSOCIATED','BUILDERS','CAR WASH','CARIOCA CO','CAS','CCF SILVA','COMPANIES','COMPANY','COMMUNITIES','COMMUNITY','CONSTRUCTION','CONSULTANTS','CORP','DESIGN BUILD','DEVELOPMENT',' ENTERPRISE','FINANCIAL SERVICES','FUND',' INC','INDUSTRIES','IPX','LLC','LIFE INSURANCE','LAND CO',' LP',' LLP',' LLLP','LUXURY LIVING','MANAGEMENT','MARKETPLACE ONE','PLHAZ','REAL ESTATE','REALTY','RESIDENTIAL','RETREAT','RMG','SENIOR LIVING','SECTION']
	if any(x in Company for x in Names):
		CoType = 'OTHER'
		return CoType
	CoType = 'PERSON'
	return CoType

# Convert 2 letter state code to full state name and vice versa
def convertState(state):
	if len(state) > 2:
		state = state.title()
	dStateAbbr = {
		'AK': 'Alaska',
		'AL': 'Alabama',
		'AZ': 'Arizona',
		'AR': 'Arkansas',
		'CA': 'California',
		'CO': 'Colorado',
		'CT': 'Connecticut',
		'DE': 'Delaware',
		'FL': 'Florida',
		'GA': 'Georgia',
		'HI': 'Hawaii',
		'ID': 'Idaho',
		'IL': 'Illinois',
		'IN': 'Indiana',
		'IA': 'Iowa',
		'KS': 'Kansas',
		'KY': 'Kentucky',
		'LA': 'Louisiana',
		'ME': 'Maine',
		'MD': 'Maryland',
		'MA': 'Massachusetts',
		'MI': 'Michigan',
		'MN': 'Minnesota',
		'MS': 'Mississippi',
		'MO': 'Missouri',
		'MT': 'Montana',
		'NE': 'Nebraska',
		'NV': 'Nevada',
		'NH': 'New Hampshire',
		'NJ': 'New Jersey',
		'NM': 'New Mexico',
		'NY': 'New York',
		'NC': 'North Carolina',
		'ND': 'North Dakota',
		'OH': 'Ohio',
		'OK': 'Oklahoma',
		'OR': 'Oregon',
		'PA': 'Pennsylvania',
		'RI': 'Rhode Island',
		'SC': 'South Carolina',
		'SD': 'South Dakota',
		'TN': 'Tennessee',
		'TX': 'Texas',
		'UT': 'Utah',
		'VT': 'Vermont',
		'VA': 'Virginia',
		'WA': 'Washington',
		'WV': 'West Virginia',
		'WI': 'Wisconsin',
		'WY': 'Wyoming',
		'Alaska': 'AK',
		'Alabama': 'AL',
		'Arizona': 'AZ',
		'Arkansas': 'AR',
		'California': 'CA',
		'Colorado': 'CO',
		'Connecticut': 'CT',
		'Delaware': 'DE',
		'Florida': 'FL',
		'Georgia': 'GA',
		'Hawaii': 'HI',
		'Idaho': 'ID',
		'Illinois': 'IL',
		'Indiana': 'IN',
		'Iowa': 'IA',
		'Kansas': 'KS',
		'Kentucky': 'KY',
		'Louisiana': 'LA',
		'Maine': 'ME',
		'Maryland': 'MD',
		'Massachusetts': 'MA',
		'Michigan': 'MI',
		'Minnesota': 'MN',
		'Mississippi': 'MS',
		'Missouri': 'MO',
		'Montana': 'MT',
		'Nebraska': 'NE',
		'Nevada': 'NV',
		'New Hampshire': 'NH',
		'New Jersey': 'NJ',
		'New Mexico': 'NM',
		'New York': 'NY',
		'North Carolina': 'NC',
		'North Dakota': 'ND',
		'Ohio': 'OH',
		'Oklahoma': 'OK',
		'Oregon': 'OR',
		'Pennsylvania': 'PA',
		'Rhode Island': 'RI',
		'South Carolina': 'SC',
		'South Dakota': 'SD',
		'Tennessee': 'TN',
		'Texas': 'TX',
		'Utah': 'UT',
		'Vermont': 'VT',
		'Virginia': 'VA',
		'Washington': 'WA',
		'West Virginia': 'WV',
		'Wisconsin': 'WI',
		'Wyoming': 'WY'
	}
	state = state.strip()
	return dStateAbbr[state]

# Check if company name is a company type is company type
def coTF(TEXT):
	TEXT = TEXT.upper()
	# print(TEXT)
	with open('{0}company_type_name_list.txt'.format(getPath('pyData')), 'r') as f:
		for row in f:
			if row.strip('\n') in TEXT:
				return True
		if TEXT[:4] == 'THE ':
			return True
		elif ' THE ' in TEXT:
			return True
		elif ', THE' in TEXT:
			return True
		elif TEXT[-5:] == ' CITY':
			return True
		elif TEXT[-3:] == ' CO':
			return True
		elif TEXT[-4:] == ' THE':
			return True
		elif TEXT[-3:] == ' JT':
			return True
		elif TEXT[-3:] == ' TR':
			return True
		elif TEXT[-3:] == ' RS':
			return True
		elif TEXT[-4:] == ' DEV':
			return True
		# elif TEXT[-5:] == ' PTRS':
		# 	return True
		else:
			return False

# Formats large numbers to comma & decimal thousdands
def thousandsFormat(num):
	numComma = '{:,}'.format(num)
	lnumComma = numComma.split(',')
	if len(lnumComma) == 2:
		numFormatted = '{0}.{1}'.format(lnumComma[0], lnumComma[1][:1])
	else:
		numFormatted = '{0},{1}.{2}'.format(lnumComma[0], lnumComma[1], lnumComma[2][:1])
	return numFormatted

#pass a int or decimal currency and return $###,###
def curFormat(NUMBER):
	td.warningMsg('\n lao.curFormat is now td.currency_format_from_number')
	print(f'\n NUMBER: {NUMBER}')
	exit('\n Terminating program...')
	# if NUMBER == 'None':
	# 	return 'None'
	# NUMBER = int(float(NUMBER))
	# NUMBER = '$'+'{:,.0f}'.format(NUMBER)
	# return NUMBER

# Format date types TF, slash, dash, 6 characters, 8 characters
def dateFormat(date, outformat='TF', informat='unknown'):
	td.warningMsg('\n lao.dateFormat is now td.date_engine')
	print(f'\n Date: {date}')
	exit('\n Terminating program...')
	# if informat == 'unknown':
	# 	if '12:00:00' in date and '/' in date:
	# 		informat = 'CoStar'
	# 	elif '/' in date:
	# 		informat = 'slash'
	# 	elif '-' in date:
	# 		if len(date) == 10 and (date[:2] == '19' or date[:2] == '20'):
	# 			informat = 'TF'
	# 		else:
	# 			informat = 'dash'
	# 	elif len(date) == 6:
	# 		if date[-4:-2] == '20' or date[-4:-2] == '19':
	# 			informat = 'monthyear'
	# 		else:
	# 			informat = 'six'
	# 	elif len(date) == 8:
	# 		informat = 'eight'
	# 	elif date == '':
	# 		informat = 'today'
	# if informat == 'slash':
	# 	lDate = date.split('/')
	# elif informat == 'dash':
	# 	lDate = date.split('-')
	# elif informat == 'CoStar':
	# 	date = date.replace(' 12:00:00 AM', '')
	# 	lDate = date.split('/')
	# elif informat == 'TF':
	# 	lDateTF = date.split('-')
	# 	lDate = [lDateTF[1], lDateTF[2], lDateTF[0]]
	# elif informat == 'monthyear':
	# 	lDate = [date[:2], '01', date[4:]]
	# elif informat == 'six' or informat == 'eight':
	# 	lDate = [date[:2], date[2:4], date[4:]]
	# elif informat == 'today':
	# 	from datetime import date
	# 	today = date.today()
	# 	return today.strftime('%m-%d-%Y')
	# else:
	# 	print('Unknown format')
	# 	return 'error unknown format'

	# if len(lDate[0]) == 1:
	# 	lDate[0] = '0{0}'.format(lDate[0])
	# if len(lDate[1]) == 1:
	# 	lDate[1] = '0{0}'.format(lDate[1])
	# if len(lDate[2]) == 2:
	# 	lDate[2] = '20{0}'.format(lDate[2])

	# if outformat == 'TF':
	# 	date = '{0}-{1}-{2}'.format(lDate[2], lDate[0], lDate[1])
	# elif outformat == 'slash':
	# 	date = '{0}/{1}/{2}'.format(lDate[0], lDate[1], lDate[2])
	# elif outformat == 'dash':
	# 	date = '{0}-{1}-{2}'.format(lDate[0], lDate[1], lDate[2])

	# return date

# Format any phone number to (###) ###-#### Ext ### (default)
def phoneFormat(phone, phoneFormat='default'):
	td.warningMsg('\n lao.phoneFormat is now td.phoneFormat')
	print(f'\n Phone: {phone}')
	exit('\n Terminating program...')

# select a file to open using the windows gui
def guiFileOpen(path='F:/Research Department/Code/RP3/data/', titlestring='', extension=[('csv files', '.csv'), ('txt files', '.txt'), ('Excel files', '.xlsx'), ('all files', '.*')]):
	print('\n Select file from GUI...')
	import tkinter as tk
	from tkinter import Tk
	from tkinter import filedialog as fd

	filename = fd.askopenfilename(initialdir=path, filetypes=extension, title=titlestring) # show an "Open" dialog box and return the path to the selected file
	if filename == '':
		td.warningMsg('\n No selection made...terminating program...')
		from sys import exit
		exit()
	return filename

# Save file as gui
def guiFileSaveAs(path = 'F:/Research Department/Code/Research/data/CompsFiles/', titlestring = '', extension = [('csv files', '.csv'), ('all files', '.*')]):
	import tkinter as tk
	from tkinter import filedialog as fd

	filename = fd.asksaveasfilename(initialdir=path, filetypes=extension, title=titlestring) # show an "Open" dialog box and return the path to the selected file
	return filename

#clears the console screen and prints a blank line
def lineFeed():
	system('cls')
	print

# Returns the current line number in our program.
def lineno():
	from inspect import currentframe
	return currentframe().f_back.f_lineno

#convert PUC from number to definition
def LookupPUC(PUC):
	PUC1 = PUC
	PUC2 = PUC[:2] + '**'
	with open('F:\\Research Department\\scripts\\Data\\puc_maricopa.csv', 'r') as f:
		PUCLOOKUP = csv.reader(f)
		for row in PUCLOOKUP:
			if PUC == row[0]:
				PUC1 = row[1]
			if PUC2 == row[0]:
				PUC2 = row[1]
	if PUC == PUC1:
		return PUC2
	else:
		return PUC1

# Calculates Acres based on lot demintions and count
def lotAcresCalc():
	banner('Lot Acres Calculator')
	ui = input(' Enter Count, Width & Depth separated by spaces.\n > ')
	acres = 0
	while 1:
		data = ui.split()
		acres = round(acres + ((float(data[1]) * float(data[2]) * float(data[0])) / 43560), 2)
		ui = input(' Enter more Lots or [Enter] to Total.\n > ')
		if ui == '':
			return acres

# Clears the console and creates a title based on argument
def message(title):
	print
	print('*' * (len(title)+6))
	print('*'+ (' ' * (len(title)+4)) + '*')
	print('*  '+title+'  *')
	print('*'+ (' ' * (len(title)+4)) + '*')
	print('*' * (len(title)+6))
	print

# converts text into standard format like T. W. to TW
def nameStandards(text):
	if 'K HOVNANIAN' in text: text = text.replace('K H','K. H')
	if 'T.W. L' in text: text = text.replace('T.W.','TW')
	if 'T. W. L' in text: text = text.replace('T. W.','TW')
	text = text.replace(',','')
	text = text.replace("'","")
	return text

# opens url in browser
def openURL(URL):
	from webbrowser import open as openbrowser
	openbrowser(URL)

# opens url in browser
def openbrowser(url):
	from webbrowser import open as openbrow
	openbrow(url)

# parses and address into street, city, state and zip
def parseAddress(ADDRESS):
	ADDRESS = ADDRESS.upper()
	str = ADDRESS.splitlines()
	ST1 = str[0]
	CSZ = str[1]
	if ',' in CSZ:
		CSZ = CSZ.split(',')
		CTY = CSZ[0]
		SZC = CSZ[1]
		SZC = SZC.split()
		STT = SZC[0]
		ZIP = SZC[1]
	else:
		print(CSZ)
		CTY = (td.uInput(' Enter City  > ')).upper()
		STT = (td.uInput(' Enter State > ')).upper()
		ZIP = CSZ[-5:]
	return ST1, CTY, STT, ZIP

def chooseFromList(lLIST, FULLLIST='None', LABEL='Item'):
	print('\nFix {0}\n'.format(LABEL))
	if FULLLIST == 'None':
		lLIST.insert(0, 'No Match')
	else:
		lLIST.insert(0, FULLLIST)
	lenLIST = len(lLIST)
	for i in range(0, lenLIST):
		print('{0:2})  {1}'.format(i, lLIST[i]))
	while 1:
		ui = td.uInput('\n Select {0} > '.format(LABEL))
		if ui == '0':
			return FULLLIST
		else:
			return lLIST[int(ui)]

def parceMailChimpPerson(FIRST, LAST):
	# print('Fix Name\n')
	# print('First Name: {0}'.format(FIRST))
	# print('Last Name:  {0}'.format(LAST))
	FIRST = FIRST.replace('.', '').replace("'", '').replace(',', '')
	LAST = LAST.replace('.', '').replace("'", '').replace(',', '')
	lSUFFIX = ['JR', 'SR', 'II', 'III', 'IV', 'V']
	if ' ' in FIRST:
		lInitialCheck = FIRST.split(' ')
		for initial in lInitialCheck:
			if len(initial) == 1:
				print(FIRST)
				FIRST = FIRST.replace(initial, '').strip()

	if ' ' in LAST:
		# Has Suffix
		lSUFFIX = [' JR', ' SR', ' II', ' III', ' IV', ' V']
		for suffix in lSUFFIX:
			if suffix in LAST:
				return FIRST, LAST

		lInitialCheck = LAST.split(' ')
		for initial in lInitialCheck:
			if len(initial) == 1:
				print(LAST)
				LAST = LAST.replace(initial, '').strip()

	# if ' ' in FIRST:
	# 	lFIRST = chooseFromList(FIRST.split(' '), FIRST, 'First Name')
	# if ' ' in LAST:
	# 	lLAST = chooseFromList(LAST.split(' '), LAST, 'Last Name')
	return FIRST, LAST

def SalesForceCSVClean(file):
	from shutil import copyfile
	input = open(file, 'r')
	output = open('C:/TEMP/eraseme.csv', 'w', newline='')
	writer = csv.writer(output)
	for row in csv.reader(input):
		if row:
			writer.writerow(row)
		else:
			break
	input.close()
	output.close()
	copyfile('C:/TEMP/eraseme.csv',file)

# Clears the console and create a title based on argument
def titler(title):
	starline = '*' * (len(title)+6)
	blankline = ' ' * (len(title)+4)
	system('cls')
	print
	print(' {0}'.format(starline))
	print(' *{0}*'.format(blankline))
	print(' *  ' + title + '  *')
	print(' *{0}*'.format(blankline))
	print(' {0}'.format(starline))
	print
	# print('*' * (len(title)+6))
	# print('*' + (' ' * (len(title)+4)) + '*')
	# print('*  '+title+'  *')
	# print('*'+ (' ' * (len(title)+4)) + '*')
	# print('*' * (len(title)+6))
	# print

# Highlights the input text
def highlight(title):
	print
	print('*' * (len(title)+6))
	print('*'+ (' ' * (len(title)+4)) + '*')
	print('*  '+title+'  *')
	print('*'+ (' ' * (len(title)+4)) + '*')
	print('*' * (len(title)+6))
	print

# remove commas and periods from text
def pcr(text):
	text = text.replace(',','').replace('.','').replace("'","")
	return text

# Add or remove comma to try and get a match in TF
def coComma(text):
	text = text.upper()
	if ' LLC' in text:
		if ', LLC' in text:
			text = text.replace(', LLC',' LLC')
		else:
			text = text.replace(' LLC',', LLC')
	if ' INC' in text:
		if ', INC' in text:
			text = text.replace(', INC',' INC')
		else:
			text = text.replace(' INC',', INC')
	return text

def upperCaseTitleCase(text):
	text = text.replace(' Llc', ' LLC')
	text = text.replace(' Lp', ' LP')
	text = text.replace(' Llp', ' LLP')
	text = text.replace(' Lllp', ' LLLP')
	text = text.replace('Nwc ','NWC ')
	text = text.replace('Nec ','NEC ')
	text = text.replace('Swc ','SWC ')
	text = text.replace('Sec ','SEC ')
	return text

# find & return city, state & country based on user input zip code
def zipCodeFindCityStateCountry(ZIP):
	CTY, STT, USA = 'None', 'None', 'None'
	ZIP = ZIP.upper().strip()
	if '-' in ZIP:
		lZIP = ZIP.split('-')
		ZIP = lZIP[0]
	elif len(ZIP) == 9:
		ZIP = ZIP[:5]
	# ZIP = ZIP.replace('-', '')
	
	# td.uInput(ZIP)
	ZIPFILE = 'F:/Research Department/Code/Research/data/north_american_zipcodes.csv'
	with open(ZIPFILE) as Z:
		ZF = csv.reader(Z)
		for row in ZF:
			if row[0] == ZIP:
				CTY = (row[2]).title()
				STT = row[3]
				USA = 'United States'
				for i in ZIP:
					if i.isalpha():
						USA = 'Canada'
				break
	return CTY, STT, USA

# Create Parcel numbers from Lot Numbers
def LotNoToParcelNo(APN, BKMP, P1):
	PRL = APN
	P2 = int(P1)
	banner('Lot Number to APN Calculator')
	firstlot = (input('Enter the first lot number > '))
	nextlot = 'no value'
	while nextlot != 'D':

		ui = (td.uInput(' Enter the next lot number or [F]inish > ')).upper()

		if ui == 'F' or ui == 'N':
			print('Fin...')
			return PRL
		else:
			nextlot = int(ui)

		lotdiff = nextlot - firstlot

		# check if lotdiff is greater than 1 and if it is ask if the difference is continuous or a jump
		if lotdiff > 1:

			ui = (td.uInput(' [C]ontinuous or [J]ump? > ')).upper()
			if ui == 'C':
				R1 = P2 + 1
				R2 = P2 + lotdiff + 1
				P2 = R2 - 1
				for i in range(R1, R2):
					PRL = '%s, %s%03d' % (PRL, BKMP, i)
					print('i = %d' % i)
			elif ui == 'J':
				PRL = '%s, %s%03d' % (PRL, BKMP, P2 + lotdiff)
				P2 = P2 + lotdiff
		else:
			P2 = P2 + lotdiff
			PRL = '%s, %s%d' % (PRL, BKMP, P2)

		firstlot = nextlot
		# print(P2)

		banner('Lot Number to APN Calculator')
		print('Last Lot Number: %d' % firstlot)

# User to select Classification
def chooseClassification(currentClass='None'):
	lClassification = dict.get_classification_list()
	# lClassification = [
	# 					'blank',
	# 					'Residential',
	# 					'Commercial',
	# 					'Industrial',
	# 					'Office',
	# 					'Retail',
	# 					'Agricultural',
	# 					'Apartment Horizontal',
	# 					'Apartment Traditional',
	# 					'Build for Rent (platted)',
	# 					'Church',
	# 					'Conservation', 
	# 					'High Density Assisted Living',
	# 					'High Density Residential',
	# 					'Hospitality',
	# 					'Manufactured Home',
	# 					'Master Planned Community',
	# 					'Medical',
	# 					'Mixed-Use',
	# 					'Not for Sale',
	# 					'Other',
	# 					'Public',
	# 					'Ranch and Recreation',
	# 					'Residence Included',
	# 					'Resort or Golf',
	# 					'School',
	# 					'School Charter',
	# 					'Solar Wind',
	# 					'Speculative',
	# 					'Storage',
	# 					'Timber'
	# 					]

	lenClassList = len(lClassification)

# User to select Development Status for Aparments & BTR
def choose_Development_Status():
	# lDevStat = [
	# 				'blank',
	# 				'Planned',
	# 				'Proposed',
	# 				'Stablized',
	# 				'Under Construction']
	print_function_name(' [lao def choose_Development_Status]')
	lDevStat = dict.get_development_status_list()
	lenDevStatList = len(lDevStat)

	while 1:
		banner('Select Development Status')
		print('\n    - Development Status -\n')
		for i in range(1, lenDevStatList):
			print(' {0:>2}) {1}'.format(i, lDevStat[i]))
		ui = td.uInput('\n Select Development Status > ').upper()
		try:
			if ui == '' or int(ui) < 1 or int(ui) > (lenDevStatList - 1):
				td.warningMsg(' Invalid input...try again...')
				sleep(1)
				continue
		except ValueError:
			td.warningMsg(' Invalid input...try again...')
			sleep(1)
			continue
		print('\n Development Status Selected')
		td.warningMsg('\n {0}'.format(lDevStat[(int(ui))]))
		uiConfirm = td.uInput('\n Use this Status? [0/1] > ')
		if uiConfirm != '1':
			print('\n Try again...')
			sleep(1)
			continue

		DEVSTAT = lDevStat[(int(ui))]
		if DEVSTAT == 'Leave Blank':
			DEVSTAT = 'None'
		
		return DEVSTAT

# Ask for Number of Lots if Classification is housing
def askForNumberOfLots(classification):
	lClassificationsWithLots = [
		'Residential',
		'Apartment Horizontal',
		'Apartment Traditional',
		'Build for Rent (platted)',
		'High Density Assisted Living',
		'High Density Residential',
		'Mixed-Use',
	]
	if classification in lClassificationsWithLots:
		return True
	else:
		return False

# User to select Lot Type
def chooseLotType(classification=['None'], lottype=None):
	if not type(classification) is list:
		classification = [classification]
	# List of Raw Acerage Classifications
	lRawAc = ['Agricultural', 'Church', 'Commercial', 'Hospitality', 'Industrial', 'Medical', 'Mixed-Use', 'Office', 'Office;Retail', 'Other', 'Public', 'Retail', 'School', 'School Charter', 'Solar Wind', 'Speculative', 'Storage']
	if classification[0] in lRawAc:
		return 'Raw Acreage'
	# List of Lot Types
	lLotType = ['blank', 'Covered Land', 'Finished Lots', 'Initial Lot Option', 'Partially Improved', 'Platted and Engineered', 'Raw Acreage']
	if lottype== [] or lottype == '' or lottype == None or lottype == 'None':
		pass
	else:
		ui = td.uInput('\n Change Type: {0} [0/1/00] > '.format(lottype))
		if ui == '0':
			return lottype
		elif ui == '00':
			exit('\n Terminating program...')
	while 1:
		print('\n    - Lot Type -\n')
		for i in range(1, 7):
			print('{0:>2}) {1}'.format(i, lLotType[i]))
		ui = td.uInput('\n Select Lot Type > ')
		if ui == '' or (int(ui)) < 1 or (int(ui)) > 6:
			td.warningMsg('\n Select a Lot Type...try again...')
			continue
		LTY = lLotType[(int(ui))]
		return LTY

# User to select the value of Buyer Acting As field
def chooseBuyerActingAs():
	lotTypeList = ['blank', 'Homebuilder', 'Inv/Dev', 'Lot Banker', 'User']
	while 1:
		print('\n    - Buyer Acting As -\n')
		for i in range(1, 5):
			print(' {0:>2}) {1}'.format(i, lotTypeList[i]))
		ui = td.uInput('\n Select Buyer Acting As > ')
		if int(ui) >= 1 and int(ui) <= 4:
			BAA = lotTypeList[(int(ui))]
			return BAA
		else:
			td.warningMsg('\n Invalid input...try again...')

def chooseSource():
	print(' 1) Affidavit')
	print(' 2) CoStar')
	print(' 3) LAO')
	print(' 4) MLS')
	print(' 5) RED News')
	print(' 6) Reonomy')
	print(' 7) Vizzda')
	while 1:
		ui = td.uInput('\n Enter Source > ')
		if ui == '1':
			source = 'County Recorder'
			break
		elif ui == '2':
			source = 'Costar'
			break
		elif ui == '3':
			source = 'LAO'
			break
		elif ui == '4':
			source = 'MLS'
			break
		elif ui == '5':
			source = 'RED News'
			break
		elif ui == '6':
			source = 'Reonomy'
			break
		elif ui == '7':
			source = 'Vizzda'
			break
		else:
			td.warningMsg('\n Invalid input...try again...')
	if ui != '3':
		sourceID = td.uInput('\n Enter Source ID > ')
	else:
		sourceID = ''
	return source, sourceID

# User to select from List Classification, Lot Type, Development Status, Buyer Acting As
def select_from_list(dTF, tf_field):

	start_over = False

	# Assign variable list(s) and title
	if tf_field == 'Classification__c':
		title = 'Classification'
		lList = dicts.get_classification_list()

	elif tf_field == 'Lot_Type__c':
		# Lot Type is always Raw Acreage if not Residential or Apartment
		lList = dicts.get_non_raw_acreage_classifications_list()
		if dTF['Classification__c'] not in lList:
			dTF['Lot_Type__c'] = 'Raw Acreage'
			return dTF,start_over
		# Not Raw Acreage so get Lot Type list
		title = 'Lot Type'
		lList = dicts.get_lot_type_list()

	elif tf_field == 'Development_Status__c':
		# Development Status only appies to Multifamily
		lList = dicts.get_multifamily_classifications_list()
		if dTF['Classification__c'] not in lList:
			dTF['Development_Status__c'] = 'None'
			return dTF, start_over
		# Is Multifamily so get Development Status list
		title = 'Development Status'
		lList = dicts.get_development_status_list()

	elif tf_field == 'Buyer_Acting_As__c':
		title = 'Buyer Acting As'
		lList = dicts.get_buyer_acting_as_list()
	
	# Ask user to select from list
	lenList = len(lList)
	while 1:
		deals_details_menu(dTF, title)

		print(f'\n    - {title} -\n')
		# Print list
		for i in range(1, lenList):
			print(' {0:>2}) {1}'.format(i, lList[i]))
		print('\n 99) Start Over')
		print(' 00) Quit')
		ui = td.uInput(f'\n Select {title} > ').upper()
		if ui == '00':
			exit('\n Terminating program...')
		elif ui == '99':
			start_over = True
			return dTF, start_over
		try:
			if ui == '' or int(ui) < 1 or int(ui) > (lenList - 1):
				td.warningMsg(' Invalid input...try again...')
				sleep(1)
				continue
		except ValueError:
			td.warningMsg(' Invalid input...try again...')
			sleep(1)
			continue
		
		# Assign selected value to field
		dTF[tf_field] = lList[(int(ui))]
		print(f'\n {title} Selected')
		td.colorText('\n {0}'.format(lList[(int(ui))]), 'ORANGE')
		if dTF[tf_field] == 'Commercial':
			dTF[tf_field] = 'Office;Retail'
		elif dTF[tf_field] == 'Leave Blank':
			dTF[tf_field] = 'None'
		break
	# Add the number of lots if the Lot Type is not Raw Acreage or Covered Land
	if tf_field == 'Lot_Type__c':
		if dTF['Lot_Type__c'] != 'Raw Acreage' and dTF['Lot_Type__c'] != 'Covered Land':
			print('\n Enter the number of Lots')
			print('\n [Enter] for unknown')
			print(' 00) Quit')
			ui_lots = td.uInput('\n Number of Lots > ')
			if ui_lots == '00':
				exit(' Terminating program...')
			elif ui_lots != '':
				dTF['Lots__c'] = int(0)
			else:
				dTF['Lots__c'] = int(ui_lots)

	return dTF, start_over

# Print the Deal Details Menu
def deals_details_menu(dTF, title):
	print_function_name(' [lao def deals_details_menu]')
	td.banner(f'Select {title}')
	# Print the values of all fields
	print(' PID:                {0}'.format(dTF['PID__c']))
	if title == 'Classification':
		td.colorText('\n Classification:     {0}'.format(dTF['Classification__c']), 'YELLOW')
	else:
		print('\n Classification:     {0}'.format(dTF['Classification__c']))

	if title == 'Lot Type':
		td.colorText(' Lot Type:           {0}'.format(dTF['Lot_Type__c']), 'YELLOW')
	else:
		print(' Lot Type:           {0}'.format(dTF['Lot_Type__c']))
	# Print Lots if not Cover Land or Raw Acreage
	if dTF['Lot_Type__c'] != 'Cover Land' and dTF['Lot_Type__c'] != 'Raw Acreage':
		if title == 'Lots':
			td.colorText(' Lots:               {0}'.format(dTF['Lots__c']), 'YELLOW')
		else:
			print(' Lots:               {0}'.format(dTF['Lots__c']))
	# Development Status only applies to Multifamily
	lList = dicts.get_multifamily_classifications_list()
	if dTF['Classification__c'] in lList:
		# Add Development Status to dTF if not already there
		if 'Development_Status__c' not in dTF:
			dTF['Development_Status__c'] = 'None'
		if title == 'Development Status':
			td.colorText(' Development Status:  {0}'.format(dTF['Development_Status__c']), 'YELLOW')
		else:
			print(' Development Status: {0}'.format(dTF['Development_Status__c']))
	
	if title == 'Buyer Acting As':
		td.colorText(' Buyer Acting As:    {0}'.format(dTF['Buyer_Acting_As__c']), 'YELLOW')
	else:
		print(' Buyer Acting As:    {0}'.format(dTF['Buyer_Acting_As__c']))

	# Print Zoning if Classification
	if title == 'Classification':
		td.colorText('\n Zoning: {0}'.format(dTF['Zoning__c']), 'ORANGE')

def VizzdaReportLookup():
	from webbrowser import open as openbrowser
	while 1:
		banner('Vizzda Report Lookup by VID')
		ui = (td.uInput(' Enter Vizzda ID or [00] Quit > ')).upper()
		if ui == 'Q' or ui == '00':
			break
		openbrowser(r'https://www2.vizzda.com/map#/event/{0}'.format(ui))

# Download and unzip Pinal County affidavits
def unzipPinal():
	from glob import glob
	from selenium import webdriver
	from selenium.webdriver.common.keys import Keys
	from selenium.common.exceptions import SessionNotCreatedException
	from datetime import date
	from datetime import timedelta

	banner('Unzipping Pinal')

	email_files = glob(r'F:\Research Department\PiRec\*.txt')
	# Close program if only one file exists (waiting for second file)

	print(email_files)
	pw_files = []
	link_files = []
	isPassword = False
	password = 'None'
	for file in email_files:
		flink = file
	
	with open(flink, 'r') as f:
		for line in f:
			print(line)
			if 'https:' in line:
				url = line
				break

	if 'google' in url:
		remove(flink)
		return 'Not the email with the link.'
	driver = getSeleniumDriver()
	driver.get(url)
	
	for i in range (6, 1, -1):
		banner('Unzipping Pinal')
		print('here2')
		print(password)
		print(' Waiting for window to fully open in...{0}'.format(i))
		sleep(1)
	if password != 'None':
		elem = driver.find_element_by_xpath('//*[@id="PasswordInput"]')
		elem.send_keys(password)
		elem.send_keys(Keys.RETURN)
	sleep(2)
	# elem = driver.find_element_by_xpath('//*[@id="SelectAll"]')
	elem = driver.find_element('xpath', '//*[@id="SelectAll"]')
	elem.click()
	# elem = driver.find_element_by_xpath('//*[@id="DownloadBtn"]/span')
	elem = driver.find_element('xpath', '//*[@id="DownloadBtn"]/span')
	elem.click()

	multiFiles = 'C:/Users/blandis/Downloads/SharedFiles.zip'

	# wait for file to download
	print('\n Waiting for file to download...')
	sleep(10)
	print('\n Accessing zip file...')
	files = 'None'
	while 1:
		if path.isfile(multiFiles) is True:
			banner('Unzipping Pinal')
			print('File downloaded...')
			files = multiFiles
			break
		for i in range(0, 14):
			day = (date.today() - timedelta(i)).isoformat()
			singleFileDay = 'C:/Users/blandis/Downloads/APV {0}.zip'.format(day)
			# print('\n Zip file name: {0}'.format(singleFileDay))
			if path.isfile(singleFileDay) is True:
				banner('Unzipping Pinal')
				print('File downloaded...')
				files = singleFileDay
				break
		if files != 'None':
			break
		else:
			sleep(2)

	# Unzip the SharedFiles.zip downloaded from Pinal website
	print('\n Uzipping SharedFiles.zip...')
	unzip(files)
	remove(files)
	# Unzip the zip files extracted from the SharedFile.zip
	files = glob('F:/Research Department/PiRec/*.zip')
	print('\n Unzipping TIFF files...')
	for zfiles in files:
		unzip(zfiles)

	driver.close()
	# remove(fpw)
	remove(flink)
	return files

def unzip(files, my_dir='F:/Research Department/PiRec'):
	from zipfile import ZipFile
	from shutil import copyfileobj
	# my_dir = r"F:\Research Department\PiRec"

	my_zip = files

	with ZipFile(my_zip) as zip_file:
		for member in zip_file.namelist():
			filename = path.basename(member)
			# skip directories
			if not filename:
				continue

			# copy file (taken from zipfile's extract)
			source = zip_file.open(member)
			# target = file(path.join(my_dir, filename), "w")
			target = open(path.join(my_dir, filename), "wb")
			with source, target:
				copyfileobj(source, target)

# Check if record is in Skip File or Write to Skip File
def SkipFile(idtext, county, action, usefile='None'):
	# print('\n ID    : {0}\n County: {1}\n Action: {2}\n'.format(idtext, county, action))
	if county == 'skipMailChimpReport':
		skipfile = 'skipMailChimpReport.txt'
	elif county == 'skipMailClickOpen':
		skipfile = 'skipMailClickOpen.txt'
	elif county == 'skipListings':
		skipfile = 'skipListings.txt'
	elif county == 'skipVizzdaAgents':
		skipfile = 'skipVizzdaAgents.txt'
	elif county == 'skipLeadToPID':
		skipfile = 'skipLeadToPID.txt'
	elif county == 'skipOrphanedPIDs':
		skipfile = 'skipOrphanedPIDs.txt'
	elif county == 'skipStandardizationCaps':
		skipfile = 'skipStandardizationCaps.txt'
	elif county == 'skipMarketMailer2023':
		skipfile = 'skipMarketMailerTaskAddedToContact2023.txt'
	elif county == 'skipAxioPipeline':
		skipfile = 'skipAxioPipeline.txt'
	elif county == 'skipTF_RQ_Reonomy':
		skipfile = 'skipTF_RQ_Reonomy_Download_Zips.txt'
	elif county == 'skip_Missing_Polygon.txt':
		skipfile = 'skip_Missing_Polygon.txt'
	else:
		# td.uInput(county)
		skipfile = getCounties('SkipFile', 'None', county)

	if usefile == 'None':
		skipfilewithpath = '{0}{1}'.format(getPath('skipfile'),skipfile)
	else:
		skipfilewithpath = usefile

	if action == 'CHECK':
		with open(skipfilewithpath, 'r') as f:
			skipit = f.read()
			if idtext in skipit:
				return True
			else:
				return False
	elif action == 'WRITE':
		with open(skipfilewithpath, 'a') as f:
			f.write('{0}\n'.format(idtext))

# Return common file paths
def getPath(path):
	user = getUserName()
	
	if path == 'axio':
		return 'F:/Research Department/MIMO/zData/AxioMetrics/'
	if path == 'comps':
		# return 'F:/Research Department/Code/Research/data/CompsFiles/'
		# return 'F:/Research Department/Code/Comps Files/'
		return 'F:/Research Department/scripts/Projects/Research/data/CompsFiles/'
	if path == 'filecabinet':
		return 'C:/Users/Public/Public Mapfiles/FileCabinet.gdb/'
	if path == 'llr':
		return 'F:/Research Department/Lot Comps Components/'
	if path == 'mapfiles':
		return 'C:/Users/Public/Public Mapfiles/'
	if path == 'm1_files':
		return 'C:/Users/Public/Public Mapfiles/M1_Files/'
	if path == 'metstud':
		return 'F:/Research Department/MIMO/zData/MetroStudy/'
	if path == 'pyData':
		return 'F:/Research Department/Code/Research/data/'
	if path == 'py3Data':
		return 'F:/Research Department/Code/RP3/data/'
	if path == 'pyEXE':
		return 'C:/Python27/ArcGIS10.8/python.exe "F:/Research Department/Code/Research/'
	if path == 'py3EXE':
		return 'C:/Program Files/Python312/python.exe "F:/Research Department/Code/RP3/'
	if path == 'rlb':
		return 'F:/Research Department/MIMO/zData/RL Brown/'
	if path == 'skipfile':
		return 'F:/Research Department/Code/Research/data/SkipFiles/'
	if path == 'temp':
		return 'C:/TEMP/'
	if path == 'TF':
		return 'https://landadvisors.my.salesforce.com/'	
	if path == 'zdata':
		return 'F:/Research Department/MIMO/zData/'
	if path == 'zonda':
		return 'F:/Research Department/MIMO/zData/Zonda/'
	if path == 'zonda_mimo':
		return 'F:/Research Department/MIMO/zData/Zonda/MIMO/'
	if path == 'pathlist':
		td.banner('Path Folders List')
		print('axio       - F:/Research Department/MIMO/zData/AxioMetrics/\n')
		print('comps      - F:/Research Department/Code/Comps Files/\n')
		print('filecabinet - C:/Users/Public/Public Mapfiles/FileCabinet.gdb')
		print('llr        - F:/Research Department/Lot Comps Components/\n')
		print('mapfiles   - C:/Users/Public/Public Mapfiles/\n')
		print('metstud    - F:/Research Department/MIMO/zData/MetroStudy/')
		print('pyEXE      - C:/Python27/ArcGIS10.8/python.exe "F:/Research Department/Code/Research/\n')
		print('py3EXE      - C:/Program Files/Python312/python.exe "F:/Research Department/Code/RP3/\n')
		print('pyData     - F:/Research Department/Code/Research/data/\n')
		print('py3Data    - F:/Research Department/Code/RP3/data/\n')
		print('rlb         - F:/Research Department/MIMO/zData/RL Brown/')
		print('skipfile   - F:/Research Department/Code/Research/data/SkipFiles/\n')
		print('temp       - C:/TEMP/\n')
		print('TF         - https://landadvisors.my.salesforce.com/\n')
		print('zdata      - F:/Research Department/MIMO/zData/\n')
		print('zonda      - F:/Research Department/MIMO/zData/Zonda/\n')
		
def costarParcelFormatter():
	from webbrowser import open as openbrowser
	banner('CoStar Parcel Formatter')
	parcels = td.uInput(' Paste parcles here > ')
	parcels = parcels.replace(", ", "', '")
	ui = td.uInput('\n Remove dashes? [0/1] > ')
	if ui == '1':
		parcels = parcels.replace('-', '')
	while 1:
		ui = td.uInput(' \n1) apn\n2) altapn\n\nSelect Field > ')
		if ui == '1':
			parcels = "apn in ('{0}')".format(parcels)
			break
		elif ui == '2':
			parcels = "altapn in ({0}')".format(parcels)
			break
		else:
			td.warningMsg('\n Invalid input try again...\n')
	with open('C:/TEMP/CoStarParcelFormatterResults.txt', 'w') as f:
		f.write(parcels)
	openbrowser('C:/TEMP/CoStarParcelFormatterResults.txt')

def propercase(name, confirm = True):

	# Check to see if name is already proper case
	for i in range(97, 122):
		if chr(i) in name:
			for i in range(65, 90):
				if chr(i) in name:
					return name
	name = name.upper()
	print(name)
	# make exception lists
	threeLetterTitleList = ['AIR', 'ANN', 'AVE', 'BAY', 'BIG', 'CIR', 'CO', 'DEV', 'DR', 'ELM', 'FWY', 'GEN', 'HOF', 'HWY',  'INC', 'JR', 'LAS', 'LIV', 'LN', 'LTD', 'LOT', 'MTN', 'NEW', 'NY', 'OAK', 'OLD', 'ONE', 'PHX', 'PL', 'REV', 'ROY', 'RD', 'RD,', 'RIM', 'RIO', 'RON', 'SAN', 'SKY', 'SON', 'SR', 'ST', 'STE', 'SUN', 'TAN', 'TEN', 'THE', 'TOM', 'TOP', 'TR', 'TWO', 'VAN', 'WAY']
	threeLetterCapsList = ['AZ', 'BOX', 'FBO', 'IH', 'IRA', 'JEN', 'IMH', 'PO', 'RE', 'USA', 'II', 'III', 'IV', 'VI', 'VII', 'IX', 'XI', 'XII', 'XIII', 'XIV', 'XVI', 'XIX', 'NE', 'SE']
	threeLetterLowerList = ['AN', 'AND', 'OF', 'OF,', 'LAW', '1ST', '2ND', '3RD', '4TH', '5TH', '6TH', '7TH', '8TH', '9TH']
	fourLetterCapsList = ['XIII', 'XVII', 'VIII', 'GCHI', 'NSHE', 'REIT']
	fourLetterLowerList = ['11TH', '12TH', '13TH']
	fourLetterTitleList = ['BLVD', 'PTNR']
	numberLowerList = ['1ST', '2ND', '3RD', '4TH', '5TH', '6TH', '7TH', '8TH', '9TH']

	print(' --------------------------------------------------------------------------------------------\n')
	print('\n Title Casing Name\n')
	print(' Upper :   {0}'.format(name))

	# Strip commas and periods
	name = name.replace(',', '').replace('.', '')

	# Key no vowel combinations uppercase
	splitName = name.split()
	fixedName = ''
	for sn in splitName:
		if len(sn) == 1:
			fixedName = '{0} {1}'.format(fixedName, sn)
			continue

		found = False
		if len(sn) <= 3:
			for n3 in threeLetterTitleList:
				if sn == n3:
					fixedName = '{0} {1}'.format(fixedName, sn.title())
					found = True
			if found:
				continue
			for n3 in threeLetterCapsList:
				if sn == n3:
					fixedName = '{0} {1}'.format(fixedName, sn)
					found = True
			if found:
				continue
			for n3 in threeLetterLowerList:
				if sn == n3:
					fixedName = '{0} {1}'.format(fixedName, sn.lower())
					found = True
			if found:
				continue

		if len(sn) == 4:
			for n4 in fourLetterCapsList:
				if sn == n4:
					fixedName = '{0} {1}'.format(fixedName, sn)
					found = True
			if found:
				continue
			for n4 in fourLetterLowerList:
				if sn == n4:
					fixedName = '{0} {1}'.format(fixedName, sn.lower())
					found = True
			if found:
				continue
			for n4 in fourLetterTitleList:
				if sn == n4:
					fixedName = '{0} {1}'.format(fixedName, sn.title())
					found = True
			if found:
				continue

		if 'A' in sn or 'E' in sn or 'I' in sn or 'O' in sn or 'U' in sn or 'Y' in sn:
			if len(sn) <= 3:
				# ui = td.uInput('\n-- > {0}\n\n1) All Caps\n2) All Lower\n3) Title case\n\nSelect > '.format(sn))
				print('\n -- > Choose format\n\n1) {0}   [Title Case]\n2) {1}   [Lower Case]\n3) {2}   [All Caps]\n\n'.format(sn.title(), sn.lower(), sn.upper()))
				ui = td.uInput(' Select > ')
				if ui == '1':
					fixedName = '{0} {1}'.format(fixedName, sn.title())
				elif ui == '2':
					fixedName = '{0} {1}'.format(fixedName, sn.lower())
				elif ui == '3':
					fixedName = '{0} {1}'.format(fixedName, sn.upper())
			else:
				fixedName = '{0} {1}'.format(fixedName, sn.title())
		else:
			for n3 in numberLowerList:
				if n3 in sn:
					fixedName = '{0} {1}'.format(fixedName, sn.lower())
					found = True
			if found:
				continue

			fixedName = '{0} {1}'.format(fixedName, sn)

	fixedName = fixedName.strip()

	print('\n Fixed :   {0}'.format(fixedName))
	print

	if confirm:
		print(' 1) Yes')
		print(' 2) No')
		print(' 3) All Caps')
		print(' 4) Quit program')
		print(' ...or type name manually...')
		ui = td.uInput('\n Select or Type > ')

		while 1:
			if ui == '1':
				return fixedName
			elif ui.upper() == 'C' or ui == '2':
				return name
			elif ui == '3':
				return name
			elif ui == '4':
				return 'QUIT'
			elif len(ui) > 2:
				return ui
			else:
				td.warningMsg('\nInvalid input...try again...')
				ui = td.uInput('\n Select or Type > ')
	else:
		return  fixedName

# Make Dictionary of Spreadsheet (CSV or Excel where Header is the dictionary keys
def spreadsheetToDict(filename, sheetname='None', capitalize_keys=False):
	# Determine if CSV or XLSX
	if '.CSV' in filename.upper():
		with open(filename, 'r') as f:
			fin = csv.reader(f)
			if 'LAO_Sale_Eblast' in filename:
				next(fin)
				next(fin)
				next(fin)
			header = next(fin)
			if capitalize_keys:
				header_upper = [x.upper() for x in header]
				header = header_upper
			colcount = len(header)
			index = 0
			dSht = {}
			for row in fin:
				index += 1
				dSht[index] = {}
				for i in range(0, colcount):
					# try:
						# print('{0} : {1}'.format([header[i]], row[i]))
					dSht[index][header[i]] = row[i]
					# except IndexError:
					# 	print(index)
					# 	print(i)
					# 	exit()
		
		return dSht
	
	elif '.XLSX' in filename.upper():
		from openpyxl import load_workbook
		wb = load_workbook(filename)
		# Set header row and first row of data if LLR or not
		if 'Land_Lot_Report' in filename:
			headerrow = 3
			firstrow = 4
			if sheetname == 'None':
				sheet = wb.worksheets[0]
			else:
				sheet = wb[sheetname]
		elif 'Market Supply & Demand' in filename:
			# headerrow = 4
			# firstrow = 5
			# xl_sheet = xl_wb.sheet_by_name('Supply and Demand')
			headerrow = 5
			firstrow = 6
			sheet = wb['Supply and Demand']
		elif 'MO Python Chart Placement' in filename:
			# headerrow = 0
			# firstrow = 1
			# xl_sheet = xl_wb.sheet_by_name('Main')
			headerrow = 1
			firstrow = 2
			sheet = wb['Main']
		elif 'Annoyatron' in filename:
			headerrow = 1
			firstrow = 2
			sheet = wb[sheetname]
		elif 'MailChimp Admin Report' in filename:
			headerrow = 3
			firstrow = 4
			sheet = wb['Group Membership']
		else:
			headerrow = 1
			firstrow = 2
			if sheetname == 'None':
				sheet = wb.worksheets[0]
			else:
				sheet = wb[sheetname]
		colcount = sheet.max_column
		rowcount = sheet.max_row
		header = ['']
		for i in range(1, colcount + 1):
			header.append(sheet.cell(headerrow, i).value)
		index = 0
		dSht = {}
		for row_idx in range(firstrow, rowcount + 1):
			index += 1
			dSht[index] = {}
			for i in range(1, colcount + 1):
				dSht[index][header[i]] = sheet.cell(row_idx, i).value
		return dSht

	elif '.XLS' in filename.upper():
		import xlrd
		xl_wb = xlrd.open_workbook(filename)
		# Set header row and first row of data if LLR or not
		if 'Land_Lot_Report' in filename:
			headerrow = 2
			firstrow = 3
			xl_sheet = xl_wb.sheet_by_index(0)
		elif 'Market Supply & Demand' in filename:
			headerrow = 4
			firstrow = 5
			xl_sheet = xl_wb.sheet_by_name('Supply and Demand')
		elif 'MO Python Chart Placement' in filename:
			headerrow = 0
			firstrow = 1
			xl_sheet = xl_wb.sheet_by_name('Main')
		elif 'Annoyatron' in filename:
			headerrow = 0
			firstrow = 1
			xl_sheet = xl_wb.sheet_by_name(sheetname)
		elif 'census_permits' in filename:
			headerrow = 7
			firstrow = 8
			xl_sheet = xl_wb.sheet_by_name('MSA Units')
		else:
			headerrow = 0
			firstrow = 1
			xl_sheet = xl_wb.sheet_by_index(0)
		
		colcount = len(xl_sheet.row(0))
		# print(f'Col Count: {colcount}')
		header = []
		for i in range(0, colcount):
			header.append(xl_sheet.cell_value(headerrow, i))
		index = 0
		dSht = {}
		rowcount = xl_sheet.nrows
		# print(f'Row Count: {rowcount}')
		for row_idx in range(firstrow, rowcount + 1):
			index += 1
			dSht[index] = {}
			for i in range(0, colcount):
				try:
					dSht[index][header[i]] = xl_sheet.cell_value(row_idx, i)
				except IndexError:
					pass
		return dSht
	else:
		input(' Not a spreadsheet file...\n{0}').format(filename)

# Save dictionary to csv file
def dict_to_csv(data, out_file=r'C:\TEMP\erase_me.csv', open_file=True):
	# Extract headers from the first inner dictionary
	headers = list(data[next(iter(data))].keys())

	# Open CSV file for writing
	with open(out_file, 'w', newline='') as f:
		fout = csv.DictWriter(f, fieldnames=['Key'] + headers)

		# Write headers
		fout.writeheader()

		# Write data rows
		for key, inner_dict in data.items():
			row = {'Key': key}
			row.update(inner_dict)
			try:
				fout.writerow(row)
			except ValueError:
				print(row)
				exit()
	
	if open_file:
		openFile(out_file)

# Convert Excel date into YYYY-MM-DD HH:MM:SS
def excelDataConverter(xldate):
	import datetime
	return (datetime.datetime(1899, 12, 30) + datetime.timedelta(days=xldate + 1462 * 0))
	
# Make a hyperlink formatted cell for CSV or Excel
def formatAsExcelHyperlink(url, id='', name=''):
	if url == 'TerraForce':
		url = 'https://landadvisors.my.salesforce.com/'
		return '=HYPERLINK(\"{0}{1}\",\"{2}\")'.format(url, id, name)
	elif url == 'Email':
		id = id.replace('elipar', 'eric')
		return '=HYPERLINK(\"mailto:{0}\",\"{0}\")'.format(id)
	elif url == 'PATH':
		return '=HYPERLINK(\"{0}\",\"{1}")'.format(id, name)
	elif url == 'MVP':
		t100url = 'https://landadvisors.my.salesforce.com/apex/Top100Account?Id='
		return '=HYPERLINK(\"{0}{1}\",\"{2}")'.format(t100url, id, name)

def selectMarket(includeUSA = False):
	lMarkets = getCounties('Market')
	count = 1
	td.banner('Select LAO Market')
	for market in lMarkets:
		print(' {0:2}) {1}'.format(count, market))
		count += 1

	# Include USA for GV National Report
	if includeUSA:
		print(' {0:2}) {1}'.format(count, 'USA'))
	print(' 00) Quit')
	ui = td.uInput('\n Select Market > ')
	# Quit program
	if ui == '00':
		exit(' Terminating program...')
	# GV Whale Sales report which covers multiple USA markets
	elif int(ui) == count:
		return 'USA'
	ui = int(ui) - 1
	market = lMarkets[ui]
	return market

# Select a Market for 'WHERE' in query based on county or state depending on the market
def getCountiesInMarketWhereClause():

	market = selectMarket()
	wc = ''
	if market == 'Albuquerque':  # Albuquerque
		wc = "State__c = 'New Mexico'"
	elif market == 'Atlanta':  # Atlanta
		wc = "State__c = 'Georgia'"
	elif market == 'Boise':  # Boise
		wc = "State__c = 'Idaho'"
	elif market == 'Charlotte':  # Charlotte
		wc = "(State__c = 'North Carolina' or State__c = 'South Carolina')"
	elif market == 'Kansas City':  # Charlotte
		wc = "(State__c = 'Kansas' or State__c = 'Missouri')"
	elif market == 'Nashville':  # Charlotte
		wc = "(State__c = 'Tennessee' or State__c = 'Alabama' or State__c = 'Kentucky')"
	elif market == 'Las Vegas':  # Las Vegas
		wc = "State__c = 'Nevada'"
	elif market == 'Salt Lake City':  # Salt Lake City
		wc = "State__c = 'Utah'"

	# Build county wc for Arizona, Florida, Nevada, Texas, Kansas, Missouri
	if wc == '':
		lMarkets = getCounties('Counties', market)
		wc = "("
		for mrkt in lMarkets:
			wc = "{0}County__c = '{1}' or ".format(wc, mrkt)
		wc = wc[::-1].replace(' or '[::-1], ')'[::-1], 1)[::-1]
	
	# Counties in multiple markets define add State to query
	if market == 'DFW' or market == 'Austin' or market == 'Houston':
		wc = "State__c = 'Texas' and {0}".format(wc)
	elif market == 'Jacksonville' or market == 'Orlando' or market == 'Tampa':
		wc = "State__c = 'Florida' and {0}".format(wc)

	return wc, market

# Open a file in its default program
def openFile(filename):
	from subprocess import Popen
	# Trap &
	if '&' in filename:
		td.warningMsg('\n File name contains "&" which will not open in this program...')
		ui = td.uInput('\n Continue [00]... > ')
		if ui == '00':
			exit('\n Terminating program...')

	while 1:
		# Open file in default program
		try:
			Popen(filename, shell=True)
			print(f'\n Opening file: {filename}\n')
			break
		# Capture if file is open
		except PermissionError:
			td.warningMsg('\n The file is open...close it and press [Enter] or [00]')
			ui = td.uInput('\n  > ')
			if ui == '00':
				exit('\n Terminating program...')

# Rename Pinal TIFF files
def renamePinalTIFFs():
	import glob
	import os

	for file in glob.glob(r'F:\Research Department\PiRec\*.tif'):
		if len(file) > 20:
			folder = os.path.dirname(file)
			newFile = os.path.basename(file)
			newFile = '{0}\\{1}'.format(folder, newFile[-15:])
			try:
				os.rename(file, newFile)
			except WindowsError:
				continue

def getPCName():
	import socket
	return socket.gethostname()

def getUser_tfUserID():
	from getpass import getuser
	user = getuser().lower()
	dAgents = getAgentDict(dict_type='full', version='v2')
	for agent in dAgents:
		if user == dAgents[agent]['Email']:
			tfUserID = dAgents[agent]['UserID']
			return tfUserID

def getUser_tfID():
	from getpass import getuser
	user = getuser().lower()
	# dAgents = getAgentDict(dict_type='full', version='v2')
	dStaff = dicts.get_staff_dict()
	for staff in dStaff:
		if user == dStaff[staff]['Email']:
			tfID = dStaff[staff]['Id']
			return tfID

# Get the user name of the opperator
def getUserName(initials=False):
	from getpass import getuser
	user = getuser().lower()
	if initials:
		initials = user[:2].upper()
		return user, initials
	else:
		return user

# Get LAO County information
def getCounties(returnList, market='None', ArcName='None', MarketAbb='None'):
	# print_function_name('lao def getCounties')
	# print(' def getcouties\n {0}\n {1}\n {2}\n {3}\n'.format(returnList, market, ArcName, MarketAbb))
	dCounties = dicts.spreadsheet_to_dict('F:/Research Department/Code/Databases/LAO_Counties.xlsx')
	d = {}
	l = []
	if returnList == 'FullDict':
		return dCounties
	for row in dCounties:
		# Return list of LAO Markets
		if returnList == 'Market' and not dCounties[row]['Market'] in l:
			l.append(dCounties[row]['Market'])
		# Return list of LAO MSAs
		if returnList == 'MSA' and not dCounties[row]['Market'] in l:
			l.append(dCounties[row]['MSA'])
		# Return list of Market's Parcel ArcName in a Market
		elif returnList == 'Counties' and dCounties[row]['Market'] == market and not dCounties[row]['ArcName'] in l:
			l.append(dCounties[row]['ArcName'])
		# Return list of Market's Parcel ArcName in a Market for 52
		elif returnList == 'Counties_52' and dCounties[row]['Market'] == market:
			listname = '{0}, {1}'.format(dCounties[row]['ArcName'], dCounties[row]['State'])
			l.append(listname)
		# Return Maret of ParcelsName
		elif returnList == 'MarketOfParcelsName' and market.upper() == dCounties[row]['ParcelsName'].upper():
			return dCounties[row]['Market']
		# Return County's State
		elif returnList == 'State' and dCounties[row]['ArcName'].upper() == ArcName.upper():
			return dCounties[row]['State']
		# Return Market's State
		elif returnList == 'MarketsState' and dCounties[row]['Market'] == market:
			return dCounties[row]['State']
		# Return list of two letter States
		elif returnList == 'StateAbb' and not dCounties[row]['State'] in l:
			l.append(dCounties[row]['State'])
		# Return list of Counties without an AltAPN
		elif returnList == 'NoAltAPN' and dCounties[row]['AltAPN'] == 'No':
			l.append(dCounties[row]['ParcelsName'])
		# Return list of Market's Parcel Names
		elif returnList == 'MarketParcelNames' and dCounties[row]['Market'] == market:
			l.append(dCounties[row]['ParcelsName'])
		elif returnList == 'MarketParcelLeadNames' and dCounties[row]['Market'] == market:
			l.append(dCounties[row]['ParcelsName'])
			l.append(dCounties[row]['ParcelsName'].replace('Parcels', 'Leads'))
		# Return list of all Parcel Names
		elif returnList == 'AllParcelsNames':
			l.append(dCounties[row]['ParcelsName'])
		# Return Name of Skipfile based on Parcel ArcName
		elif returnList == 'SkipFile' and dCounties[row]['ArcName'].upper() == ArcName.upper():
			return dCounties[row]['skipFile']
		# Return if County is Non-Disclosure based on ArcName
		elif returnList == 'IsDisclosure' and dCounties[row]['StateFull'] == market and dCounties[row]['ArcName'] == ArcName:
			return dCounties[row]['Disclosure']
		elif returnList == 'CountysMarket':
			county = dCounties[row]['ArcName']
			d[county] = dCounties[row]['Market']
		# Return Market name based on 3 Letter Market Abbriviation
		elif returnList == 'MarketAbb' and dCounties[row]['MarketAbb'] == MarketAbb:
			return dCounties[row]['Market']
		# Return list of all counties
		elif returnList == 'AllCounties ArcName':
			l.append(dCounties[row]['ArcName'])
		# Return list of all counties
		elif returnList == 'AllCounties County Name':
			l.append(dCounties[row]['County'])
		elif returnList == 'ArcParcelLayers':
			d[dCounties[row]['ParcelsName']] = 'owner'
		elif returnList == 'CountyInfo' and ArcName == dCounties[row]['ArcName']:
			return dCounties[row]
		# Open LAO_Counties.csv in Excel
		elif returnList == 'Modify':
			openFile('F:/Research Department/Code/Databases/LAO_Counties.xlsx')
			break

	if l != []:
		l.sort()
		return l
	if d != {}:
		return d

# Open file for writing and loop if it is already open and ask to close
def check_if_file_is_open(filename):
	while 1:
		try:
			with open(filename, 'w') as f:
				break
		except IOError:
			filename_wo_path = path.basename(filename)
			td.warningMsg('\n File {0} is open...close and continue...'.format(filename_wo_path))
			ui = td.uInput('\n Continue or [00] to quit... > ')
			if ui == '00':
				exit('\n Terminating program...')

def doh():
	# from colored import fg, attr
	print("{0}                              __".format(fg('yellow_2')))
	print("                    _ ,___,-'',-=-.")
	print("        __,-- _ _,-'_)_  (""`'-._  `.")
	print("     _,'  __ |,' ,-' __)  ,-     /. |")
	print("   ,'_,--'   |     -'  _)/         `\\ ")
	print(" ,','      ,'       ,-'_,`           :")
	print(" ,'     ,-'       ,(,-(              :")
	print("      ,'       ,-' ,    _            ;")
	print("     /        ,-._/`---'            /")
	print("    /        (____)(----. )       ,'")
	print("   /         (      `.__,     /\\ /,")
	print("  :           ;-.___         /__\\/|")
	print("  |         ,'      `--.      -,\\ |")
	print("  :        /            \\    .__/")
	print("   \\      (__            \\    |_")
	print("    \\       ,`-, *       /   _|,\\ ")
	print("     \\    ,'   `-      ,'_,-'    \\ ")
	print("    (_\\,-'    ,'\\\")--,'-'       __\\ ")
	print("     \\       /  /  ,'|      ,--'  `-.")
	print("      `-.    `-/ \\\'  |   _,'         `.")
	print("         `-._ /      `--'/             \\ ")
	print("            ,'           |              \\ ")
	print("           /             |               \\ ")
	print("        ,-'              |               /")
	print("       /                 |             -'{0}".format(attr(0)))

	td.uInput('\n Continue... > ')

# Return list of Fields for CSV to TF spreadsheet
def headerTFCSV():
	td.warningMsg('\n lao.headerTFCSV() has been replace by dicts.get_tf_csv_header()')
	print('\n Terminating program...')

# Get the market abbreviation and state abbreviation
def get_market_abbreviation(market, county='None'):
		
	# Temp fix
	if market == 'Georgia':
		return 'Altanta', 'GA'
	elif market == 'Knoxville':
		return 'Nashville', 'TN'


	dCounties = getCounties('FullDict')
	marketAbb = 'None'
	county = county.replace(' ', '').replace('.', '')

	for row in dCounties:
		if county != 'None':
			if market.upper() == dCounties[row]['Market'].upper():
				# print(market)
				# print(county)
				# print(dCounties[row]['County'])
				# print
				if county.upper() == dCounties[row]['County'].upper() or county.upper() == dCounties[row]['ArcName'].upper():
					marketAbb = dCounties[row]['MarketAbb']
					stateAbb = dCounties[row]['State']
					break
		else:
			if market.upper() == dCounties[row]['Market'].upper():
				marketAbb = dCounties[row]['MarketAbb']
				stateAbb = dCounties[row]['State']
				break
	# Error check
	if marketAbb == 'None':
		td.warningMsg('\n Could not determine marketAbb...')
		print('\n Market: {0}'.format(market))
		print(' County: {0}'.format(county))
		td.warningMsg('\n Terminating program...')
		lao.holdup()
		exit()

	return marketAbb, stateAbb

def get_state_of_market(market, county='None'):
	dMarkets = getCounties('FullDict')
	for mkt in dMarkets:
		if market == dMarkets[mkt]['Market'] and county == 'None':
			return dMarkets[mkt]['StateFull']
		elif market == dMarkets[mkt]['Market'] and county == dMarkets[mkt]['County']:
			return dMarkets[mkt]['StateFull']
	td.warningMsg('\n Could not identify State by Market {0}\n\nTerminating program...'.format(market))
	exit()

# Set ArcMap pop-out console window size & color to green
def consoleArcMapSizeColor(color='default'):
	# Format 'color [background][text]'
	if color == 'MailChimp':
		system('color e0')
		rect = wintypes.SMALL_RECT(0, 0, 70, 40)  # (left, top, right, bottom)
	elif color == 'ArcMakeLeadStart':
		system('color 2f')  # green on white
		rect = wintypes.SMALL_RECT(0, 0, 50, 10)  # (left, top, right, bottom)
	elif color == 'ArcMakeLeadFinish':
		system('color 0')  # black on white
		rect = wintypes.SMALL_RECT(0, 0, 50, 10)  # (left, top, right, bottom)
	else:
		system('color 2f')  # green on white
		rect = wintypes.SMALL_RECT(0, 0, 90, 65)  # (left, top, right, bottom)
	STDOUT = -11
	hdl = windll.kernel32.GetStdHandle(STDOUT)
	windll.kernel32.SetConsoleWindowInfo(hdl, True, byref(rect))

# Sets position of the console window
def consoleWindowPosition(position='default'):
	user = getUserName()
	import win32gui
	import win32api

	# Get screen resolution using Windows API
	screenwidth = win32api.GetSystemMetrics(0)  # SM_CXSCREEN
	screenheight = win32api.GetSystemMetrics(1)  # SM_CYSCREEN
	

	# MoveWindow(hwnd, x, y, width, height, bRepaint)
	hwnd = win32gui.GetForegroundWindow()
	strInterns = 'tkarber, kwhite, oabdelmawgoud, lcofeen, cthompson, aroe, lsweetser, ejusto'
	if position == 'default':
		if user in strInterns:
			# Intern 2 monitor
			win32gui.MoveWindow(hwnd, 50, 50, 830, 750, True)
		# else:
		# 	# Research big monitor
		# 	win32gui.MoveWindow(hwnd, 250, 250, 830, 1000, True)
	elif position == 'MailChimp5':
		win32gui.MoveWindow(hwnd, 2242, 0, 830, 1000, True)

	elif position == 'MailChimp3':
		if screenwidth == 1920 and screenheight == 1080: # Bill Home
			print('home')
			win32gui.MoveWindow(hwnd, 3017, 0, 830, 1000, True)
		else:
			print('office')
			win32gui.MoveWindow(hwnd, 2242, 1010, 830, 675, True)

	elif position == 'Bill Marvelous Menu':
		if screenwidth == 1920 and screenheight == 1080: # Bill Home
			win32gui.MoveWindow(hwnd, 1920, 0, 830, 1000, True)
		else:
			win32gui.MoveWindow(hwnd, 2242, 0, 830, 1000, True)


# Sets color of the console
def consoleColor(color='BLACK'):
	# import Color_Console as c_c
	if color.upper() == 'BLACK':
		system('color 07') # black with white letters
	if color.upper() == 'BLUE':
		system('color 1f') # blue with white letters
	elif color.upper() == 'GREEN':
		system('color 2f')  # green with white letters
	elif color.upper() == 'PURPLE':
		system('color 5f') # purple with white letters
	elif color.upper() == 'RED':
		system('color 4f') # red with white letters
	elif color.upper() == 'YELLOW':
		system('color e0') # yellow with black letters

def playSound(sound):
	import winsound
	if sound == 'ERROR':
		duration = 1000 # milliseconds
		freq = 440 # Hz
		winsound.Beep(freq, duration)
		winsound.Beep(freq, duration)
		winsound.Beep(freq, duration)

# Return CSV as a List of Lists where each line is a list withing a master list
def getCSVListofList(csvFile):
	import csv
	lList = []
	with open(csvFile, 'r') as f:
		fin = csv.reader(f)
		for row in fin:
			# print(row)
			lList.append(row)
	return lList

# Convert UTF-8 special characters to ASCII and set case
def charactersToASCII(txt, charCase='None'):

	txt = txt.replace('±', '+/-')
	txt = txt.replace('±', '+/-') # MailChimp +/- character
	txt = txt.replace('—', '-')
	txt = txt.replace('½', '1/2')
	txt = txt.replace('¼', '1/4')
	txt = txt.replace('¾', '3/4')
	txt = txt.replace("’", "'")
	txt = txt.replace('Ń', 'N')
	txt = txt.replace('ń', 'n')
	txt = txt.replace('Ň', 'N')
	txt = txt.replace('ň', 'n')
	txt = txt.replace('©', '(c)')
	txt = txt.replace('®', '(R)')
	txt = txt.replace("’s", "'s")
	if charCase.upper() == 'UPPER':
		txt = txt.upper()
	elif charCase.upper() == 'TITLE':
		txt = txt.title()
	elif charCase.upper() == 'LOWER':
		txt = txt.lower()
	return txt

def getTruePeopleSearchPhoneNumbers():
	import pyperclip
	while 1:
		print('\n\n  1) Get phone numbers from clipboard')
		print('  2) Skip')
		print('  Type Phone Number')
		print(' 00) Quit')
		ui = td.uInput('\n > ')
		if ui == '1':
			getphonenumbers = True
			break
		elif ui == '2':
			getphonenumbers = False
			break
		elif len(ui) > 2:
			landline = phoneFormat(ui)
			return landline, 'None', 'None'
		elif ui == '00':
			exit('\n Terminating program...')
		else:
			td.warningMsg('\n Invalid input...try again...')
	if getphonenumbers:
		while 1:
			ui = pyperclip.paste()
			# td.uInput(ui)
			if 'Landline' not in ui and 'Wireless' not in ui:
				td.warningMsg('\n No phone numbers found in clipboard!\n Copy the phone numbers and try again...')
				td.uInput('\n Continue > ')
			else:
				break
		phonenumbers = ui.split('\r\n')
		# td.uInput(phonenumbers)
		# print(phonenumbers)
		landline, mobile, phoneDescriptionFieldString = 'None', 'None', 'None'
		for phone in phonenumbers:
			if 'Landline' in phone and landline == 'None':
				landline = phone[:14]
			elif 'Wireless' in phone and mobile == 'None':
				mobile = phone[:14]
			if phoneDescriptionFieldString == 'None' and ('Landline' in phone or 'Wireless' in phone):
				phoneDescriptionFieldString = 'Possible Phone Numbers:\n {0}'.format(phone)
			elif 'Landline' in phone or 'Wireless' in phone:
				phoneDescriptionFieldString = '{0}\n {1}'.format(phoneDescriptionFieldString, phone)
		# print(landline)
		# print(mobile)
		# print(phoneDescriptionFieldString)
		return landline, mobile, phoneDescriptionFieldString
	else:
		return 'None', 'None', 'None'

def getDictValue(dict, inKey, inValue, outKey='All'):
	for row in dict:
		if dict[row][inKey] == inValue:
			if outKey == 'All':
				return dict[row]
			else:
				return dict[row][outKey]

# Run script in Python 2
def run_python_3(py_script):
	print('\n Launching {0} in Python 3...'.format(py_script))
	pyPath = 'C:/"Program Files/ArcGIS/Pro/bin/Python/envs/arcgispro-py3/python.exe" "F:/Research Department/Code/Research/{0}"'.format(py_script)
	system(pyPath)

# Copy file and rename it appending the current date and time
def make_file_backup(file_path):
	import os
	import shutil
	from datetime import datetime
	directory = os.path.dirname(file_path)
	filename, extension = os.path.splitext(os.path.basename(file_path))

	now = datetime.now()
	formatted_date = now.strftime('%Y-%m-%d-%H-%M-%S')
	new_name = "{0}_{1}{2}".format(filename, formatted_date, extension)
	new_file_path = os.path.join(directory, new_name)


	shutil.copy2(file_path, new_file_path)
	print("File '{0}' copied and renamed as '{1}'.".format(file_path, new_name))

def print_py_3_string(strText):
	print(f'\n This is the string: {strText}')
	exit()

# Pause the script with the option to quit
def holdup():
	from colored import Fore, Back, Style
	ui = input(f'{Fore.cyan_1}\n Continue...[00]... > {Style.reset}')
	if ui == '00':
		exit('\n Terminating program...')
	else:
		return

# Print the function name for Bill
def print_function_name(fun_name, skipit=False):
	import fjson
	function_status = fjson.read_bill_script_msgs_on_off('Functions')
	if function_status == 'Off':
		return
	if skipit:
		return
	filename = 'F:/Research Department/Code/Show Script Functions.txt'
	with open(filename, 'r') as f:
		contents = f.read()
	if contents == 'True' and getUserName() == 'blandis':
			td.colorText(' [{0}]'.format(fun_name), 'YELLOW', colorama=True)

# Check if all MetroStudy reports have been generated
def metrostudy_reports_exist(year_qtr):
	import os.path
	MS_folder = 'F:/Research Department/MIMO/zData/Metrostudy/'
	lMarkets = ['ATL',
				'AUS',
				'BOI',
				'CLT',
				'DEN',
				'DFW',
				'HOU',
				'IEP',
				'JAX',
				'LVS',
				'NSH',
				'ORL',
				'PHX',
				'PRC',
				'RNO',
				'SAC',
				'SEA',
				'SLC',
				'SRQ',
				'TPA',
				'TUC']
	dReports = {'subs': 'xls',
				'HistoricalHousingActivity': 'xls',
				'Qtrly_Activity_Builder': 'pdf',
				'Qtrly_Plan_Sum_Builder': 'pdf'}
	
	missing_reports = False
	for mkt in lMarkets:
		for rpt in dReports:
			file_extention = dReports[rpt]
			report_name = '{0}{1}_{2}_{3}.{4}'.format(MS_folder, mkt, rpt, year_qtr, file_extention)
			if os.path.isfile(report_name) is False:
				report_name = report_name.replace(MS_folder, '')
				print(f'\n Missing Report: {report_name}')
				missing_reports = True
	
	if missing_reports:
		ui = td.uInput('\n Continue...[00] > ')
		if ui == '00':
			exit('\n Terminating program...')

# Get LAO Market from County & State
def get_market_from_county_state(county, state):
	# print(f'{county} : {state}')
	dCounties = getCounties('FullDict')
	for row in dCounties:
		if dCounties[row]['ArcName'] == county and dCounties[row]['StateFull'] == state:
			return dCounties[row]['Market']

# Check if file exists
def does_file_exists(file_path):
	import os.path
	if os.path.isfile(file_path):
		return True
	else:
		return False

def delete_file(file_path):
	import os
	os.remove(file_path)

# Archives the Comps and Listings files
def clean_comps_listings_files_folders():
	import os
	import shutil
	import time

	print('\n Cleaning comps & listins files folders...')

	# Define the age threshold in seconds (5 weeks)
	five_weeks_in_seconds = 5 * 7 * 24 * 60 * 60  # 5 weeks

	# Get the current time
	current_time = time.time()

	# Iterate through the Comps files in the source directory
	print('\n Archiving Comps Files...')
	# Define source and destination directories
	source_dir = r"F:\Research Department\scripts\Projects\Research\data\CompsFiles"
	archive_dir = r"F:\Research Department\scripts\Projects\Research\data\CompsFiles\Archive"
	for filename in os.listdir(source_dir):
		file_path = os.path.join(source_dir, filename)

		# Check if it's a file and has the correct extension
		if os.path.isfile(file_path) and (filename.endswith('.zip') or filename.endswith('.csv')):
			# Get the file's last modification time
			file_age = current_time - os.path.getmtime(file_path)

			# Check if the file is older than 5 weeks
			if file_age > five_weeks_in_seconds:
				# Move the file to the archive directory
				shutil.move(file_path, os.path.join(archive_dir, filename))
				print(f"Moved: {filename} to {archive_dir}")
			else:
				print(f"File {filename} is not older than 5 weeks.")

	print('\n Archiving Comps Files...')
	# Define source and destination directories
	source_dir = r"F:\Research Department\Listings"
	archive_dir = r"F:\Research Department\Listings\Archive"

	# Iterate through the Listings files in the source directory
	for filename in os.listdir(source_dir):
		file_path = os.path.join(source_dir, filename)

		# Check if it's a file and has the correct extension
		if os.path.isfile(file_path) and (filename.endswith('.xls') or (filename.endswith('.xlsx'))):
			# Get the file's last modification time
			file_age = current_time - os.path.getmtime(file_path)

			# Check if the file is older than 5 weeks
			if file_age > five_weeks_in_seconds:
				# Move the file to the archive directory
				shutil.move(file_path, os.path.join(archive_dir, filename))
				print(f"Moved: {filename} to {archive_dir}")
			else:
				print(f"File {filename} is not older than 5 weeks.")

# OLD FUNCTIONS REPLACED BY NEW FUNCTIONS ##############################################

def less24months():
	td.warningMsg('lao.less24months is replaced with td.less_time_ago')
	exit('\n Terminating program...')
	# from datetime import date
	# from datetime import timedelta
	# return (date.today() - timedelta(2*365.25)).isoformat()

def less84months():
	td.warningMsg('lao.less84months is replaced with td.less_time_ago')
	exit('\n Terminating program...')
	# from datetime import date
	# from datetime import timedelta
	# return (date.today() - timedelta(7*365.25)).isoformat()	

def todayLessNdays(days):
	td.warningMsg('lao.todayLessNdays is replaced with td.less_time_ago')
	exit('\n Terminating program...')
	# from datetime import date
	# from datetime import timedelta
	# return (date.today() - timedelta(days)).isoformat()

# Print yellow warning message
def warningMsg(text):
	td.warningMsg('lao.warningMsg is replaced with td.warningMsg')
	print(f'\n Text sent: {text}')
	exit('\n Terminating program...')
	# from colored import Fore, Back, Style
	# print(f'{Fore.yellow_2}{text}{Style.reset}')

def instrMsg(text):
	td.warningMsg('lao.instrMsg is replaced with td.instrMsg')
	exit('\n Terminating program...')
	# from colored import Fore, Back, Style
	# print(f'{Fore.green_1}{text}{Style.reset}')

def colorText(text, color):
	td.warningMsg('lao.colorText is replaced with td.colorText')
	print(f'\n Text: {text}')
	print(f' Color: {color}')
	exit('\n Terminating program...')
	# from colored import Fore, Back, Style
	# print(f'{Fore.yellow_2}{'lao.colorText is replaced with td.colorText'}{Style.reset}')
	
	# color = color.upper()
	# if color == 'BLUE':
	# 	print(f'{Fore.dodger_blue_2}{text}{Style.reset}')
	# if color == 'BROWN':
	# 	print(f'{Fore.sandy_brown}{text}{Style.reset}')
	# if color == 'CYAN':
	# 	print(f'{Fore.cyan_1}{text}{Style.reset}')
	# if color == 'GREY' or color == 'GRAY':
	# 	print(f'{Fore.grey_54}{text}{Style.reset}')
	# if color == 'GREEN':
	# 	print(f'{Fore.green_1}{text}{Style.reset}')
	# if color == 'ORANGE':
	# 	print(f'{Fore.orange_1}{text}{Style.reset}')
	# if color == 'PINK':
	# 	print(f'{Fore.hot_pink_1b}{text}{Style.reset}')
	# if color == 'PURPLE':
	# 	print(f'{Fore.purple_1b}{text}{Style.reset}')
	# if color == 'RED':
	# 	print(f'{Fore.indian_red_1c}{text}{Style.reset}')
	# if color == 'YELLOW':
	# 	print(f'{Fore.yellow_2}{text}{Style.reset}')

# Sets the title of the console window to use for finding it with os
def consoleTitle(title):
	td.warningMsg('\n lao.consoleTitle is replaced by td.console_title')
	exit('\n Terminating program...')
	# import ctypes
	# ctypes.windll.kernel32.SetConsoleTitleA(title)

def setActiveWindow(window_title):
	td.warningMsg('\n lao.setActiveWindow() has been replaced by td.set_active_window()')
	exit('\n Terminating program...')

# Select LAO Advisor name from list
def select_LAO_advisor_name():
	dStaff = dicts.get_staff_dict()
	lAdvisors = []
	for name in dStaff:
		if dStaff[name]['LAO'] == 'Yes' and dStaff[name]['Roll'] == 'Agent':
			lAdvisors.append(name)

	lAdvisors.sort()

	advisors = 'None'
	while 1:
		# Menu
		print('\n Select an Advisor')
		for i in range(0, len(lAdvisors)):
			print(f' {i+1:2}) {lAdvisors[i]}')
		print('\n 99) Submit / Finished selecting')
		print(' 00) Quit')
		td.colorText(f'\n Selected: {advisors}', "GREEN")

		
		ui = td.uInput('\n Select > ')
		if ui == '00':
			exit('\n Fin...')
		elif ui == '99' or ui == '':
			return advisors
		else:
			if advisors == 'None':
				advisors = lAdvisors[int(ui)-1]
			else:
				advisors = f'{advisors};{lAdvisors[int(ui)-1]}'

