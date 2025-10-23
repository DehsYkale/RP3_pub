#!/usr/bin/env python
# -*- coding: utf-8 -*-

# Make CSV of Market Insight Excel file's table data/stats for InDesign templates

import csv
import datetime
import fun_text_date as td
import lao
import xlrd
import openpyxl
import pandas as pd
import shutil
import os
from pprint import pprint

def getHeader():
	header = ['Cover_Title_1',
			  'Cover_Val_1',
			  'Cover_Title_2',
			  'Cover_Val_2',
			  'Cover_Title_3',
			  'Cover_Val_3',
			  'Cover_Title_4',
			  'Cover_Val_4',
			  'Housing_Title_1',
			  'Housing_Date_Last_1',
			  'Housing_Date_Cur_1',
			  'Housing_Val_Last_1',
			  'Housing_Val_Cur_1',
			  'Housing_Chg_Val_1',
			  'Housing_Title_2',
			  'Housing_Date_Last_2',
			  'Housing_Date_Cur_2',
			  'Housing_Val_Last_2',
			  'Housing_Val_Cur_2',
			  'Housing_Chg_Val_2',
			  'Housing_Title_2',
			  'Housing_Date_Last_3',
			  'Housing_Date_Cur_3',
			  'Housing_Val_Last_3',
			  'Housing_Val_Cur_3',
			  'Housing_Chg_Val_3',
			  'Emp_Rate_Title',
			  'Emp_Rate_Date_Last',
			  'Emp_Rate_Date_Cur',
			  'Emp_Rate_Mkt',
			  'Emp_Rate_Mkt_Val_Last',
			  'Emp_Rate_Mkt_Val_Cur',
			  'Emp_Rate_Mkt_Val_Chg',
			  'Emp_Rate_State',
			  'Emp_Rate_State_Val_Last',
			  'Emp_Rate_State_Val_Cur',
			  'Emp_Rate_State_Val_Chg',
			  'Emp_Rate_State2',
			  'Emp_Rate_State2_Val_Last',
			  'Emp_Rate_State2_Val_Cur',
			  'Emp_Rate_State2_Val_Chg',
			  'Emp_Total_Title',
			  'Emp_Total_Date_Last',
			  'Emp_Total_Date_Cur',
			  'Emp_Total_Mkt',
			  'Emp_Total_Mkt_Val_Last',
			  'Emp_Total_Mkt_Val_Cur',
			  'Emp_Total_Mkt_Val_Chg',
			  'Emp_Total_State',
			  'Emp_Total_State_Val_Last',
			  'Emp_Total_State_Val_Cur',
			  'Emp_Total_State_Val_Chg',
			  'Emp_Total_State2',
			  'Emp_Total_State2_Val_Last',
			  'Emp_Total_State2_Val_Cur',
			  'Emp_Total_State2_Val_Chg',
			  'Emp_Grwth_Title',
			  'Emp_Grwth_Mkt',
			  'Emp_Grwth_Mkt_Chg',
			  'Emp_Grwth_State_Mkt',
			  'Emp_Grwth_State_Chg',
			  'Emp_Grwth_State2_Mkt',
			  'Emp_Grwth_State2_Chg',
			  'MLS_Title_1',
			  'MLS_Date_Last_1',
			  'MLS_Date_Cur_1',
			  'MLS_Val_Last_1',
			  'MLS_Val_Cur_1',
			  'MLS_Val_Chg_1',
			  'MLS_Title_2',
			  'MLS_Date_Last_2',
			  'MLS_Date_Cur_2',
			  'MLS_Val_Last_2',
			  'MLS_Val_Cur_2',
			  'MLS_Val_Chg_2',
			  'MLS_Title_3',
			  'MLS_Date_Last_3',
			  'MLS_Date_Cur_3',
			  'MLS_Val_Last_3',
			  'MLS_Val_Cur_3',
			  'MLS_Val_Chg_3',
			  'MLS_Title_4',
			  'MLS_Date_Last_4',
			  'MLS_Date_Cur_4',
			  'MLS_Val_Last_4',
			  'MLS_Val_Cur_4',
			  'MLS_Val_Chg_4',
			  'Emp_Cat_Title_1',
			  'Emp_Cat_Val_1',
			  'Emp_Cat_Title_2',
			  'Emp_Cat_Val_2',
			  'Emp_Cat_Title_3',
			  'Emp_Cat_Val_3',
			  'Emp_Cat_Title_4',
			  'Emp_Cat_Val_4',
			  'Chart_Title_1',
			  'Chart_Title_2',
			  'Chart_Title_3',
			  'Chart_Title_4',
			  'Chart_Title_5',
			  'Chart_Title_6',
			  'Chart_Title_7',
			  'Chart_Title_8',
			  'Chart_Title_9',
			  'Chart_Title_10',
			  'Year',
			  'Quarter'
			  ]
	return header

def formatCell(rawCell, datatype, titleCell):
	# print('FormatCell')
	# print('  datatype: {0}'.format(datatype))
	# print('  rawCell.value: {0}'.format(rawCell.value))
	# print('  titleCell: {0}'.format(titleCell))
	# Format Year Only
	if datatype == 'Title':
		if type(rawCell.value) is float:
			if rawCell.value < 2030:  # Year only
				return str(int(rawCell.value))
			else:  # Month-Year
				yr, mo, day, hr, min, sec = xlrd.xldate_as_tuple(rawCell.value, 0)
				x = datetime.datetime(yr, mo, day, hr, min, sec)
				return x.strftime('%b %Y')
		else:
			return rawCell.value
	if datatype == 'Value':
		tcv = titleCell.value
		rcv = rawCell.value
		if 'Price' in tcv and rawCell.value > 1:
			return lao.curFormat(rcv)
		if 'Days' in tcv:
			return '{0} Days'.format(int(rcv))
		if 'Months of Inventory' in tcv:
			return '{0:.1f} Mo'.format(float(rcv))
		if 'Listings' in tcv and rcv > 1:
			return '{:,}'.format(int(rcv))
		if 'Annualized Closed Sales' in tcv and rcv > 1:
			return '{:,}'.format(int(rcv))
		# if 'Months' in tcv:
		# 	return '{0:.1f}'.format(float(rcv))
		if ('Starts' in tcv or 'Closings' in tcv or 'Permits' in tcv or 'Closed' in tcv or 'New Listings' in tcv) and rcv > 1:
			numFormatted = '{:,}'.format(int(rcv))
			return numFormatted
		if 'Volume' in tcv and rcv > 1:
			numraw =  '{:,}'.format(int(rcv))
			lnumraw = numraw.split(',')
			# td.uInput(len(lnumraw))
			if len(lnumraw) == 3:
				numFormatted = '${0}M'.format(lnumraw[0])
			if len(lnumraw) == 4:
				numFormatted = '${0}.{1}B'.format(lnumraw[0], lnumraw[1][:1])
			return numFormatted
		if type(rcv) is float:
			if rcv < 1:
				return '{0:.1%}'.format(rcv)
			if rawCell.value > 9999:
				numFormatted = lao.thousandsFormat(int(rcv))
				return numFormatted
			else:
				return '{:,}'.format(int(rcv))
		return rcv  # return string value

# Get the Cover Data for a Market
def getCoverData():
	# print('\n getCoverData')
	lRow = []

	rowStart = 6
	rowEnd = 12
	# colTitle = 0
	colTitle = 1
	colValue = 2
	for i in range(rowStart, rowEnd, 2):
		# Title
		cell = sheet.cell(i, colTitle)
		title = '{0}'.format(cell.value)
		lRow.append(title)
		# Value
		cell = sheet.cell(i, colValue)
		print('here1')
		print(cell.value)
		if cell.value < 1:
			wcell = '{0:.1%}'.format(cell.value)
		elif 'PRICE' in title.upper():
			wcell = lao.curFormat(cell.value)
		elif cell.value >= 1000:
			wcell = '{:,}'.format(int(cell.value))
		else:
			wcell = '{0}'.format(cell.value)
		lRow.append(wcell)
	return lRow

def getChartTitles():
	lRow = []

	rowStart = 4
	rowEnd = 14
	colTitle = 6  # 
	for i in range(rowStart, rowEnd, 1):
		# Title
		cell = sheet.cell(i, colTitle)
		title = '{0}'.format(cell.value)
		lRow.append(title)
	# td.uInput(lRow)
	return lRow

# Get the Housing Data for a Market
def getHousingData():
	# print('\n getHousingData')
	lRow = []
	rowStart = 15
	rowEnd = 21

	titleVals = True
	for r in range(rowStart, rowEnd):  # Cycle through rows
		# Value
		if titleVals:
			titleCell = sheet.cell(r - 1, 0)
			colValue = 0
			for c in range(colValue, 3):  # Cycle through columns
				cell = sheet.cell(r, c)
				lRow.append(formatCell(cell, 'Title', titleCell))
			titleVals = False
		else:
			colValue = 1
			titleCell = sheet.cell(r - 1, 0)
			for c in range(colValue, 4):  # Cycle through columns
				cell = sheet.cell(r, c)
				lRow.append(formatCell(cell, 'Value', titleCell))
			titleVals = True
	return lRow

# Get the Employment Data for a Market
def getEmploymentData():
	# print('\n getEmploymentData')
	
	lRow = []
	rowStart = 25
	rowEnd = 35

	titleVals = True
	for r in range(rowStart, rowEnd):  # Cycle through rows
		# print(titleVals)
		# Skip blank rows
		cell = sheet.cell(r, 0)
		if cell.value == '':
			continue
		if cell.value == 'STATE 2 NOT USED':
			lRow.append('')
			lRow.append('')
			lRow.append('')
			lRow.append('')
		# Value
		elif titleVals:
			titleCell = sheet.cell(r - 1, 0)
			colValue = 0
			for c in range(colValue, 3):  # Cycle through columns
				cell = sheet.cell(r, c)	
				# print(cell.value)
				lRow.append(formatCell(cell, 'Title', titleCell))
				# print formatCell(cell, 'Title', titleCell)
			titleVals = False
			isFirstTime = True
			countKCI = 0
		else:
			colValue = 0
			titleCell = sheet.cell(r - 1, 0)
			for c in range(colValue, 4):  # Cycle through columns
				cell = sheet.cell(r, c)
				# print(cell.value)
				vcell = formatCell(cell, 'Value', titleCell)
				# print(vcell)
				# Remove decimal from large total employment numbers
				if ',' in vcell:
					vcell = vcell[:-2]
				lRow.append(vcell)
			if market == 'KCI':
				if countKCI < 2:
					countKCI += 1
				else:
					titleVals = True
			elif isFirstTime:
				isFirstTime = False
			else:
				titleVals = True
	return lRow

# Get the Employment Growth % for a Market
def getEmploymentGrowthData():
	# print('\n getEmploymentGrowthData')
	lRow = []
	rowStart = 35
	rowEnd = 39

	titleVals = True
	for r in range(rowStart, rowEnd):  # Cycle through rows
		# print('\nstart')
		celltemp = sheet.cell(r - 1, 0)
		# print(celltemp.value )
		# if celltemp.value == '':
		# 	print('here4')
		# 	continue
		if celltemp.value == 'STATE 2 NOT USED':
			pass
		# Title
		elif titleVals:
			# print('here6')
			titleCell = sheet.cell(r - 1, 0)
			cell = sheet.cell(r, 0)
			# print(cell.value)
			lRow.append(formatCell(cell, 'Title', titleCell))
			titleVals = False
		# Values
		else:
			# print('here7')
			titleCell = sheet.cell(r - 1, 0)
			colValue = 0
			for c in range(colValue, 2):  # Cycle through columns
				cell = sheet.cell(r, c)
				# print(cell.value)
				lRow.append(formatCell(cell, 'Value', titleCell))
	# td.uInput('\n Continue... > ')
	return lRow

# Get the MLS Data for a Market
def getMLSData():
	lRow = []
	rowStart = 43
	rowEnd = 54

	titleVals = True
	for r in range(rowStart, rowEnd):  # Cycle through rows
		# Skip blank rows
		cell = sheet.cell(r, 1)
		if cell.value == '':
			continue
		# Value
		if titleVals:
			titleCell = sheet.cell(r - 1, 0)
			colValue = 0
			for c in range(colValue, 3):  # Cycle through columns
				cell = sheet.cell(r, c)
				lRow.append(formatCell(cell, 'Title', titleCell))
			titleVals = False
		else:
			titleCell = sheet.cell(r - 1, 0)
			colValue = 1
			for c in range(colValue, 4):  # Cycle through columns
				cell = sheet.cell(r, c)
				cell = formatCell(cell, 'Value', titleCell)
				lRow.append(cell)
			titleVals = True
	# td.uInput(lRow)
	return lRow

def employmentGrowthByCategory():
	lRow = []
	rowStart = 58
	rowEnd = 62
	colTitle = 0
	colValue = 3
	iconNumber = 1
	iconSourcePath = 'F:/Research Department/MIMO/Market Insights/Template Data/Icons/'
	iconDesinationPath = 'F:/Research Department/MIMO/Market Insights/Template Data/'
	for i in range(rowStart, rowEnd):
		# Title
		cell = sheet.cell(i, colTitle)
		title = '{0}'.format(cell.value)
		lRow.append(title)
		cell = sheet.cell(i, colValue)
		wcell = '{0:.1%}'.format(cell.value)
		lRow.append(wcell)

		# Format Icon
		iconDesinationPNG = '{0}{1}_Emp_Cat_Icon_{2}.png'.format(iconDesinationPath, market, iconNumber)
		if 'Education' in title:
			iconSourcePNG = '{0}EducationHealthServices.png'.format(iconSourcePath)
		elif 'Financial' in title:
			iconSourcePNG = '{0}FinancialActivities.png'.format(iconSourcePath)
		elif 'Government' in title:
			iconSourcePNG = '{0}Government.png'.format(iconSourcePath)
		elif 'Information' in title:
			iconSourcePNG = '{0}Information.png'.format(iconSourcePath)
		elif 'Leisure' in title:
			iconSourcePNG = '{0}LeisureHospitality.png'.format(iconSourcePath)
		elif 'Manufacturing' in title:
			iconSourcePNG = '{0}Manufacturing.png'.format(iconSourcePath)
		elif 'Mining' in title:
			if 'Construction' in title:
				iconSourcePNG = '{0}MiningLoggingConstruction.png'.format(iconSourcePath)
			else:
				iconSourcePNG = '{0}MiningLogging.png'.format(iconSourcePath)
		elif 'Construction' in title:
			iconSourcePNG = '{0}Construction.png'.format(iconSourcePath)
		elif 'Personal' in title:
			iconSourcePNG = '{0}PersonalServices.png'.format(iconSourcePath)
		elif 'Business' in title:
			iconSourcePNG = '{0}ProfessionalBusinessServices.png'.format(iconSourcePath)
		elif 'Trade' in title:
			iconSourcePNG = '{0}TradeTransportationUtilities.png'.format(iconSourcePath)
		shutil.copy(iconSourcePNG, iconDesinationPNG)
		iconNumber += 1
	return lRow

lao.banner('Table Stats to InDesign CSV v05')

yearquarter = lao.getDateQuarter(lastquarter=True)
miexcelpath = 'F:/Research Department/MIMO/Spreadsheets/{0}'.format(yearquarter)

# Make Year & Quarter list to extend to end of csv row 2
yr = yearquarter[2:4]
qtr = '{0}{1}'.format(yearquarter[5:], yearquarter[4:5])
lYrQtr = [yr, qtr]

mitemplatepath = 'F:/Research Department/MIMO/Market Insights/Template Data/'
while 1:
	miexcelpathfilename = lao.guiFileOpen(miexcelpath, 'Open Market Insights Spreadsheet', [('MI Excel files', '.xlsm')])
	if 'CPI Mort Pop' in miexcelpathfilename:
		td.warningMsg('\n Wrong file selected...')
		print('  Try again...')
		lao.sleep(2)
	else:
		break
# miexcelpath = os.path.dirname(miexcelpathfilename)
miexportcsvpath = '{0}/{1}/'.format(os.path.dirname(miexcelpathfilename), 'InDesign CSVs')
miexcelfilename = os.path.basename(miexcelpathfilename)

# book = xlrd.open_workbook(miexcelpathfilename)
book = openpyxl.load_workbook(miexcelpathfilename)

# lMarketSheets = ['ABQ']
lMarketSheets = ['ATL', 'AUS', 'BOI', 'CLT', 'DEN', 'DFW', 'GSP', 'HOU', 'HNT', 'JAX', 'KCI', 'LVS', 'NAZ', 'NSH', 'ORL', 'PHX', 'PIN', 'RNO', 'SLC', 'SRQ', 'TPA', 'TUC']

for market in lMarketSheets:
	print(f' Market: {market}')
	# if market != 'KCI':
	# 	continue
	# Build csv line
	# sheet = book.sheet_by_name(market)
	sheet = book[market]
	lLineOut = getCoverData()
	lLine = getHousingData()
	lLineOut.extend(lLine)
	lLine = getEmploymentData()
	lLineOut.extend(lLine)
	lLine = getEmploymentGrowthData()
	lLineOut.extend(lLine)
	lLine = getMLSData()
	lLineOut.extend(lLine)
	lLine = employmentGrowthByCategory()
	lLineOut.extend(lLine)
	lLine = getChartTitles()
	lLineOut.extend(lLine)
	lLineOut.extend(lYrQtr)

	# Make CSV files in the Excel folder and the Template Folder
	outFileNameExcelPath = '{0}{1}_inDesign_Mail_Merge.csv'.format(miexportcsvpath, market)
	outFileNameTemplatePath = '{0}{1}_inDesign_Mail_Merge.csv'.format(mitemplatepath, market)
	with open(outFileNameExcelPath, 'wb') as e, open(outFileNameTemplatePath, 'wb') as t:
		eout = csv.writer(e, delimiter='\t')
		tout = csv.writer(t, delimiter='\t')
		eout.writerow(getHeader())
		tout.writerow(getHeader())
		eout.writerow(lLineOut)
		tout.writerow(lLineOut)

	# Print out export results
	with open(outFileNameTemplatePath, 'rb') as f:
		fin = csv.reader(f, delimiter='\t')
		lHeader = fin.next()
		lData = fin.next()
	numCols = len(lHeader)
	# print numCols
	# for i in range(0, numCols):
	# 	print '{0:30}: {1}'.format(lHeader[i], lData[i])
	# exit('\n Fin...')
	# td.uInput('continue...')
# lao.openFile('C:/TEMP/inDesign File.csv')