# E-Blast OPR Generator
# Sends a One Page Report on an E-Blast Campaign with an attached spreadsheet
#   of the results of the E-Blast

import aws
import bb
import csv
import datetime
import dicts
import django
from django.template import Template, Context
from django.conf import settings
import emailer
import fun_login
import fun_text_date as td
import lao
import mc
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import pandas as pd
from pprint import pprint
import sys
# from webs import awsUpload
import win32gui

def createRow(table, name, top100, company, phone, email):
	table = '{0}<tr>'.format(table)
	table = '{0}<td width="160">{1}</td>\r\n'.format(table, name)
	table = '{0}<td width="80">{1}</td>\r\n'.format(table, top100)
	table = '{0}<td width="260">{1}</td>\r\n'.format(table, company)
	table = '{0}<td width="160">{1}</td>\r\n'.format(table, phone)
	table = '{0}<td width="40">{1}</td>\r\n'.format(table, email)
	table = '{0}</tr>'.format(table)
	return table

def create_excel_report(dSentTo, campTitle):
	print('\n Creating Excel report...')
	# Create dataframes for each category
	clicks_data = []
	opens_data = []
	not_opened_data = []
	bounced_data = []
	unsub_data = []
	
	for row in dSentTo:
		dSentTo[row]['Email']
		try:
			top_100_url = lao.formatAsExcelHyperlink('MVP', dSentTo[row]['TFID'], dSentTo[row]['Top100'])
		except KeyError:
			td.warningMsg(' Could not create Top 100 link for:')
			pprint(dSentTo[row])
			continue
			
		data_row = {
			'MVP Status': top_100_url,
			'Name': dSentTo[row]['Person'],
			'Company': dSentTo[row]['Company'],
			'Phone': dSentTo[row]['Phone'],
			'Email': dSentTo[row]['Email']
		}

		if dSentTo[row]['Action'] == 'CLICKED':
			clicks_data.append(data_row)
		elif dSentTo[row]['Action'] == 'OPENED':
			opens_data.append(data_row)
		elif dSentTo[row]['Action'] == 'NOT OPENED':
			not_opened_data.append(data_row)
		elif 'BOUNCED' in dSentTo[row]['Action']:
			bounced_data.append(data_row)
		elif 'UNSUBSCRIBED' in dSentTo[row]['Action']:
			unsub_data.append(data_row)

	# Create Excel writer object
	excel_file = f'C:/TEMP/{campTitle}_results.xlsx'
	writer = pd.ExcelWriter(excel_file, engine='openpyxl')

	# Write each dataframe to a separate sheet with title
	sheets_data = {
		'Clicks': clicks_data,
		'Opens': opens_data,
		'Not Opened': not_opened_data, 
		'Bounced': bounced_data,
		'Unsubscribed': unsub_data
	}

	for sheet_name, data in sheets_data.items():
		# Create DataFrame
		df = pd.DataFrame(data)
		
		# Write to Excel with offset for title
		df.to_excel(writer, sheet_name=sheet_name, startrow=2, index=False)
		
		# Get worksheet
		worksheet = writer.sheets[sheet_name]
		
		# Add campaign title and sheet name
		worksheet['A1'] = campTitle
		campaign_title_cell = worksheet['A1']
		campaign_title_cell.font = Font(size=16, bold=True)
		campaign_title_cell.alignment = Alignment(horizontal='left')

		worksheet['A2'] = sheet_name
		sheet_name_cell = worksheet['A2']
		sheet_name_cell.font = Font(size=16, bold=True)
		sheet_name_cell.alignment = Alignment(horizontal='left')

		# Freeze panes below header row (row 4 since we have campaign title, sheet name, blank row, and header)
		worksheet.freeze_panes = 'A4'

	# Access workbook and format sheets
	workbook = writer.book
	for sheet_name in workbook.sheetnames:
		worksheet = workbook[sheet_name]
		
		# Format header
		for cell in worksheet[1]:
			cell.font = Font(bold=True)
			cell.fill = PatternFill(start_color='551A8B', end_color='551A8B', fill_type='solid')
			cell.font = Font(color='FFFFFF', bold=True)

		# Format Hyperlink columns hyperlinks
		for row in worksheet.iter_rows(min_row=5):  # Start after header row
			for i in range(0, len(row)):
				if row[i].value and '=HYPERLINK' in str(row[i].value):
					row[i].font = Font(color='0000FF', underline='single')
			
		# Adjust column widths
		for idx, column in enumerate(worksheet.columns):
			if idx == 0:  # First column
				worksheet.column_dimensions[column[0].column_letter].width = 12
				continue
				
			max_length = 0
			column = list(column)
			for cell in column:
				try:
					cell_value = str(cell.value)
					if '=HYPERLINK' in cell_value:
						# Extract just the display text between the last quote marks
						display_text = cell_value.split('"')[-2]
						length = len(display_text)
					else:
						length = len(str(cell_value))
					if length > max_length:
						max_length = length
				except:
					pass
			adjusted_width = (max_length + 2)
			worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

	writer.close()
	return excel_file

def choose_report_recipient(senderName, campTitle):

	dStaff = dicts.get_staff_dict(dict_type='full', skipFormerEmployees=True)

	lEblastTeams = []
	for row in dStaff:
		if row == None:
			break
		lEblastTeamCell = dStaff[row]['MC Audience'].split(',')
		for cell in lEblastTeamCell:
			if cell in lEblastTeams:
				continue
			else:
				lEblastTeams.append(cell)
	lEblastTeams.sort()
	dEblastTeams = {}
	count = 1
	for team in lEblastTeams:
		if team != 'None':
			dEblastTeams[count] = team
			count += 1

	lRecipient = []
	OPRsender = 'LAO Research <research@landadvisors.com>'
	while 1:

		# Print Menu
		td.banner('MC OPR v08')
		print('\n Campaign: {0}'.format(campTitle))
		eblastTeam = dStaff[senderName]['MC Audience']

		# Print the menu of MC Groups
		print(' Sent by: {0} : {1}\n'.format(senderName, eblastTeam))
		for i in range(1, len(dEblastTeams)+1):
			if eblastTeam == dEblastTeams[i]:
				td.colorText(' {0:3})  {1}'.format(i, dEblastTeams[i]), 'YELLOW')
			else:
				print(' {0:3})  {1}'.format(i, dEblastTeams[i]))

		td.colorText('\n 111)  Send EBlast OPR', 'CYAN')
		td.colorText(' 999)  Send Test OPR', 'CYAN')
		print('\n  00)  Quit')

		ui = td.uInput('\n Select Team > ')

		if ui == '00':
			exit('\n Terminating program...')
		else:
			ui = int(ui)

		if ui == 999:
			# OPRsender = 'LAO Research <research@landadvisors.com>'
			lRecipient = ['Bill Landis <blandis@landadvisors.com>']
			print('\n Sending Test EBlast OPR...')
			return lRecipient, OPRsender, True
		elif ui == 111:
			try:
				lRecipient.remove('Land Advisors Events <events@landadvisors.com>')
			except ValueError:
				pass
			for rec in lRecipient:
				print(rec)
			return lRecipient, OPRsender, False

		elif ui <= len(lEblastTeams)+1:
			team = dEblastTeams[ui]
			for agent in dStaff:
				if (team in dStaff[agent]['MC Audience'] or 'Marketing' in dStaff[agent]['MC Audience']) and dStaff[agent]['LAO'] == 'Yes' and dStaff[agent]['Roll'] != 'MC Group':
					agent = agent.replace('.', '')
					lRecipient.append('{0} <{1}@landadvisors.com>'.format(agent, dStaff[agent]['Email']))
		else:
			td.warningMsg('\n Invalid entry...try again...')
			lao.holdup()

client = fun_login.MailChimp()

daysAgo60 = datetime.datetime.now() + datetime.timedelta(-60)

TEMPLATES = [
	{
		'BACKEND': 'django.template.backends.django.DjangoTemplates',
		'DIRS': 'F:/Research Department/Code/RP3/templates'
	}
]
settings.configure(TEMPLATES=TEMPLATES)
django.setup()

mcCamps = 'None'

# Set console to MailChimp format
user = lao.getUserName()

if user == 'blandis':
	lao.consoleWindowPosition(position='MailChimp3')

td.console_title('MC OPR v08')

# Get Contacts from TF
td.banner('MC OPR v08')
# Get all TF Accounts
print(' Generating TF Account dictionary\n This takes a minute, please standby...')
service = fun_login.TerraForce()

dContacts = bb.get_contacts_dict(service)


while 1:
	td.banner('MC OPR v08')

	# Select the Campaign
	campTitle, CID, listID, senderEmail, senderName, mcCamps = mc.selectCampaign(client, 'MailChimp Report', mcCamps)
	
	campTitleCSVFileName = campTitle.replace('/', '-')

	recipients, OPRsender, oprTest = choose_report_recipient(senderName, campTitle)
	lao.consoleColor('BLUE')

	# Get the dictionary of Members sent the Campaign
	dSentTo = mc.getReportSentTo(client, CID)

	# Get the dictionary of Members Actions
	dRecipiant_actions = mc.get_campaign_actions_dict(client, CID, listID)

	# If not Sent To anyone remove it from the list of sent campaigns
	if dSentTo == {}:
		with open('{0}skipMailChimpReport.txt'.format(lao.getPath('skipfile')), 'ab') as cs:
			cs.write('{0}\r\n'.format(campTitle))
		lao.consoleColor('RED')
		td.banner('MC OPR v08')
		print('Campaign has no recipients.')
		sys.exit('\nfin')

	# Get Sender name
	# TerraForce Query
	fields = 'default'
	wc = "PersonEmail = '{0}'".format(senderEmail)
	results = bb.tf_query_3(service, rec_type='Person', where_clause=wc, limit=None, fields=fields)
	if results == []:
		sender = senderEmail
	else:
		sender = results[0]['Name']

	sender = sender.title()

	for row in dSentTo:
		print(f' {dSentTo[row]['Name']} : {dSentTo[row]['Email']}')
		# Add Action taken by List Member
		try:
			dSentTo[row]['Action'] = dRecipiant_actions[dSentTo[row]['Email'].lower()]
		except KeyError:
			continue
		# Get TF Info
		# Add Fields
		dSentTo[row]['TFID'] = 'No TF Record'
		email = dSentTo[row]['Email'].replace("'", "")

		# Find contact by email address and set to variable tfRow
		tfRow = {'Id': '', 'Name': '', 'Phone': '', 'Company__r': 'None', 'Top100__c': [], 'PersonEmail': ''}
		# for contact in dContacts:
		# 	if contact['PersonEmail'] == email:
		# 		tfRow = contact
		if email in dContacts:
			tfRow = dContacts[email]
		# sentToNameClean = dSentTo[row]['Name'].decode('ascii', 'ignore')
		sentToNameClean = dSentTo[row]['Name']
		dSentTo[row]['Person'] = lao.formatAsExcelHyperlink('TerraForce', tfRow['Id'], sentToNameClean)
		# Is Top 100?
		dSentTo[row]['Top100'] = '.'
		if tfRow['Top100__c'] == [] or tfRow['Top100__c'] == 'None':
			dSentTo[row]['Top100'] = 'Add MVP'
		else:
			dSentTo[row]['Top100'] = 'MVP'
		# Add Company if exists
		if dSentTo[row]['Company'] == '':
			if tfRow['Company__r'] == 'None':
				dSentTo[row]['Company'] = 'None'
			else:
				dSentTo[row]['Company'] = tfRow['Company__r']['Name']
		# dSentTo[row]['TFID'] = lao.formatAsExcelHyperlink('TerraForce', tfRow['Id'], 'TF Link')
		dSentTo[row]['TFID'] = tfRow['Id']
		dSentTo[row]['Phone'] = tfRow['Phone']
		dSentTo[row]['Email'] = lao.formatAsExcelHyperlink('Email', dSentTo[row]['Email'])
		

		if 'UNSUBSCRIBED' in dSentTo[row]['Action']:
			lUnsub = (dSentTo[row]['Action']).split(':')
			dSentTo[row]['Action'] = 'UNSUBSCRIBED'
			dSentTo[row]['Company'] = 'Reason: {0}'.format(lUnsub[1])
		
	# Write the info to a csv
	outCSV = 'C:/TEMP/{0}.csv'.format(campTitleCSVFileName)
	with open(outCSV, 'w', newline='') as f:
		fout = csv.writer(f)
		fout.writerow(['Campaign: {0}'.format(campTitle)])
		fout.writerow(['Sender: {0}'.format(sender)])
		fout.writerow([])
		fout.writerow(['Name', 'Top 100', 'Company', 'Phone', 'Email', 'Action'])
		totalSentToLAO = 0

		for row in dSentTo:
			if 'LAND ADVISORS' in dSentTo[row]['Company'].upper():
				totalSentToLAO += 1
				continue
			line = []
			try:
				line.append(dSentTo[row]['Person'])
				# print(dSentTo[row]['Person'])
			except KeyError:
				line.append(dSentTo[row]['Name'])
				# print(dSentTo[row]['Name'])
			try:
				line.append(dSentTo[row]['Top100'])
			except KeyError:
				line.append('Add MVP')
			line.append(dSentTo[row]['Company'])
			line.append(dSentTo[row]['Phone'])
			line.append(dSentTo[row]['Email'])
			if dSentTo[row]['Action'] == 'SENT':
				dSentTo[row]['Action'] = 'NOT OPENED'
			line.append(dSentTo[row]['Action'])
			
			try:
				fout.writerow(line)
			except UnicodeEncodeError:
				print(line)
				td.warningMsg('UnicodeEncodeError')

	# Get Campaign Report data
	dCampResults = mc.getCampReport(client, CID)

	# Get the Email Subject
	emailSubject = (dCampResults['subject_line'])

	# Calculate the Send Date
	sendDate = (dCampResults['send_time'])
	# print sendDate
	sendHour = int(sendDate[11:13]) - 7
	if sendHour < 0:
		sendHour = sendHour + 12
	sendDate = '{0}{1}{2}'.format(sendDate[:11], str(sendHour), sendDate[13:19])
	sendDate = datetime.datetime.strptime(sendDate, '%Y-%m-%dT%H:%M:%S').strftime('%a. %b %d, %Y %I:%m%p')

	# Get the Campaign Results Stats
	totalSentTo = dCampResults['emails_sent']
	totalSentTo = totalSentTo  - totalSentToLAO
	clicksTotal = dCampResults['clicks']['clicks_total']
	clicksUniqueSubscribersTotal = dCampResults['clicks']['unique_subscriber_clicks']
	clickRate = '{0:0.1f}%'.format(float(dCampResults['clicks']['click_rate'] * 100))
	bounced = dCampResults['bounces']['hard_bounces'] + dCampResults['bounces']['soft_bounces']
	opens = dCampResults['opens']['unique_opens']
	openRate = '{0:0.1f}%'.format(float(dCampResults['opens']['open_rate'] * 100))
	notOpened = totalSentTo - opens

	# Write the Campaign Results to the csv
	with open('C:/TEMP/{0}.csv'.format(campTitleCSVFileName), 'r') as f:
		fin = csv.reader(f)
		# fin.next()
		next(fin)

		# Copy E-Blast html to FTP
		html = client.campaigns.content.get(CID)
		html = html['html']
		htmlFileName = '{0}.html'.format(campTitleCSVFileName.replace(' ', ''))
		with open('C:/Users/Public/Public Mapfiles/awsUpload/Maps/{0}'.format(htmlFileName), 'w', newline='') as f:
			f.write(html)
		# Upload to AWS
		aws.sync_opr_maps_comp_listings_folders_to_s3(delete_files=True)
		# awsUpload(delete_files=True)

		# Get the Sender
		sender = next(fin)
		sender = sender[0].replace('Sender: ', '')
		next(fin)
		next(fin)

		clickTable, openTable, sentTable, bouncedTable, unsubTable = '', '', '', '', ''
		# clickTable100, openTable100, sentTable100, bouncedTable100, unsubTable100 = '', '', '', '', ''
		tfurl = 'https://landadvisors.my.salesforce.com/'
		t100url = 'https://landadvisors.my.salesforce.com/apex/Top100Account?Id='
		unsubCount = 0
		for row in fin:
			action = row[5]
			if 'HYPERLINK' in row[0]:
				person_Link_Name = (row[0]).split(',', 1)
				tfid = person_Link_Name[0].replace('=HYPERLINK("https://landadvisors.my.salesforce.com/', '').replace('"', '')
				tfid = tfid[:-3]
				personName = person_Link_Name[1].replace('"', '').replace(')', '')
				name = "<a href='{0}{1}'>{2}</a>".format(tfurl, tfid, personName)
				top100 = "<a href='{0}{1}'>{2}</a>".format(t100url, tfid, row[1])
				company = row[2]
				tflink = "https://landadvisors.my.salesforce.com/00T/e?title=Call&who_id={0}&what_id={0}&followup=1&tsk5=Call&retURL=%2F{0}".format(tfid)
				phone = "<a href='{0}'>{1}</a>".format(tflink, row[3])
			else:
				name = row[0]
				nameSearch = name.replace(',', '+')
				top100 = row[1]
				if action == 'UNSUBSCRIBED':
					company = row[2]
				else:
					company = row[2]
					top100 = 'NOT IN TF'
				phone = "<a href='https://landadvisors.my.salesforce.com/_ui/search/ui/UnifiedSearchResults?searchType=2&sen=a0O&sen=001&sen=a0Q&sen=005&sen=00O&str={0}'>Search TF</a>".format(nameSearch)
			if 'HYPERLINK' in row[4]:
				emailLink = row[4].split(',', 1)
				email = (emailLink[1]).replace('"', '').replace(')', '')
			else:
				email = row[4]
			email = "<a href=mailto:{0}>Email</a>".format(email)

			# Create Tables for Clicks, Opens, Sent, Bounced, Unsubscribed
			if action == 'CLICKED':
				clickTable = createRow(clickTable, name, top100, company, phone, email)
			elif action == 'OPENED':
				openTable = createRow(openTable, name, top100, company, phone, email)
			elif action == 'NOT OPENED':
				sentTable = createRow(sentTable, name, top100, company, phone, email)
			elif 'BOUNCED' in action:
				bouncedTable = createRow(bouncedTable, name, top100, company, phone, email)
			elif action == 'UNSUBSCRIBED':
				unsubTable = createRow(unsubTable, name, top100, company, phone, email)
				unsubCount += 1

	htmldict = {'CAMPTITLE': campTitle, 'SEND_DATE': sendDate, 'TOTAL_SENT_TO': totalSentTo, 'EMAIL_SUBJECT': emailSubject, 'CLICKS_UNIQUE_SUBSCRIBERS': clicksUniqueSubscribersTotal, 'CLICK_RATE': clickRate, 'OPENS': opens, 'OPEN_RATE': openRate, 'NOT_OPENED': notOpened, 'BOUNCED': bounced, 'UNSUB': unsubCount, 'CLICK_TABLE': clickTable, 'OPEN_TABLE': openTable, 'SENT_TABLE': sentTable, 'BOUNCED_TABLE': bouncedTable, 'UNSUB_TABLE': unsubTable,'HTML_FILE_NAME': htmlFileName, 'SENDER': sender}

	html = open(r'F:\Research Department\Code\RP3\templates\mc_report_02.html', 'r').read()

	# assign Template
	t = Template(html)
	c = Context(htmldict)
	html = t.render(c)


	with open('C:/TEMP/MC_OPR.html', 'w', newline='') as h:
		h.write(html)

	checkForBlockedRecipient = False
	while 1:
		if checkForBlockedRecipient:
			# Load modified html
			with open('C:/TEMP/MC_OPR.html', 'r') as h:
				html = h.read()
		

		# Create Excel report
		if checkForBlockedRecipient is False:
			excel_file = create_excel_report(dSentTo, campTitleCSVFileName)

		# NEW EMAILER LIB ########################################################################
		subject = 'MailChimp Campaign Results: {0}'.format(campTitle)
		body = html
		attachments = [excel_file]
		sender_email = OPRsender
		# recipients = ['Bill Landis <blandis@landadvisors.com>', 'Bill Gmail <william.scott.landis@gmail.com>']
		
		print(' Recipients:')
		pprint(recipients)
		# ui = td.uInput('\n Continue [00]... > ')
		# if ui == '00':
		# 	exit('\n Terminating program...')
		
		print('\n Sending email report for\n\n {0}'.format(campTitle))

		emailer.send_email_ses(subject, body, sender_email, recipients, cc=None, bcc=None, attachments=attachments)

		##########################################################################
		write_soft_bounces_to_file = False
		lao.consoleColor('RED')
		ui = td.uInput('\n Did E-blast report send {0/1] > ')
		if ui == '1':
			if oprTest is False:
				lao.SkipFile(campTitle, 'skipMailChimpReport', 'WRITE')
				write_soft_bounces_to_file = True
				break
			elif oprTest is True:
				break
		elif ui == '0':
			ui = td.uInput('\n Try sending again [0/1]? > ')
			if ui == '0':
				checkForBlockedRecipient = False
				# lao.consoleColor('YELLOW')
				break
			else:
				checkForBlockedRecipient = True
				continue
		else:
			'\n Invalid input...try again...\n'

	# Writing soft bouces to the soft bounce file
	# Future MC OPRs will check this file for soft bounces and convert them to hard bounces and remove them from the list/audience in MC
	if write_soft_bounces_to_file:
		print('\n Writing soft bounces to file...')
		bounce_file = 'F:/Research Department/Code/Databases/MC_Soft_Bounces.txt'
		with open(bounce_file, 'a') as f:
			for email in dRecipiant_actions:
				if 'SOFT' in dRecipiant_actions[email]:
					strBounce_check = f'{email} : listID {listID}\n'
					f.write(f'{email} : listID {listID}\n')
	
print('\n Fin')

