# Dictionaries for python scripts
import lao
import fun_text_date as td
import pandas as pd
from pprint import pprint
from typing import Union, List, Dict

# DICTIONARIES ###############################################################
# Make Dictionary of Spreadsheet (CSV or Excel where Header is the dictionary keys
def spreadsheet_to_dict(filename, sheetname='None', capitalize_keys=False):	
	import csv
	# Determine if CSV or XLSX
	if '.CSV' in filename.upper():
		# with open(filename, 'r', encoding='utf-8') as f:

		with open(filename, 'r') as f:

			# Check if file is readable and has content
			first_line = f.readline().strip()
			if not first_line:  # If first line is empty
				print(f"File {filename} has no data (empty first line).")
				return {}
			
			# Reset file pointer to beginning
			f.seek(0)

			fin = csv.reader(f)

			if 'LAO_Sale_Eblast' in filename:
				next(fin)
				next(fin)
				next(fin)
			header = next(fin)
			if capitalize_keys:
				header_upper = [x.upper() for x in header]
				header = header_upper
			colcount = len(header)
			index = 0
			dSht = {}
			for row in fin:
				# Trap for Property Radar user agreement
				if 'User Agreement' in row[0]:
					break
				index += 1
				dSht[index] = {}
				for i in range(0, colcount):
					# print('{0} : {1}'.format([header[i]], row[i]))
					dSht[index][header[i]] = row[i]
					# except IndexError:
					# 	print(index)
					# 	print(i)
					# 	exit()
		
		return dSht
	
	elif '.XLSX' in filename.upper():
		from openpyxl import load_workbook
		import warnings

		# Suppress specific warnings
		warnings.filterwarnings("ignore", category=UserWarning, module='openpyxl')
		# Open the workbook, load it then close it so that it can be opened again and edited in Excel
		with open(filename, 'rb') as f:
			wb = load_workbook(f)

		# Set header row and first row of data if LLR or not
		if 'Land_Lot_Report' in filename:
			headerrow = 3
			firstrow = 4
			if sheetname == 'None':
				sheet = wb.worksheets[0]
			else:
				sheet = wb[sheetname]
		elif 'AnnualMarketPerformance' in filename:
			headerrow = 3
			firstrow = 4
			sheet = wb['Annual-Market Performance']
		elif 'Market Supply & Demand' in filename:
			headerrow = 5
			firstrow = 6
			sheet = wb['Supply and Demand']
		elif 'MO Python Chart Placement' in filename:
			headerrow = 1
			firstrow = 2
			sheet = wb['Main']
		elif 'Annoyatron' in filename:
			headerrow = 1
			firstrow = 2
			sheet = wb[sheetname]
		elif 'MailChimp Admin Report' in filename:
			headerrow = 3
			firstrow = 4
			sheet = wb['Group Membership']
		elif sheetname == 'Pipeline Project Details':
			headerrow = 3
			firstrow = 4
			sheet = wb['Pipeline Project Details']
		elif 'Staff_PTO' in filename:
			headerrow = 1
			firstrow = 2
			sheet = wb['PTO']
		else:
			headerrow = 1
			firstrow = 2
			if sheetname == 'None':
				sheet = wb.worksheets[0]
			else:
				sheet = wb[sheetname]
		colcount = sheet.max_column
		rowcount = sheet.max_row
		header = ['']
		for i in range(1, colcount + 1):
			header.append(sheet.cell(headerrow, i).value)
		index = 0
		dSht = {}
		for row_idx in range(firstrow, rowcount + 1):
			index += 1
			dSht[index] = {}
			for i in range(1, colcount + 1):
				dSht[index][header[i]] = sheet.cell(row_idx, i).value
		return dSht

	elif '.XLS' in filename.upper():
		import xlrd
		xl_wb = xlrd.open_workbook(filename)
		# Set header row and first row of data if LLR or not
		if 'Land_Lot_Report' in filename:
			headerrow = 2
			firstrow = 3
			xl_sheet = xl_wb.sheet_by_index(0)
		elif 'Market Supply & Demand' in filename:
			headerrow = 4
			firstrow = 5
			xl_sheet = xl_wb.sheet_by_name('Supply and Demand')
		elif 'MO Python Chart Placement' in filename:
			headerrow = 0
			firstrow = 1
			xl_sheet = xl_wb.sheet_by_name('Main')
		elif 'Annoyatron' in filename:
			headerrow = 0
			firstrow = 1
			xl_sheet = xl_wb.sheet_by_name(sheetname)
		elif 'census_permits' in filename:
			headerrow = 7
			firstrow = 8
			xl_sheet = xl_wb.sheet_by_name('MSA Units')
		else:
			headerrow = 0
			firstrow = 1
			xl_sheet = xl_wb.sheet_by_index(0)
		
		colcount = len(xl_sheet.row(0))
		# print(f'Col Count: {colcount}')
		header = []
		for i in range(0, colcount):
			header.append(xl_sheet.cell_value(headerrow, i))
		index = 0
		dSht = {}
		rowcount = xl_sheet.nrows
		# print(f'Row Count: {rowcount}')
		for row_idx in range(firstrow, rowcount + 1):
			index += 1
			dSht[index] = {}
			for i in range(0, colcount):
				try:
					dSht[index][header[i]] = xl_sheet.cell_value(row_idx, i)
				except IndexError:
					pass
		return dSht
	else:
		input(' Not a spreadsheet file...\n{0}').format(filename)

def load_staff_db(skip_former_employees=True) -> pd.DataFrame:
	"""
	Load the LAO staff database from an Excel file.
	
	Args:
		filepath: Path to the Excel file (default: 'LAO_Staff_Db_v03.xlsx')
		
	Returns:
		A pandas DataFrame containing the staff database
	"""
	filepath = r"F:\Research Department\Code\Databases\LAO_Staff_Db_v03.xlsx"

	# Load the Excel file
	df = pd.read_excel(filepath, sheet_name='staff')

	# Apply LAO filter if requested
	if skip_former_employees:
		df = df[df['LAO'] == 'Yes']
		
	return df

def get_staff_dict_2(dict_key='Name', skip_former_employees=True) -> Dict[str, Dict[str, str]]:
	"""
	Return a dictionary of staff member details where Name is the key.
	
	Returns:
		A dictionary where:
			- Key is Name
			- Value is a dictionary containing Roll, Office, and marketAbb...
	"""
	df = load_staff_db(skip_former_employees=skip_former_employees)
	
	# Initialize the result dictionary
	dResults = {}
	
	# Include all columns as fields
	fields = df.columns.tolist()
	
	# For each person
	for _, row in df.iterrows():
		# Create sub-dictionary with required fields
		person_dict = {field: row[field] for field in fields}
		dResults[row[dict_key]] = person_dict
	
	return dResults

# Get Dictionary of agents & staff
def get_staff_dict(dict_type='full', skipFormerEmployees=True):
	# Do not use this function for because it crashes the GUI
	# lao.print_function_name('dict get_staff_dict')

	inFile = 'F:/Research Department/Code/Databases/LAO_Staff_Db_v03.xlsx'
	dTemp = spreadsheet_to_dict(inFile)
	dStaff = {}
	for i in dTemp:
		if skipFormerEmployees:
			if dTemp[i]['LAO'] == 'No':
				continue
		name = dTemp[i]['Name']
		dStaff[name] = {}
		dStaff[name]['City'] = dTemp[i]['City']
		dStaff[name]['MC Audience'] = dTemp[i]['MC Audience']
		dStaff[name]['Email'] = dTemp[i]['Email']
		dStaff[name]['Id'] = dTemp[i]['Id']
		dStaff[name]['UserID'] = dTemp[i]['UserID']
		dStaff[name]['LAO'] = dTemp[i]['LAO']
		dStaff[name]['MC Aud ID'] = dTemp[i]['MC Aud ID']
		dStaff[name]['MC New Subscriber Market'] = dTemp[i]['MC New Subscriber Market']
		dStaff[name]['Markets'] = dTemp[i]['Markets']
		dStaff[name]['Office'] = dTemp[i]['Office']
		dStaff[name]['Phone'] =dTemp[i]['Phone']
		dStaff[name]['Phn Desk'] =dTemp[i]['Phn Desk']
		dStaff[name]['Phn Mobile'] =dTemp[i]['Phn Mobile']
		dStaff[name]['Phn Extention'] =dTemp[i]['Phn Extention']
		dStaff[name]['Roll'] = dTemp[i]['Roll']
		dStaff[name]['State'] = dTemp[i]['State']
		dStaff[name]['State2'] = dTemp[i]['State2']
		dStaff[name]['Zipcode'] = dTemp[i]['Zipcode']
		dStaff[name]['MktMailerMkt'] = dTemp[i]['MktMailerMkt']
		dStaff[name]['marketAbb'] = dTemp[i]['marketAbb']
		dStaff[name]['SF Password'] = dTemp[i]['SF Password']
		dStaff[name]['SF Token'] = dTemp[i]['SF Token']
		dStaff[name]['PC Name'] = dTemp[i]['PC Name']
	
	if dict_type == 'full':
		return dStaff
	# Return dict of MailChimp Report agent selection list
	elif dict_type == 'mc':
		lStaff = [] # Must create a list to sort before adding numbers
		for staff in dStaff:
			if dStaff[staff]['Roll'] == 'Agent' and dStaff[staff]['State2'] != 'SC':
				string = '{0}-{1:20} {2}'.format(dStaff[staff]['State2'], dStaff[staff]['MC Audience'], staff)
				# print(string)
				lStaff.append(string)
		lStaff.sort()
		d = {}
		i = 0
		for staff in lStaff:
			i += 1
			stri = str(i)
			d[stri] = staff
		return d
	elif dict_type == 'mcDict':
		dict = {}
		for staff in dStaff:
			d = dStaff[staff]
			dict[d['Email']] = d['MC Aud ID']
		return dict
	# Return dict of Terraforce Ids and staff name
	elif dict_type == 'tfid':
		d = {}
		for staff in dStaff:
			d[dStaff[staff]['Id']] = staff
		return d
	elif dict_type == 'userid':
		d = {}
		for staff in dStaff:
			d[dStaff[staff]['UserID']] = staff
		return d
	# Return dict of Email names and staff name
	elif dict_type == 'email':
		d = {}
		for staff in dStaff:
			d[dStaff[staff]['Email']] = staff
		return d
	# Return list of staff names
	elif dict_type == 'namesonly':
		l= []
		for staff in dStaff:
			if dStaff[staff]['Roll'] == 'Agent' and dStaff[staff]['LAO'] == 'Yes':
				l.append(staff)
		return l
	elif dict_type == 'modify':
		lao.openFile(inFile)
	elif dict_type == 'market':
		d = {}
		for staff in dStaff:
			d[dStaff[staff]['Email']] = dStaff[staff]['Markets']
		return d
	elif dict_type == 'researchers':
		d = {}
		for staff in dStaff:
			if staff == 'Ethan Granger':
				continue
			# if dStaff[staff]['Roll'] == 'Research' or 'Intern' in dStaff[staff]['Roll']:
			if dStaff[staff]['Roll'] == 'Research':
				if dStaff[staff]['UserID'] == 'None':
					continue
				d[staff] = dStaff[staff]['UserID']
		return d
	elif dict_type == 'userids':
		d = {}
		for staff in dStaff:
			if dStaff[staff]['UserID'] == 'None':
				continue
			d[staff] = dStaff[staff]['UserID']
		return d
	elif dict_type == 'staffoffice':
		d = {}
		for staff in dStaff:
			if dStaff[staff]['Roll'] == 'Agent':
				d[staff] = dStaff[staff]['marketAbb']
		return d
	elif dict_type == 'marketabblist':
		lMarkets = []
		for staff in dStaff:
			if not dStaff[staff]['marketAbb'] in lMarkets:
				lMarkets.append(dStaff[staff]['marketAbb'])
		return lMarkets
	elif dict_type == 'marketfulllist':
		lMarkets = []
		for staff in dStaff:
			ltemp = dStaff[staff]['Markets'].split(',')
			for market in ltemp:
				if market == 'None' or market == 'Conservation' or market == 'Seattle':
					continue
				elif market == 'Agriculture':
					market = 'Yuma'
				elif market == 'Northern Arizona':
					market = 'Prescott'
				if not market in lMarkets:
					lMarkets.append(market)
		lMarkets.sort()
		return lMarkets

# State Abbreviations Dictionary
def get_state_abbriviations_dict():
	dStateAbbr = {
		'AK': 'Alaska',
		'AL': 'Alabama',
		'AZ': 'Arizona',
		'AR': 'Arkansas',
		'CA': 'California',
		'CO': 'Colorado',
		'CT': 'Connecticut',
		'DE': 'Delaware',
		'FL': 'Florida',
		'GA': 'Georgia',
		'HI': 'Hawaii',
		'ID': 'Idaho',
		'IL': 'Illinois',
		'IN': 'Indiana',
		'IA': 'Iowa',
		'KS': 'Kansas',
		'KY': 'Kentucky',
		'LA': 'Louisiana',
		'ME': 'Maine',
		'MD': 'Maryland',
		'MA': 'Massachusetts',
		'MI': 'Michigan',
		'MN': 'Minnesota',
		'MS': 'Mississippi',
		'MO': 'Missouri',
		'MT': 'Montana',
		'NE': 'Nebraska',
		'NV': 'Nevada',
		'NH': 'New Hampshire',
		'NJ': 'New Jersey',
		'NM': 'New Mexico',
		'NY': 'New York',
		'NC': 'North Carolina',
		'ND': 'North Dakota',
		'OH': 'Ohio',
		'OK': 'Oklahoma',
		'OR': 'Oregon',
		'PA': 'Pennsylvania',
		'RI': 'Rhode Island',
		'SC': 'South Carolina',
		'SD': 'South Dakota',
		'TN': 'Tennessee',
		'TX': 'Texas',
		'UT': 'Utah',
		'VT': 'Vermont',
		'VA': 'Virginia',
		'WA': 'Washington',
		'WV': 'West Virginia',
		'WI': 'Wisconsin',
		'WY': 'Wyoming',
		'Alaska': 'AK',
		'Alabama': 'AL',
		'Arizona': 'AZ',
		'Arkansas': 'AR',
		'California': 'CA',
		'Colorado': 'CO',
		'Connecticut': 'CT',
		'Delaware': 'DE',
		'Florida': 'FL',
		'Georgia': 'GA',
		'Hawaii': 'HI',
		'Idaho': 'ID',
		'Illinois': 'IL',
		'Indiana': 'IN',
		'Iowa': 'IA',
		'Kansas': 'KS',
		'Kentucky': 'KY',
		'Louisiana': 'LA',
		'Maine': 'ME',
		'Maryland': 'MD',
		'Massachusetts': 'MA',
		'Michigan': 'MI',
		'Minnesota': 'MN',
		'Mississippi': 'MS',
		'Missouri': 'MO',
		'Montana': 'MT',
		'Nebraska': 'NE',
		'Nevada': 'NV',
		'New Hampshire': 'NH',
		'New Jersey': 'NJ',
		'New Mexico': 'NM',
		'New York': 'NY',
		'North Carolina': 'NC',
		'North Dakota': 'ND',
		'Ohio': 'OH',
		'Oklahoma': 'OK',
		'Oregon': 'OR',
		'Pennsylvania': 'PA',
		'Rhode Island': 'RI',
		'South Carolina': 'SC',
		'South Dakota': 'SD',
		'Tennessee': 'TN',
		'Texas': 'TX',
		'Utah': 'UT',
		'Vermont': 'VT',
		'Virginia': 'VA',
		'Washington': 'WA',
		'West Virginia': 'WV',
		'Wisconsin': 'WI',
		'Wyoming': 'WY'
	}
	return dStateAbbr

# Make dictionary of Account fields (dAcc) for TerraForce
def get_blank_account_dict():
	dAcc = {'ADDRESSFULL':'None',
		'AGENT':'None',
		'AID': 'None',			# Account Person ID
		'AUTOFORMAT':False,
		'CATEGORY': 'None',
		'CCID':'None',
		'CITY':'None',
		'COUNTRY': 'None',
		'DELETECONTACT':False,
		'DEPARTMENT':'None',
		'DESCRIPTION':'None',
		'EID':'None',				# Entity ID
		'EMAIL':'None',
		'EMAILDOMAIN':'None',		# Email Domain
		'EMAILUSERNAME':'None',		# Email Username
		'EMAILOPTOUT': False,
		'ENTITY':'None',			# Company
		'FASTMODE': False,
		'FAX':'None',
		'IP':'None', 				# MC New Subscriber IP
		'LINKEDIN': 'None',			# Is Account Active
		'ISBROWSEROPEN': False,		# Track if browser opened for Enitity in Google/Entity domain
		'MCAUDID': 'None',			# MC Audience ID
		'MCAUDNAME': 'None',		# MC Audience Name
		'MCLCAMPAIGNIDS': 'None',	# MC List of Campaigns
		'MCSTATUS': 'subscribed', 	# MC default subscribed
		'MCSUBMSG': 'None',			# MC Subscirber message/comments
		'MIMARKETS': 'None',	 	# list of MI markets subscriber wants
		'MARKETMAILER':False,
		'MOBILE':'None',			# Person Mobile Phone
		'NAME':'None',				# Person Name
		'NF':'None',				# First Name
		'NM':'None',				# Middle Name
		'NL':'None',				# Last Name
		'NOTECONTENT':'None',		# Content for Note
		'NOTETITLE':'None',			# Title for Note
		'PHONE':'None',
		'PHONEENTITY':'None',
		'PHONEHOME':'None',
		'PHONEOTHER':'None',
		'SOURCE': 'None',			# Source of the date (i.e. parcel, TPS, DnB...)
		'STATE':'None',
		'STATEOFORIGIN':'None',
		'STREET':'None',
		'STRAIGHT2TF': False,
		'TITLEPERSON': 'None',
		'TOP100AGENT':'None',
		'UPDATERECORD':False,
		'URL':'None',
		'WEBSITE': 'None',
		'ZIPCODE':'None'}
	return dAcc

# Get TF fields dict
def get_blank_tf_deal_dict():
	dTF = {
		'Id': 'None',
		'AccountId__c': 'None',
		'Acres__c': 0,
		'Beneficiary__c': 'None',
		'Beneficiary_Contact__c': 'None',
		'Buyer_Acting_As__c': 'None',
		'City__c': 'None',
		'Classification__c': 'None',
		'County__c': 'None',
		'Description__c': 'None',
		'Development_Status__c': 'None',
		'Latitude__c': 0,
		'Lead_Parcel__c': 'None',
		'Legal_Description__c': 'None',
		'List_Date__c': 'None',
		'Listing_Expiration_Date__c': 'None',
		'List_Price__c': 0,
		'Loan_Amount__c': 0,
		'Loan_Date__c': 'None',
		'Location__c': 'None',
		'Longitude__c': 0,
		'Lot_Description__c': 'None',
		'Lot_Type__c': 'None',
		'Lots__c': 0,
		'Market__c': 'None',
		'Name': 'None',
		'OwnerId': 'None',
		'Owner_Entity__c': 'None',
		'OPR_Sent__c': 'None',
		'Parcels__c': 'None',
		'PID__c': 'None',
		'Recorded_Instrument_Number__c': 'None',
		'RecordTypeID': '012a0000001ZSS5AAO',
		'Sale_Date__c': 'None',
		'Sale_Price__c': 0,
		'Source__c': 'None',
		'Source_ID__c': 'None',
		'StageName__c': 'None',
		'State__c': 'None',
		'Subdivision__c': 'None',
		'Submarket__c': 'None',
		'type': 'lda_Opportunity__c',
		'Zipcode__c': 'None',
		'Zoning__c': 'None'
	}
	return dTF

# Get Buyer Seller (dbS) dict
def get_blank_buyer_seller_dict():
	dBS = {
		'BUYERENTITYID': 'None',
		'BUYERENTITYNAME':  'None',
		'BUYERPERSONID': 'None',
		'BUYERPERSONNAME': 'None',
		'SELLERENTITYID': 'None',
		'SELLERENTITYNAME': 'None',
		'SELLERPERSPNID': 'None',
		'SELLERPERSONNAME': 'None',
		}

	return dBS

# Raw Data Deal Classification Dictionary for Phoenix
def get_phx_raw_data_deal_classification_dict():
	dClass = {'Agricultural':'AGRICULTURAL',
	'Apartment Traditional':'APARTMENT',
	'Apartment Horizontal':'APARTMENT',
	'Apartment':'APARTMENT',
	'Build For Rent (platted)':'APARTMENT',
	'Church':'OTHER',
	'Conservation':'OTHER',
	'Data Center':'DATA CENTER',
	'Finished Lots':'RESIDENTIAL',
	'Golf':'COMMERCIAL',
	'Golf and Resort':'COMMERCIAL',
	'High Density Assisted Living':'APARTMENT',
	'High Density Residential':'RESIDENTIAL',
	'Homebuilder':'RESIDENTIAL',
	'Hospitality':'COMMERCIAL',
	'Industrial':'INDUSTRIAL',
	'Large Lot':'RESIDENTIAL',
	'Manufactured Home':'RESIDENTIAL',
	'Master Planned Community':'RESIDENTIAL',
	'Medical':'COMMERCIAL',
	'Medium Density Residential':'RESIDENTIAL',
	'Mixed Use':'MIXED USE',
	'Mixed-Use':'MIXED USE',
	'Multifamily':'APARTMENT',
	'Office':'COMMERCIAL',
	'Other':'OTHER',
	'Partially Improved':'RESIDENTIAL',
	'Partially Improved Lots':'RESIDENTIAL',
	'Platted':'RESIDENTIAL',
	'Platted and Engineered Lots':'RESIDENTIAL',
	'Public':'OTHER',
	'Residential':'RESIDENTIAL',
	'Retail':'COMMERCIAL',
	'Resort or Golf':'COMMERCIAL',
	'Ranch and Recreation':'COMMERCIAL',
	'Sand and Gravel':'INDUSTRIAL',
	'School':'SCHOOL',
	'School Charter':'SCHOOL',
	'Single Family':'RESIDENTIAL',
	'Solar Wind': 'INDUSTRIAL',
	'Speculative':'SPEC',
	'Storage':'INDUSTRIAL',
	'TIMBER':'AGRICULTURAL'}
	
	return dClass

# Get dicts of fields and keys for Debt_Processor
def get_debt_processor_dicts():
		# Property Radar = PR, Reonomy = RY, Vizzda = VZ
	dDebt_fields = {
		'APN': {'PR': 'APN', 'RY': 'apn', 'VZ': ''},
		'Acres': {'PR': 'Lot Acres', 'RY': 'lot_area', 'VZ': 'General Land Area'},
		'Zoning': {'PR': 'Zoning', 'RY': 'zoning', 'VZ': 'Property Description'},
		'County': {'PR': 'County', 'RY': 'county', 'VZ': 'County Id'},
		'State': {'PR': 'State', 'RY': 'address_state', 'VZ': 'Property State'},
		'Owner Entity': {'PR': 'Owner', 'RY': 'sale_buyer_name', 'VZ': 'True Owner/Buyer - Name'},
		'Owner Person': {'PR': 'Primary Name', 'RY': 'contact_name', 'VZ': 'Buyer Contact Point/Signatory - Name'},
		'Mail Street': {'PR': 'Mail Address', 'RY': 'reported_mailing_address_line_1', 'VZ': 'Legal Owner/Buyer - Addresses'},
		'Mail City': {'PR': 'Mail City', 'RY': 'reported_mailing_address_city', 'VZ': ''},
		'Mail State': {'PR': 'Mail State', 'RY': 'reported_mailing_address_state', 'VZ': ''},
		'Mail Zip': {'PR': 'Mail ZIP', 'RY': 'reported_mailing_address_postal_code', 'VZ': ''},
		'Lender': {'PR': '1st Orig Lender', 'RY': 'mortgage_recorded_name', 'VZ': 'Legal Lender - Name'},
		'Loan Amount': {'PR': '1st Amount', 'RY': 'mortgage_amount', 'VZ': 'Notes'},
		'Loan Date': {'PR': '1st Rec Date', 'RY': 'mortgage_recorded_date', 'VZ': 'Recorded Date'},
		'Loan Maturity Date': {'PR': '', 'RY': 'mortgage_maturity_date', 'VZ': ''},
		'Loan Purpose': {'PR': '1st Purpose', 'RY': '', 'VZ': 'Notes'},
		'Prop Type': {'PR': 'Type', 'RY': 'property_type', 'VZ': ''},
		'Prop Subtype': {'PR': '', 'RY': 'property_subtype', 'VZ': 'Property Sub Types'},
		'Last Sale Date': {'PR': 'Purchase Date', 'RY': 'sale_recorded_date', 'VZ': 'Recorded Date'},
		'Last Sale Amount': {'PR': 'Purchase Amt', 'RY': 'sale_amount', 'VZ': 'Sale Price'},
		'Foreclosure': {'PR': 'Foreclosure?', 'RY': 'preforeclosure_active_flag', 'VZ': ''},
		'FCL Doc Type': {'PR': 'FCL Doc Type', 'RY': 'preforeclosure_doc_type', 'VZ': ''},
		'FCL Rec Date': {'PR': 'FCL Rec Date', 'RY': 'preforeclosure_recording_date', 'VZ': ''},
		'FCL Stage': {'PR': 'FCL Stage', 'RY': '', 'VZ': ''},
		'Default Date': {'PR': 'Default Date', 'RY': '', 'VZ': ''},
		'Default Amt': {'PR': 'Default Amt', 'RY': '', 'VZ': ''},
		'Source': {'PR': 'Property Radar', 'RY': 'Reonomy', 'VZ': 'Vizzda'},
		'Source ID': {'PR': 'Radar ID', 'RY': 'reonomy_id', 'VZ': 'Event Id'},
		'Latitude': {'PR': 'Latitude', 'RY': 'latitude', 'VZ': 'Latitude'},
		'Longitude': {'PR': 'Longitude', 'RY': 'longitude', 'VZ': 'Longitude'}
	}
	dDup_keys = {
		'PR': ('Owner', 'County', 'State', 'Purchase Date', '1st Amount', '1st Rec Date', '1st Orig Lender'),
		'RY': ('sale_buyer_name', 'county', 'sale_document_id', 'mortgage_recorded_name', 'mortgage_amount', 'mortgage_recorded_date'),
		'VZ': ('Event Id', 'Property Name'),
		'MSTR': ('County', 'State', 'Loan Amount', 'Loan Date')
	}
	dSources = {
		'Property Radar Debt': {'src': 'PR', 'data_type': 'debt'},
		'Reonomy Debt': {'src': 'RY', 'data_type': 'debt'},
		'Property Radar Foreclosures': {'src': 'PR', 'data_type': 'foreclosure'},
		'Reonomy Foreclosures': {'src': 'RY', 'data_type': 'foreclosure'}
		# 'Vizzda Debt': {'src': 'VZ', 'data_type': 'debt'},
	}
	dViz_patterns = {
		'in new construction debt': 'backward',
		'in new Industrial Development Authority debt': 'backward',
		'in new Fanne Mae debt': 'backward',
		'revolving line of credit': 'backward',
		'in construction debt': 'backward',
		'in related construction debt': 'backward',
		'new construction debt': 'backward',
		'in cross-collateralized construction debt': 'backward',
		'in cross-collateralized': 'backward',
		'in new debt': 'backward', 'in seller-carried debt': 'backward',
		'in seller carried debt': 'backward',
		'in new lending from': 'backward',
		'seller-carry debt': 'backward',
		'new funding in the form of': 'backward',
		'new debt with': 'backward',
		'note with': 'backward',
		'in debt with entity': 'backward',
		'in debt with': 'backward',
		'debt with': 'backward',
		'seller-carried': 'backward',
		'seller-carrie.': 'backward',
		'seller-carry': 'backward',
		'loan with': 'backward',
		'construction mortgage with': 'backward',
		'mortgage with': 'backward',
		'an entity tracing to M & T Bank.': 'backward',
		'with a new debt of': 'forward',
		'encumbered by a new loan of': 'forward',
		'new Construction Loan of up to': 'forward',
		'new construction loan in the amount of': 'forward',
		'new construction loan of': 'forward',
		'group borrowed': 'forward',
		'construction loan of': 'forward',
		'with seller carrying': 'forward',
		'new loan in the amount of': 'forward',
		'new loan of': 'forward',
		'new total debt of': 'forward',
		'new contruction mortgages totalling': 'forward',
		'revolving construction loan of up-to': 'forward',
		'mortgage of up to': 'fowrard',
		'with a new construction mortgage of': 'forward',
		'new private loan of ': 'forward',
		'carrying a note of': 'forward',
		'loan in the amount of up to': 'forward',
		'loan in the amount of': 'forward',
		'loans in the total amount of': 'forward',
		'principal amount of': 'forward',
		'new construction loan for acquisition and development in the amount of up to': 'forward',
		'new construction mortgage in an amount of up to': 'forward',
		'new loans totaling': 'forward',
		'with an aggregate debt of': 'forward',
		'construction mortgage in the amount of': 'forward',
		'two new notes of': 'forward',
		'loan in the principal sum of': 'forward',
		'construction mortgages totaling': 'forward',
		'mortgage of up to': 'forward',
		'two construction loans of': 'forward',
		'new debt': 'backward'
	}
	# To rename PropertyRadar FCL Doc Type field
	dFCL_doc_type = {
		'COP': 'CERTIFICATE OF PARTICIPATION',
		'FJF': 'FINAL JUDGMENT OF FORECLOSURE',
		'FJG': 'FINAL JUDGMENT',
		'LIS': 'LIS PENDENS',
		'NOD': 'NOTICE OF DEFAULT',
		'NSL': 'NOTICE OF SALE',
		'NTS': 'NOTICE OF TRUSTEE SALE',
		'SBT': 'APPOINTMENT OF SUBSTITUTE TRUSTEE'
	}

	# To populate Reonomy FCL Stage field
	dFCL_stage = {
		'APPOINTMENT OF SUBSTITUTE TRUSTEE': 'PREFORECLOSURE',
		'FINAL JUDGMENT': 'PREFORECLOSURE',
		'LIS PENDENS': 'PREFORECLOSURE',
		'NOTICE OF DEFAULT': 'PREFORECLOSURE',
		'NOTICE OF FORECLOSURE': 'BANK OWNED',
		'NOTICE OF SALE': 'AUCTION',
		"NOTICE OF TRUSTEE'S SALE": 'AUCTION',
		'RELEASE OF PRE-FORECLOSURE': 'NOT IN FCL',
		'UNKNOWN': 'PREFORECLOSURE'	
	}
	# 		'an entity tracing to': 'backward',
	return dDebt_fields, dDup_keys, dSources, dViz_patterns, dFCL_doc_type, dFCL_stage

# Get Foreclosure Processor Dicts
def get_foreclosure_processor_dicts():
	# Property Radar = PR, Reonomy = RY, Vizzda = VZ
	dFCL_fields = {
		'Map': {'PR': 'Map', 'RY': 'Map'},
		'State': {'PR': 'State', 'RY': 'address_state'},
		'LAO Market': {'PR': 'LAO Market', 'RY': 'LAO Market'},
		'County': {'PR': 'County', 'RY': 'county'},
		'LAO Submarket': {'PR': 'LAO Submarket', 'RY': 'LAO Submarket'},
		'City': {'PR': 'City', 'RY': 'address_city'},
		'Acres': {'PR': 'Lot Acres', 'RY': 'lot_area'},
		'Zoning': {'PR': 'Zoning', 'RY': 'zoning'},
		'FCL Doc Type': {'PR': 'FCL Doc Type', 'RY': 'preforeclosure_doc_type'},
		'FCL Rec Date': {'PR': 'FCL Rec Date', 'RY': 'preforeclosure_recording_date'},
		'FCL Stage': {'PR': 'FCL Stage', 'RY': 'flc_stage'},
		'Default Date': {'PR': 'Default Date', 'RY': 'None'},
		'Default Amount': {'PR': 'Default Amt', 'RY': 'None'},
		'Owner Entity': {'PR': 'Owner', 'RY': 'sale_buyer_name'},
		'Owner Person': {'PR': 'Primary Name', 'RY': 'contact_name'},
		'Mail Street': {'PR': 'Mail Address', 'RY': 'reported_mailing_address_line_1'},
		'Mail City': {'PR': 'Mail City', 'RY': 'reported_mailing_address_city'},
		'Mail Address': {'PR': 'Mail State', 'RY': 'reported_mailing_address_state'},
		'Mail Zip': {'PR': 'Mail ZIP', 'RY': 'reported_mailing_address_postal_code'},
		'Phone': {'PR': 'Primary Phone1', 'RY': 'contact_phone_1'},
		'Mobile': {'PR': 'Primary Mobile Phone1', 'RY': 'contact_phone_2'},
		'Email': {'PR': 'Primary Email1', 'RY': 'contact_email_1'},
		'Lender': {'PR': '1st Orig Lender', 'RY': 'mortgage_recorded_name'},
		'Loan Amount': {'PR': '1st Amount', 'RY': 'mortgage_amount'},
		'Loan Date': {'PR': '1st Rec Date', 'RY': 'mortgage_recorded_date'},
		'Loan Purpose': {'PR': '1st Purpose', 'RY': 'None'},
		'Prop Type': {'PR': 'Type', 'RY': 'property_type'},
		'Prop Subtype': {'PR': 'None', 'RY': 'property_subtype'},
		'Loan Maturity Date': {'PR': 'None', 'RY': 'mortgage_maturity_date'},
		'Last Sale Date': {'PR': 'Purchase Date', 'RY': 'sale_recorded_date'},
		'Last Sale Amount': {'PR': 'Purchase Amt', 'RY': 'sale_amount'},
		'APN': {'PR': 'APN', 'RY': 'apn'},
		'Source': {'PR': 'PropertyRadar', 'RY': 'Reonomy'},
		'Source ID': {'PR': 'Radar ID', 'RY': 'reonomy_id'},
		'Latitude': {'PR': 'Latitude', 'RY': 'latitude'},
		'Longitude': {'PR': 'Longitude', 'RY': 'longitude'}
	}

		# To rename PropertyRadar FCL Doc Type field
	dFCL_doc_type = {
		'COP': 'Certificate Of Participation',
		'FJF': 'Final Judgment of Foreclosure',
		'FJG': 'Final Judgment',
		'LIS': 'Lis Pendens',
		'NOD': 'Notice Of Default',
		'NSL': 'Notice Of Sale',
		'NTS': 'Notice Of Trustee Sale',
		'SBT': 'Appointment Of Substitute Trustee'
	}

	# To populate Reonomy FCL Stage field
	dFCL_stage = {
		'APPOINTMENT OF SUBSTITUTE TRUSTEE': 'Preforeclosure',
		'FINAL JUDGMENT': 'Preforeclosure',
		'LIS PENDENS': 'Preforeclosure',
		'NOTICE OF DEFAULT': 'Preforeclosure',
		'NOTICE OF FORECLOSURE': 'Bank Owned',
		'NOTICE OF SALE': 'Auction',
		"NOTICE OF TRUSTEE'S SALE": 'Auction',
		'RELEASE OF PRE-FORECLOSURE': 'Not In Foreclosure',
		'UNKNOWN': 'Preforeclosure'	
	}
	return dFCL_fields, dFCL_doc_type, dFCL_stage

# Get Buyer Seller dict
def get_blank_buyer_seller_dict():
	dBS = {
		'BUYERENTITYID': 'None',
		'BUYERENTITYNAME':  'None',
		'BUYERPERSONID': 'None',
		'BUYERPERSONNAME': 'None',
		'SELLERENTITYID': 'None',
		'SELLERENTITYNAME': 'None',
		'SELLERPERSPNID': 'None',
		'SELLERPERSONNAME': 'None'
	}
	return dBS

# Make Dictionary for updating Deal from Lead to Closed Lost
def get_lead_to_closed_lost_dict(DID):
	# Brokerage: 012a0000001ZSS5AAO Research: 012a0000001ZSS8AAO
	dUP = {
		'Id': DID,
		'type': 'lda_Opportunity__c',
		'RecordTypeID': '012a0000001ZSS8AAO',
		'StageName__c': 'Closed Lost'
	}
	return dUP

# Get the dictionary of M1 parameters
def get_blank_m1_params_dict():
	# Assign user and initials and dateupdated
	user, initials = lao.getUserName(initials=True)
	dateupdated = td.date_engine(date='today', outformat='slash', informat='unknown')

	m1_params = {
		"action": "None",
		"county": "None",
		"dateupdated": dateupdated,
		"fieldName": "None",
		"initials": initials,
		"layername": "None",
		"lat": "None",
		"lid": "None",
		"lon": "None",
		"oid": "None",
		"parcels": "None",
		"pid": "None",
		"pidnew": "None",
		"propertyid": "None",
		"user": user
	}
	return m1_params

# Get the dictionary for M1_Deal_Updater owner values
def get_m1_deal_updater_owner_dict():
	dOS_val = {
		'email': 'None',
		'entity': 'None',
		'person': 'None',
		'phone': 'None',
		'street': 'None'
	}
	return dOS_val

# Get blank Deal Update dict
def get_blank_deal_update_dict(DID='None'):
	user = lao.getUserName()
	ownerid = lao.getUser_tfUserID()
	today_date = td.today_date()
	dUpdate = {
		'OwnerId': ownerid,
		'type': 'lda_Opportunity__c',
		'Verified__c': True,
	}
	if DID == 'None':
		dUpdate['Id'] = 'None'
	else:
		dUpdate['Id'] = DID

	return dUpdate

# Get TF Record Type
def get_record_type(recordid):
	# Brokerage: 012a0000001ZSS5AAO Research: 012a0000001ZSS8AAO
	if recordid == '012a0000001ZSS5AAO':
		return 'Brokerage'
	elif recordid == '012a0000001ZSS8AAO':
		return 'Research'

# Build MC Create/Update Dicts from dAcc and dTags
def get_mailchimp_add_update_dict(dAcc):
	# Input:	dAcc (Account Dict)
	# Return:	dMC (MailChimp create/update dict)
	#			dTags (Maichimp add/update Tags dict)
	
	dTags = 'None'

	# Build MailChimp dict
	dMC = {
		"status": dAcc['MCSTATUS'],
		"email_address": dAcc["EMAIL"]}

	# Only return basic dMC if contact is to be unsubscribed
	if dAcc['MCSTATUS'] == 'unsubscribed':
		return dMC, dTags
	
	dMC["merge_fields"] = {
			"FNAME": dAcc["NF"],
			"LNAME": dAcc["NL"]}

	# Add Entity/Company if populated in dAcc
	if dAcc["ENTITY"] != "" or dAcc["ENTITY"] != "None":
		dMC["merge_fields"]["COMPANY"] = dAcc["ENTITY"]

	# Add Interests/Campaign Groups
	# MCLCAMPAIGNIDS is a dict of Groups and IDs
	if dAcc["MCLCAMPAIGNIDS"] != "None":
		dMC["interests"] = {}
		for capaignID in dAcc["MCLCAMPAIGNIDS"]:
			dMC["interests"][capaignID] = True

	# Add Tags (TF Categories)
	dTags = 'None'
	if dAcc["CATEGORY"] != "None":
		dTags = {"tags": []}
		lCat = dAcc['CATEGORY'].split(';')
		for row in lCat:
			# dMC["tags"].append(row)
			dTags['tags'].append({"name": row, "status": "active"})

	return dMC, dTags

# Make dict of Markets/Advisors
def get_mm_market_dict():
	dMMMkt = {'ATL': 'Atlanta',
				'AUS': 'Austin',
				'BOI': 'Boise',
				'CLT': 'Charlotte',
				'DEN': 'Denver',
				'DFW': 'Dallas-Ft Worth',
				'GSP': 'Greenville',
				'HOU': 'Houston',
				'JAX': 'Jacksonville',
				'KCI': 'Kansas City',
				'LVS': 'Las Vegas',
				'NAZ': 'Northern Arizona',
				'NSH': 'Nashville',
				'ORL': 'Orlando',
				'PHX': 'Phoenix',
				'SLC': 'Salt Lake City',
				'TPA': 'Tampa Bay',
				'TUC': 'Tucson',
				'CAM': 'Campbell',
				'HEG': 'Heglie',
				'MCC': 'McCarville',
				'MED': 'Medical',
				'RNO': 'Reno',
				'SCH': 'Schwab',
				'VOG': 'Vogel',
				'WUE': 'Wuertz',
				'XAN': 'Xander'}
	return dMMMkt

# Return dict of Fields for CSV to TF spreadsheet
def get_tf_csv_dict():
	d = {'Seller Entity': 'None',
		'Seller Person': 'None',
		'Seller Street': 'None',
		'Seller City': 'None',
		'Seller State': 'None',
		'Seller Zip': 'None',
		'Seller Country': 'None',
		'Seller Phone': 'None',
		'Seller Email': 'None',
		'Buyer Entity': 'None',
		'Buyer Person': 'None',
		'Buyer Street': 'None',
		'Buyer City': 'None',
		'Buyer State': 'None',
		'Buyer Zip': 'None',
		'Buyer Country': 'None',
		'Buyer Phone': 'None',
		'Buyer Email': 'None',
		'Acres__c': 'None',
		'Buyer_Acting_As__c': 'None',
		'City__c': 'None',
		'Classification__c': 'None',
		'Country__c': 'None',
		'County__c': 'None',
		'Description__c': 'None',
		'General_Plan__c': 'None',
		'Keyword_Group__c': 'None',
		'Latitude__c': 'None',
		'Lead_Parcel__c': 'None',
		'Legal_Description__c': 'None',
		'Location__c': 'None',
		'Longitude__c': 'None',
		'Lot_Description__c': 'None',
		'Lot_Type__c': 'None',
		'Lots__c': 'None',
		'Market__c': 'None',
		'Deal Name': 'None',
		'Parcels__c': 'None',
		'Recorded_Instrument_Number__c': 'None',
		'Sale_Date__c': 'None',
		'Sale_Price__c': 'None',
		'Source__c': 'None',
		'Source_ID__c': 'None',
		'State__c': 'None',
		'Subdivision__c': 'None',
		'Submarket__c': 'None',
		'Zipcode__c': 'None',
		'Zoning__c': 'None',
		'List_Date__c': 'None',
		'List_Price__c': 'None',
		'List Entity': 'None',
		'List Agent': 'None',
		'List Agent Phone': 'None',
		'List Agent Email': 'None',
		'AltAPN A': 'None',
		'AltAPN B': 'None',
		'PID__c': 'None',
		'Residence Y-N': 'None',
		'Terms__c': 'None',
		'Recorded_Doc_URL__c': 'None',
		'Notes': 'None',
		'Lender': 'None',
		'Loan_Amount__c': 'None',
		'Loan_Date__c': 'None',
		'Buyer Agent Entity': 'None',
		'Buyer Agent': 'None',
		'Buyer Agent Phone': 'None',
		'Buyer Agent Email': 'None',
		'MLS Status': 'None'}
	
	return d

# get the propertyradar to tf csv field mapping dictionary
def get_proprad_to_tf_csv_field_map_dict():

	d = {
	'AdvancedPropertyType': 'Classification__c',
	'APN': 'Parcels__c',
	'City': 'City__c',
	'LastTransferSeller': 'Seller Entity',
	'County': 'County__c',
	'FirstLenderOriginal': 'Lender',
	'FirstAmount': 'Loan_Amount__c',
	'FirstDate': 'Loan_Date__c',
	'LastTransferRecDate': 'Sale_Date__c',
	'LastTransferValue': 'Sale_Price__c',
	'Latitude': 'Latitude__c',
	'Longitude': 'Longitude__c',
	'LotSizeAcres': 'Acres__c',
	'Owner': 'Buyer Entity',
	'OwnerFirstName': 'Buyer First Name',
	'OwnerLastName': 'Buyer Last Name',
	'OwnerAddress': 'Buyer Street',
	'OwnerCity': 'Buyer City',
	'OwnerState': 'Buyer State',
	'OwnerZipFive': 'Buyer Zip',
	'RadarID': 'Source_ID__c',
	'State': 'State__c',
	'Subdivision': 'Subdivision__c',
	'ZipFive': 'Zipcode__c',
	'Zoning': 'Zoning__c'
	}
	return d

# Get the Red News to tf csv field mapping dictionary
def get_red_news_to_tf_csv_field_map_dict():

	d = {
		'Seller Entity': 'Seller Entity',
		'Seller Person': 'Seller Person',
		'Seller Street': 'Seller Street',
		'Seller City': 'Seller City',
		'Seller State': 'Seller State',
		'Seller Zip': 'Seller Zip',
		'Seller Phone': 'Seller Phone',
		'Buyer Entity': 'Buyer Entity',
		'Buyer Person': 'Buyer Person',
		'Buyer Street': 'Buyer Street',
		'Buyer City': 'Buyer City',
		'Buyer State': 'Buyer State',
		'Buyer Zip': 'Buyer Zip',
		'Buyer Phone': 'Buyer Phone',
		'Sale Date': 'Sale_Date__c',
		'Type': 'Description__c',
		'Lead Parcel': 'Lead_Parcel__c',
		'Sale Price': 'Sale_Price__c',
		'Zoning': 'Zoning__c',
		'Longitude': 'Longitude__c',
		'Parcels': 'Parcels__c',
		'Rec Doc': 'Recorded_Instrument_Number__c',
		'Latitude': 'Latitude__c',
		'City': 'City__c',
		'Subdivision': 'Subdivision__c',
		'Source ID': 'Source_ID__c',
		'Acres': 'Acres__c',
		'Lots': 'Lots__c',
		'Source': 'Red News',
		'State': 'State__c',
		'County': 'County__c'
	}
	return d

def get_state_2_digit_fips_dict():
	d = 	d = {
		'Alabama': '01',
		'Alaska': '02',
		'Arizona': '04',
		'Arkansas': '05',
		'California': '06',
		'Colorado': '08',
		'Connecticut': '09',
		'Delaware': '10',
		'District of Columbia': '11',
		'Florida': '12',
		'Georgia': '13',
		'Hawaii': '15',
		'Idaho': '16',
		'Illinois': '17',
		'Indiana': '18',
		'Iowa': '19',
		'Kansas': '20',
		'Kentucky': '21',
		'Louisiana': '22',
		'Maine': '23',
		'Maryland': '24',
		'Massachusetts': '25',
		'Michigan': '26',
		'Minnesota': '27',
		'Mississippi': '28',
		'Missouri': '29',
		'Montana': '30',
		'Nebraska': '31',
		'Nevada': '32',
		'New Hampshire': '33',
		'New Jersey': '34',
		'New Mexico': '35',
		'New York': '36',
		'North Carolina': '37',
		'North Dakota': '38',
		'Ohio': '39',
		'Oklahoma': '40',
		'Oregon': '41',
		'Pennsylvania': '42',
		'Puerto Rico': '72',
		'Rhode Island': '44',
		'South Carolina': '45',
		'South Dakota': '46',
		'Tennessee': '47',
		'Texas': '48',
		'Utah': '49',
		'Vermont': '50',
		'Virgin Islands': '78',
		'Virginia': '51',
		'Washington': '53',
		'West Virginia': '54',
		'Wisconsin': '55',
		'Wyoming': '56'}
	return d

# Get State and CBSA FIPS codes dictionary
def get_state_cbsa_fips_dict():
	d = {
		'Abilene': '4810180',
		'Aguadilla': '7210380',
		'Akron': '3910420',
		'Alabama': '0100000',
		'Alaska': '0200000',
		'Albany': '4110540',
		'Albany-Schenectady-Troy': '3610580',
		'Albuquerque': '3510740',
		'Alexandria': '2210780',
		'All Metropolitan Statistical Areas': '0099999',
		'Allentown-Bethlehem-Easton': '4210900',
		'Altoona': '4211020',
		'Amarillo': '4811100',
		'Ames': '1911180',
		'Amherst Town-Northampton': '2511200',
		'Anaheim-Santa Ana-Irvine  Metropolitan Division  Metropolitan Division': '0611244',
		'Anchorage': '0211260',
		'Ann Arbor': '2611460',
		'Anniston-Oxford': '0111500',
		'Appleton': '5511540',
		'Arecibo': '7211640',
		'Arizona': '0400000',
		'Arkansas': '0500000',
		'Arlington-Alexandria-Reston Metropolitan Division': '5111694',
		'Asheville': '3711700',
		'Athens-Clarke County': '1312020',
		'Atlanta-Sandy Springs-Roswell': '1312060',
		'Atlanta-Sandy Springs-Roswell Metropolitan Division': '1312054',
		'Atlantic City-Hammonton': '3412100',
		'Auburn-Opelika': '0112220',
		'Augusta-Richmond County': '1312260',
		'Austin-Round Rock-San Marcos': '4812420',
		'Bakersfield-Delano': '0612540',
		'Baltimore-Columbia-Towson': '2412580',
		'Bangor': '2312620',
		'Barnstable Town': '2512700',
		'Baton Rouge': '2212940',
		'Battle Creek': '2612980',
		'Bay City': '2613020',
		'Beaumont-Port Arthur': '4813140',
		'Beckley': '5413220',
		'Bellingham': '5313380',
		'Bend': '4113460',
		'Billings': '3013740',
		'Binghamton': '3613780',
		'Birmingham': '0113820',
		'Bismarck': '3813900',
		'Blacksburg-Christiansburg-Radford': '5113980',
		'Bloomington': '1814020',
		'Boise City': '1614260',
		'Boston Metropolitan Division': '2514454',
		'Boston-Cambridge-Newton': '2514460',
		'Boulder': '0814500',
		'Bowling Green': '2114540',
		'Bozeman': '3014580',
		'Bremerton-Silverdale-Port Orchard': '5314740',
		'Bridgeport-Stamford-Danbury': '0914860',
		'Brownsville-Harlingen': '4815180',
		'Brunswick-St. Simons': '1315260',
		'Buffalo-Cheektowaga': '3615380',
		'Burlington': '3715500',
		'Burlington-South Burlington': '5015540',
		'California': '0600000',
		'Cambridge-Newton-Framingham Metropolitan Division': '2515764',
		'Camden Metropolitan Division': '3415804',
		'Canton-Massillon': '3915940',
		'Cape Coral-Fort Myers': '1215980',
		'Cape Girardeau': '2916020',
		'Carson City': '3216180',
		'Casper': '5616220',
		'Cedar Rapids': '1916300',
		'Chambersburg': '4216540',
		'Champaign-Urbana': '1716580',
		'Charleston': '5416620',
		'Charleston-North Charleston': '4516700',
		'Charlotte-Concord-Gastonia': '3716740',
		'Charlottesville': '5116820',
		'Chattanooga': '4716860',
		'Cheyenne': '5616940',
		'Chicago-Naperville-Elgin': '1716980',
		'Chicago-Naperville-Schaumburg Metropolitan Division': '1716984',
		'Chico': '0617020',
		'Cincinnati': '3917140',
		'Clarksville': '4717300',
		'Cleveland': '4717420',
		"Coeur d'Alene": '1617660',
		'College Station-Bryan': '4817780',
		'Colorado': '0800000',
		'Colorado Springs': '0817820',
		'Columbia': '4517900',
		'Columbus': '3918140',
		'Connecticut': '0900000',
		'Corpus Christi': '4818580',
		'Corvallis': '4118700',
		'Crestview-Fort Walton Beach-Destin': '1218880',
		'Dallas-Fort Worth-Arlington': '4819100',
		'Dallas-Plano-Irving Metropolitan Division': '4819124',
		'Dalton': '1319140',
		'Daphne-Fairhope-Foley': '0119300',
		'Davenport-Moline-Rock Island': '1919340',
		'Dayton-Kettering-Beavercreek': '3919430',
		'Decatur': '1719500',
		'Delaware': '1000000',
		'Deltona-Daytona Beach-Ormond Beach': '1219660',
		'Denver-Aurora-Centennial': '0819740',
		'Des Moines-West Des Moines': '1919780',
		'Detroit-Dearborn-Livonia Metropolitan Division': '2619804',
		'Detroit-Warren-Dearborn': '2619820',
		'District of Columbia': '1100000',
		'Dothan': '0120020',
		'Dover': '1020100',
		'Dubuque': '1920220',
		'Duluth': '2720260',
		'Durham-Chapel Hill': '3720500',
		'Eagle Pass': '4820580',
		'Eau Claire': '5520740',
		'El Centro': '0620940',
		'El Paso': '4821340',
		'Elgin Metropolitan Division': '1720994',
		'Elizabethtown': '2121060',
		'Elkhart-Goshen': '1821140',
		'Elmira': '3621300',
		'Enid': '4021420',
		'Erie': '4221500',
		'Eugene-Springfield': '4121660',
		'Evansville': '1821780',
		'Everett Metropolitan Division': '5321794',
		'Fairbanks-College': '0221820',
		'Fargo': '3822020',
		'Farmington': '3522140',
		'Fayetteville': '3722180',
		'Fayetteville-Springdale-Rogers': '0522220',
		'Flagstaff': '0422380',
		'Flint': '2622420',
		'Florence': '4522500',
		'Florence-Muscle Shoals': '0122520',
		'Florida': '1200000',
		'Fond du Lac': '5522540',
		'Fort Collins-Loveland': '0822660',
		'Fort Lauderdale-Pompano Beach-Sunrise Metropolitan Division': '1222744',
		'Fort Smith': '0522900',
		'Fort Wayne': '1823060',
		'Fort Worth-Arlington-Grapevine Metropolitan Division Worth-Arlington-Grapevine Metropolitan Division': '4823104',
		'Frederick-Gaithersburg-Bethesda Metropolitan Division': '2423224',
		'Fresno': '0623420',
		'Gadsden': '0123460',
		'Gainesville': '1323580',
		'Georgia': '1300000',
		'Gettysburg': '4223900',
		'Glens Falls': '3624020',
		'Goldsboro': '3724140',
		'Grand Forks': '3824220',
		'Grand Island': '3124260',
		'Grand Junction': '0824300',
		'Grand Rapids-Wyoming-Kentwood': '2624340',
		'Grants Pass': '4124420',
		'Great Falls': '3024500',
		'Greeley': '0824540',
		'Green Bay': '5524580',
		'Greensboro-High Point': '3724660',
		'Greenville': '3724780',
		'Greenville-Anderson-Greer': '4524860',
		'Guayama': '7225020',
		'Gulfport-Biloxi': '2825060',
		'Hagerstown-Martinsburg': '2425180',
		'Hammond': '2225220',
		'Hanford-Corcoran': '0625260',
		'Harrisburg-Carlisle': '4225420',
		'Harrisonburg': '5125500',
		'Hartford-West Hartford-East Hartford': '0925540',
		'Hattiesburg': '2825620',
		'Hawaii': '1500000',
		'Helena': '3025740',
		'Hickory-Lenoir-Morganton': '3725860',
		'Hilton Head Island-Bluffton-Beaufort': '4525940',
		'Hinesville': '1325980',
		'Homosassa Springs': '1226140',
		'Hot Springs': '0526300',
		'Houma-Bayou Cane-Thibodaux': '2226380',
		'Houston-Pasadena-The Woodlands': '4826420',
		'Huntington-Ashland': '5426580',
		'Huntsville': '0126620',
		'Idaho': '1600000',
		'Idaho Falls': '1626820',
		'Illinois': '1700000',
		'Indiana': '1800000',
		'Indianapolis-Carmel-Greenwood': '1826900',
		'Iowa': '1900000',
		'Iowa City': '1926980',
		'Ithaca': '3627060',
		'Jackson': '4727180',
		'Jacksonville': '1227260',
		'Janesville-Beloit': '5527500',
		'Jefferson City': '2927620',
		'Johnson City': '4727740',
		'Johnstown': '4227780',
		'Jonesboro': '0527860',
		'Joplin': '2927900',
		'Kahului-Wailuku': '1527980',
		'Kalamazoo-Portage': '2628020',
		'Kankakee': '1728100',
		'Kansas': '2000000',
		'Kansas City': '2928140',
		'Kennewick-Richland': '5328420',
		'Kenosha': '5528450',
		'Kentucky': '2100000',
		'Killeen-Temple': '4828660',
		'Kingsport-Bristol': '4728700',
		'Kingston': '3628740',
		'Kiryas Joel-Poughkeepsie-Newburgh': '3628880',
		'Knoxville': '4728940',
		'Kokomo': '1829020',
		'La Crosse-Onalaska': '5529100',
		'Lafayette': '2229180',
		'Lafayette-West Lafayette': '1829200',
		'Lake Charles': '2229340',
		'Lake County Metropolitan Division County Metropolitan Division': '1729404',
		'Lake County-Porter County-Jasper County Metropolitan Division': '1829414',
		'Lake Havasu City-Kingman': '0429420',
		'Lakeland-Winter Haven': '1229460',
		'Lakewood-New Brunswick Metropolitan Division Brunswick Metropolitan Division': '3429484',
		'Lancaster': '4229540',
		'Lansing-East Lansing': '2629620',
		'Laredo': '4829700',
		'Las Cruces': '3529740',
		'Las Vegas-Henderson-North Las Vegas': '3229820',
		'Lawrence': '2029940',
		'Lawton': '4030020',
		'Lebanon': '4230140',
		'Lewiston': '1630300',
		'Lewiston-Auburn': '2330340',
		'Lexington Park': '2430500',
		'Lexington-Fayette': '2130460',
		'Lima': '3930620',
		'Lincoln': '3130700',
		'Little Rock-North Little Rock-Conway': '0530780',
		'Logan': '4930860',
		'Longview': '4830980',
		'Longview-Kelso': '5331020',
		'Los Angeles-Long Beach-Anaheim': '0631080',
		'Los Angeles-Long Beach-Glendale Metropolitan Division': '0631084',
		'Louisiana': '2200000',
		'Louisville/Jefferson County': '2131140',
		'Lubbock': '4831180',
		'Lynchburg': '5131340',
		'Macon-Bibb County': '1331420',
		'Madison': '5531540',
		'Maine': '2300000',
		'Manchester-Nashua': '3331700',
		'Manhattan': '2031740',
		'Mankato': '2731860',
		'Mansfield': '3931900',
		'Marietta Metropolitan Division': '1331924',
		'Maryland': '2400000',
		'Massachusetts': '2500000',
		'Mayaguez': '7232420',
		'McAllen-Edinburg-Mission': '4832580',
		'Medford': '4132780',
		'Memphis': '4732820',
		'Merced': '0632900',
		'Miami-Fort Lauderdale-West Palm Beach': '1233100',
		'Miami-Miami Beach-Kendall Metropolitan Division Beach-Kendall Metropolitan Division': '1233124',
		'Michigan': '2600000',
		'Michigan City-La Porte': '1833140',
		'Midland': '4833260',
		'Milwaukee-Waukesha': '5533340',
		'Minneapolis-St. Paul-Bloomington': '2733460',
		'Minnesota': '2700000',
		'Minot': '3833500',
		'Mississippi': '2800000',
		'Missoula': '3033540',
		'Missouri': '2900000',
		'Mobile': '0133660',
		'Modesto': '0633700',
		'Monroe': '2633780',
		'Montana': '3000000',
		'Montgomery': '0133860',
		'Montgomery County-Bucks County-Chester County Metropolitan Division': '4233874',
		'Morgantown': '5434060',
		'Morristown': '4734100',
		'Mount Vernon-Anacortes': '5334580',
		'Muncie': '1834620',
		'Muskegon-Norton Shores': '2634740',
		'Myrtle Beach-Conway-North Myrtle Beach': '4534820',
		'Napa': '0634900',
		'Naples-Marco Island': '1234940',
		'Nashville-Davidson--Murfreesboro--Franklin': '4734980',
		'Nassau County-Suffolk County Metropolitan Division': '3635004',
		'Nebraska': '3100000',
		'Nevada': '3200000',
		'New Hampshire': '3300000',
		'New Haven': '0935300',
		'New Jersey': '3400000',
		'New Mexico': '3500000',
		'New Orleans-Metairie': '2235380',
		'New York': '3600000',
		'New York City': '3693561',
		'New York-Jersey City-White Plains Metropolitan Division': '3635614',
		'New York-Newark-Jersey City': '3635620',
		'Newark Metropolitan Division': '3435084',
		'Niles': '2635660',
		'North Carolina': '3700000',
		'North Dakota': '3800000',
		'North Port-Bradenton-Sarasota': '1235840',
		'Norwich-New London-Willimantic': '0935980',
		'Oakland-Fremont-Berkeley Metropolitan Division': '0636084',
		'Ocala': '1236100',
		'Odessa': '4836220',
		'Ogden': '4936260',
		'Ohio': '3900000',
		'Oklahoma': '4000000',
		'Oklahoma City': '4036420',
		'Olympia-Lacey-Tumwater': '5336500',
		'Omaha': '3136540',
		'Oregon': '4100000',
		'Orlando-Kissimmee-Sanford': '1236740',
		'Oshkosh-Neenah': '5536780',
		'Owensboro': '2136980',
		'Oxnard-Thousand Oaks-Ventura': '0637100',
		'Paducah': '2137140',
		'Palm Bay-Melbourne-Titusville': '1237340',
		'Panama City-Panama City Beach City': '1237460',
		'Parkersburg-Vienna': '5437620',
		'Pennsylvania': '4200000',
		'Pensacola-Ferry Pass-Brent': '1237860',
		'Peoria': '1737900',
		'Philadelphia Metropolitan Division': '4237964',
		'Philadelphia-Camden-Wilmington': '4237980',
		'Phoenix-Mesa-Chandler': '0438060',
		'Pinehurst-Southern Pines': '3738240',
		'Pittsburgh': '4238300',
		'Pittsfield': '2538340',
		'Pocatello': '1638540',
		'Ponce': '7238660',
		'Port St. Lucie': '1238940',
		'Portland-South Portland': '2338860',
		'Portland-Vancouver-Hillsboro': '4138900',
		'Prescott Valley-Prescott': '0439150',
		'Providence-Warwick': '4439300',
		'Provo-Orem-Lehi': '4939340',
		'Pueblo': '0839380',
		'Puerto Rico': '7200000',
		'Punta Gorda': '1239460',
		'Racine-Mount Pleasant': '5539540',
		'Raleigh-Cary': '3739580',
		'Rapid City': '4639660',
		'Reading': '4239740',
		'Redding': '0639820',
		'Reno': '3239900',
		'Rhode Island': '4400000',
		'Richmond': '5140060',
		'Riverside-San Bernardino-Ontario': '0640140',
		'Roanoke': '5140220',
		'Rochester': '3640380',
		'Rockford': '1740420',
		'Rockingham County-Strafford County Metropolitan Division': '3340484',
		'Rocky Mount': '3740580',
		'Rome': '1340660',
		'Sacramento-Roseville-Folsom': '0640900',
		'Saginaw': '2640980',
		'Salem': '4141420',
		'Salinas': '0641500',
		'Salisbury': '2441540',
		'Salt Lake City-Murray': '4941620',
		'San Angelo': '4841660',
		'San Antonio-New Braunfels': '4841700',
		'San Diego-Chula Vista-Carlsbad': '0641740',
		'San Francisco-Oakland-Fremont': '0641860',
		'San Francisco-San Mateo-Redwood City Metropolitan Division': '0641884',
		'San Jose-Sunnyvale-Santa Clara': '0641940',
		'San Juan-Bayamon-Caguas': '7241980',
		'San Luis Obispo-Paso Robles': '0642020',
		'San Rafael Metropolitan Division Rafael Metropolitan Division': '0642034',
		'Sandusky': '3941780',
		'Santa Cruz-Watsonville': '0642100',
		'Santa Fe': '3542140',
		'Santa Maria-Santa Barbara': '0642200',
		'Santa Rosa-Petaluma': '0642220',
		'Savannah': '1342340',
		'Scranton--Wilkes-Barre': '4242540',
		'Seattle-Bellevue-Kent Metropolitan Division': '5342644',
		'Seattle-Tacoma-Bellevue': '5342660',
		'Sebastian-Vero Beach': '1242680',
		'Sebring': '1242700',
		'Sheboygan': '5543100',
		'Sherman-Denison': '4843300',
		'Shreveport-Bossier City': '2243340',
		'Sierra Vista-Douglas': '0443420',
		'Sioux City': '1943580',
		'Sioux Falls': '4643620',
		'Slidell-Mandeville-Covington': '2243640',
		'South Bend-Mishawaka': '1843780',
		'South Carolina': '4500000',
		'South Dakota': '4600000',
		'Spartanburg': '4543900',
		'Spokane-Spokane Valley': '5344060',
		'Springfield': '3944220',
		'St. Cloud': '2741060',
		'St. George': '4941100',
		'St. Joseph': '2941140',
		'St. Louis': '2941180',
		'St. Petersburg-Clearwater-Largo Metropolitan Division Petersburg-Clearwater-Largo Metropolitan Division': '1241304',
		'State College': '4244300',
		'Staunton-Stuarts Draft': '5144420',
		'Stockton-Lodi': '0644700',
		'Sumter': '4544940',
		'Syracuse': '3645060',
		'Tacoma-Lakewood Metropolitan Division': '5345104',
		'Tallahassee': '1245220',
		'Tampa Metropolitan Division': '1245294',
		'Tampa-St. Petersburg-Clearwater': '1245300',
		'Tennessee': '4700000',
		'Terre Haute': '1845460',
		'Texarkana': '4845500',
		'Texas': '4800000',
		'Toledo': '3945780',
		'Topeka': '2045820',
		'Traverse City': '2645900',
		'Trenton-Princeton': '3445940',
		'Tucson': '0446060',
		'Tulsa': '4046140',
		'Tuscaloosa': '0146220',
		'Twin Falls': '1646300',
		'Tyler': '4846340',
		'Urban Honolulu': '1546520',
		'Utah': '4900000',
		'Utica-Rome': '3646540',
		'Valdosta': '1346660',
		'Vallejo': '0646700',
		'Vermont': '5000000',
		'Victoria': '4847020',
		'Vineland': '3447220',
		'Virgin Islands': '7800000',
		'Virginia': '5100000',
		'Virginia Beach-Chesapeake-Norfolk': '5147260',
		'Visalia': '0647300',
		'Waco': '4847380',
		'Walla Walla': '5347460',
		'Warner Robins': '1347580',
		'Warren-Troy-Farmington Hills Metropolitan Division Hills Metropolitan Division': '2647664',
		'Washington': '5300000',
		'Washington Metropolitan Division': '1147764',
		'Washington-Arlington-Alexandria': '1147900',
		'Waterbury-Shelton': '0947930',
		'Waterloo-Cedar Falls': '1947940',
		'Watertown-Fort Drum': '3648060',
		'Wausau': '5548140',
		'Weirton-Steubenville': '5448260',
		'Wenatchee-East Wenatchee': '5348300',
		'West Palm Beach-Boca Raton-Delray Beach': '1248424',
		'West Virginia': '5400000',
		'Wheeling': '5448540',
		'Wichita': '2048620',
		'Wichita Falls': '4848660',
		'Wildwood-The Villages': '1248680',
		'Williamsport': '4248700',
		'Wilmington': '3748900',
		'Wilmington Metropolitan Division': '1048864',
		'Winchester': '5149020',
		'Winston-Salem': '3749180',
		'Wisconsin': '5500000',
		'Worcester': '2549340',
		'Wyoming': '5600000',
		'Yakima': '5349420',
		'York-Hanover': '4249620',
		'Youngstown-Warren': '3949660',
		'Yuba City': '0649700',
		'Yuma': '0449740'}
		
	return d

def get_L5_url_dict():
	dL5_URL_Vals = {
	'Albuquerque': {'Name': 'Abq', 'Zoom': '8'},
	'Atlanta': {'Name': 'Atlanta', 'Zoom': '15'},
	'Austin': {'Name': 'Austin', 'Zoom': '8'},
	'Boise': {'Name': 'Boise', 'Zoom': '8'},
	'Charlotte': {'Name': 'charlotte', 'Zoom': '9'},
	'Denver': {'Name': 'Denver', 'Zoom': '23'},
	'Dallas-Ft Worth': {'Name': 'DFW', 'Zoom': '11'},
	'DFW': {'Name': 'DFW', 'Zoom': '11'},
	'Greenville': {'Name': 'greenville', 'Zoom': '9'},
	'Houston': {'Name': 'Houston', 'Zoom': '25'},
	'Huntsville': {'Name': 'Nashville', 'Zoom': '20'},
	'Jacksonville': {'Name': 'Jacksonville', 'Zoom': '8'},
	'Kansas City': {'Name': 'KansasCity', 'Zoom': '6'},
	'Las Vegas': {'Name': 'LasVegas', 'Zoom': '9'},
	'Nashville': {'Name': 'Nashville', 'Zoom': '20'},
	'Orlando': {'Name': 'Orlando', 'Zoom': '10'},
	'Scottsdale': {'Name': 'phoenix', 'Zoom': '24'},
	'Phoenix': {'Name': 'phoenix', 'Zoom': '24'},
	'Prescott': {'Name': 'Prescott', 'Zoom': '26'},
	'Northern Arizona': {'Name': 'Prescott', 'Zoom': '26'},
	'Raleigh': {'Name': 'Raleigh', 'Zoom': '8'},
	'Reno': {'Name': 'Reno', 'Zoom': '23'},
	'Salt Lake City': {'Name': 'Utah', 'Zoom': '10'},
	'Tampa': {'Name': 'Tampa', 'Zoom': '11'},
	'Tucson': {'Name': 'Tucson', 'Zoom': '22'},
	'Yuma': {'Name': 'phoenix', 'Zoom': '24'},
	'P&Z': {'Name': 'phoenix', 'Zoom': '11'}
	}
	return dL5_URL_Vals

# LISTS #####################################################################

# Get TF Classification list
def get_classification_list():
	lList = [
		'blank',
		'Residential',
		'Commercial',
		'Industrial',
		'Office',
		'Retail',
		'Agricultural',
		'Apartment Horizontal',
		'Apartment Traditional',
		'Build for Rent (platted)',
		'Church',
		'Conservation',
		'Data Center',
		'High Density Assisted Living',
		'High Density Residential',
		'Hospitality',
		'Manufactured Home',
		'Master Planned Community',
		'Medical',
		'Mixed-Use',
		'Not for Sale',
		'Other',
		'Public',
		'Ranch and Recreation',
		'Residence Included',
		'Resort or Golf',
		'Sand and Gravel',
		'School',
		'School Charter',
		'Solar Wind',
		'Speculative',
		'Storage',
		'Timber'
	]
	return lList

# Get the county FIPS codes from the LAO Counties spreadsheet given a MarketAbb.
def get_county_fips_codes(market_abb):
	"""
	Given a MarketAbb from the LAO Counties spreadsheet, returns the State 
	and all FIPS codes for that market as a comma-separated string.
	
	Args:
		market_abb (str): The MarketAbb to look up (e.g., 'ATL', 'ABQ', 'DEN')
		excel_file_path (str): Path to the Excel file (default: 'LAO_Counties.xlsx')
	
	Returns:
		tuple: (state, fips_codes_string) where:
			- state (str): The state abbreviation for the market
			- 		fips_codes_string (str): Comma-separated FIPS codes with spaces and double quotes
	
	Raises:
		ValueError: If the MarketAbb is not found in the data
	"""
	county_file = 'F:/Research Department/Code/Databases/LAO_Counties.xlsx'

	# Read the Excel file
	df = pd.read_excel(county_file, sheet_name='LAO_Counties')
	
	# Filter data for the specified MarketAbb
	market_data = df[df['MarketAbb'] == market_abb]
	
	if market_data.empty:
		raise ValueError(f"MarketAbb '{market_abb}' not found in the data")
	
	# Get the state (should be the same for all rows in a market)
	state = market_data['State'].iloc[0]
	
	# Get FIPS codes, excluding 'NA' values
	fips_codes = market_data['FIPS'].tolist()
	lCounty_fips = [f'{int(fips):05d}' for fips in fips_codes if fips != 'NA' and pd.notna(fips)]
	
	# # Create comma-separated string with no spaces
	# fips_codes_string = ','.join(valid_fips)
	
	return state, lCounty_fips

# Get TF Lot Type list
def get_lot_type_list():
	lList = [
		'blank',
		'Covered Land',
		'Finished Lots',
		'Initial Lot Option',
		'Lot Option Built Out',
		'Partially Improved',
		'Platted and Engineered',
		'Raw Acreage'
	]
	return lList

# Get list of Classifications that may not be Raw Acreage
def get_non_raw_acreage_classifications_list():
	lList = [
		'Apartment Horizontal',
		'Apartment Traditional',
		'Build for Rent (platted)',
		'High Density Assisted Living',
		'High Density Residential',
		'Manufactured Home',
		'Master Planned Community',
		'Mixed-Use',
		'Residential'
	]
	return lList

# Get list of Development Statuses
def get_development_status_list():
	lList = [
		'blank',
		'Planned',
		'Proposed',
		'Stabilized',
		'Under Construction',
		'Leave Blank'
	]
	return lList

# Get list of Multifamily Classifications for Development Status
def get_multifamily_classifications_list():
	lList = [
		'blank',
		'Apartment Horizontal',
		'Apartment Traditional',
		'Build for Rent (platted)'
	]
	return lList

# Return header list of Fields for TF to CSV spreadsheet
def get_tf_csv_header():
	header = ['Seller Entity',
			  'Seller Person',
			  'Seller Street',
			  'Seller City',
			  'Seller State',
			  'Seller Zip',
			  'Seller Country',
			  'Seller Phone',
			  'Seller Email',
			  'Buyer Entity',
			  'Buyer Person',
			  'Buyer Street',
			  'Buyer City',
			  'Buyer State',
			  'Buyer Zip',
			  'Buyer Country',
			  'Buyer Phone',
			  'Buyer Email',
			  'Acres__c',
			  'Buyer_Acting_As__c',
			  'City__c',
			  'Classification__c',
			  'Country__c',
			  'County__c',
			  'Description__c',
			  'General_Plan__c',
			  'Keyword_Group__c',
			  'Latitude__c',
			  'Lead_Parcel__c',
			  'Legal_Description__c',
			  'Location__c',
			  'Longitude__c',
			  'Lot_Description__c',
			  'Lot_Type__c',
			  'Lots__c',
			  'Market__c',
			  'Deal Name',
			  'Parcels__c',
			  'Recorded_Instrument_Number__c',
			  'Sale_Date__c',
			  'Sale_Price__c',
			  'Source__c',
			  'Source_ID__c',
			  'State__c',
			  'Subdivision__c',
			  'Submarket__c',
			  'Zipcode__c',
			  'Zoning__c',
			  'List_Date__c',
			  'List_Price__c',
			  'List Entity',
			  'List Agent',
			  'List Agent Phone',
			  'List Agent Email',
			  'AltAPN A',
			  'AltAPN B',
			  'PID__c',
			  'Residence Y-N',
			  'Terms__c',
			  'Recorded_Doc_URL__c',
			  'Notes',
			  'Lender',
			  'Loan_Amount__c',
			  'Loan_Date__c',
			  'Buyer Agent Entity',
			  'Buyer Agent',
			  'Buyer Agent Phone',
			  'Buyer Agent Email',
			  'MLS Status']
	return header

# Property Radar API fields
def get_proprad_api_fields():
	
	lList = [
		'AdvancedPropertyType',
		'APN',
		'City',
		'LastTransferSeller',
		'County',
		'FirstLenderOriginal',
		'FirstAmount',
		'FirstDate',
		'LastTransferRecDate',
		'LastTransferValue',
		'Latitude',
		'Longitude',
		'LotSizeAcres',
		'Owner',
		'OwnerFirstName',
		'OwnerLastName',
		'OwnerAddress',
		'OwnerCity',
		'OwnerState',
		'OwnerZipFive',
		'RadarID',
		'State',
		'Subdivision',
		'ZipFive',
		'Zoning'
	]

	# Convert list to a comma-separated string
	strList = ', '.join(lList)
	
	return strList

# Get list of Buyer Acting As
def get_buyer_acting_as_list():
	lList = [
		'blank',
		'Buyer',
		'Homebuilder',
		'Inv/Dev',
		'Lot Banker',
		'User'
	]	
	return lList

# Get list of Phoenix West Valley Submarkets
def get_phoenix_west_valley_submarkets_list():
	lList = [
	'West Surprise',
	'Tonopah',
	'Laveen',
	'South Peoria',
	'North I17',
	'Yavapai',
	'South Surprise',
	'North Goodyear',
	'Southwest',
	'North Phoenix',
	'Estrella',
	'Avondale-Tolleson',
	'Pecos',
	'Rainbow Valley',
	'North Peoria',
	'North Surprise',
	'Verrado',
	'Wittmann',
	'Sun Cities',
	'North Buckeye',
	'West Buckeye',
	'Northwest Maricopa County',
	'East Buckeye',
	'Lake Pleasant',
	'Glendale',
	'Central',
	'Goodyear',
	'North Central',
	'Anthem']

	return lList

# Get debt header list
def get_debt_header_list():
	headerDebtList = [
	' ', # 1 : A
	' ', # 2 : B
	'LAO Deal', # 3 : C
	'Market', # 4 : D
	'Stage', # 5 : E
	'PID', # 6 : F
	'Lot Type', # 7 : G
	'Classification', # 8 : H
	'Buyer Acting As', # 9 : I
	'Sale Date', # 10 : J
	'Sale Price', # 11 : K
	'Loan Amount', # 12 : L
	'Acres', # 13 : M
	'$/Ac', # 14 : N
	'Lot Count', # 15 : O
	'$/Lot', # 16 : P
	'$/FF', # 17 : Q
	'Beneficiary', # 18 : R
	'MVP Byr', # 19 : S
	'Buyer Enity', # 20 : T
	'Buyer Person', # 21 : U
	'Phone Byr', # 22 : V
	'Mobile Byr', # 23 : W
	'Email Byr', # 24 : X
	'MVP Slr', # 25 : Y
	'Seller Entity', # 26 : Z
	'Seller Person', # 27 : AA
	'Phone Slr', # 28 : AB
	'Mobile Slr', # 29 : AC
	'Email Slr', # 30 : AD
	'City', # 31 : AE
	'Submarket', # 32 : AF
	'Subdivision', # 33 : AG
	'Location', # 34 : AH
	'Description' # 35 : AI
	]
	return headerDebtList

# Get tax delingquency header list
def get_tax_delinquent_header_list():
	header_tax_delinquent = [
		'Map',
		'State',
		'LAO Market',
		'County',
		'LAO Submarket',
		'City',
		'APN',
		'Lot Acres',
		'Zoning',
		'Tax Delinquent Amt',
		'Purchase Date',
		'Purchase Amt',
		'Owner Entity',
		'Person Name',
		'Mail Address',
		'Mail City',
		'Mail State',
		'Mail ZIP',
		'Phone',
		'Mobile',
		'Email',
		'Lender',
		'Loan Amount',
		'Loan Rec Date',
		'Loan Purpose',
		'Is FLC',
		'FCL Doc Type',
		'FCL Rec Date',
		'FCL Stage',
		'Default Date',
		'Default Amt'
	]
	return header_tax_delinquent

# Get list of LAO markets
def get_lao_markets_list(abbreviation=True):
	import pandas as pd
	# Read the Excel file
	df = pd.read_excel(r'F:\Research Department\Code\Databases\LAO_Staff_Db_v03.xlsx')
	# Get unique values from marketAbb or Markets column and sort them
	if abbreviation:
		lMarkets = df['marketAbb'].fillna('').astype(str)
	else:
		lMarkets = df['Markets'].fillna('').astype(str)
	lMarkets = sorted([m for m in lMarkets.unique() if m != ''])
	return lMarkets

# Get LAO markets list for Market Mailers
def get_lao_markets_mailer_lists():
	# Get the markets and Scottsdale agents dict
	d = get_mm_market_dict()

	# Make list of markets and Scottsdale agents
	lOffice = []
	for office in d:
		if office == 'DFW':
			lOffice.append('DFW')
		elif office == 'NAZ':
			lOffice.append('Prescott')
		elif office == 'PHX':
			continue
		elif office == 'TPA':
			lOffice.append('Tampa')
		else:
			lOffice.append(d[office])

	# Make list of Scottsdale agents
	lScottsaleAgents = ['Campbell',
						'Heglie',
						'McCarville',
						'Medical',
						'Schwab',
						'Vogel',
						'Wuertz',
						'Xander',
						'Scottsdale']

	# Remove Vogel list
	lRemove_Vogel = ['Campbell',
				'Heglie',
				'McCarville',
				'Medical',
				'Schwab',
				'Wuertz',
				'Xander']
	
	# Schwab only list
	lSchwab_Only = ['Campbell',
				'Heglie',
				'McCarville',
				'Medical',
				'Northern Arizona',
				'Wuertz',
				'Xander']


	return lOffice, lScottsaleAgents, lRemove_Vogel, lSchwab_Only

def get_zonda_markets_list():
	lZonda_markets = [
		'ATL',
		'AUS',
		'BOI',
		'CLT',
		'DEN',
		'DFW',
		'HOU',
		'IEP',
		'JAX',
		'LVS',
		'NSH',
		'ORL',
		'PHX',
		'PRC',
		'RNO',
		'SAC',
		'SEA',
		'SLC',
		'SRQ',
		'TPA',
		'TUC',
		'IEP',
		'SAC',
		'SEA'
	]
	return lZonda_markets

def get_lao_employment_data_area_markets_list():
	"""
	Returns a list of markets from the LAO Employment Data spreadsheet.
	
	Returns:

		list: A list of market abbreviations.
	"""
	l = [
		'Alabama',
		'Huntsville',
		'California',
		'Riverside-San Bernardino-Ontario',
		'Sacramento-Roseville-Folsom',
		'Colorado',
		'Denver-Aurora-Centennial',
		'Florida',
		'Jacksonville',
		'North Port-Bradenton-Sarasota',
		'Orlando-Kissimmee-Sanford',
		'Tampa-St. Petersburg-Clearwater',
		'Georgia',
		'Atlanta-Sandy Springs-Roswell',
		'Idaho',
		'Boise City',
		'Kansas',
		'Missouri',
		'Kansas City',
		'Nevada',
		'Las Vegas-Henderson-North Las Vegas',
		'Reno',
		'New Mexico',
		'Albuquerque',
		'North Carolina',
		'Charlotte-Concord-Gastonia',
		'South Carolina',
		'Greenville-Anderson-Greer',
		'Spartanburg',
		'Tennessee',
		'Nashville-Davidson--Murfreesboro--Franklin',
		'Texas',
		'Austin-Round Rock-San Marcos',
		'Dallas-Fort Worth-Arlington',
		'Houston-Pasadena-The Woodlands',
		'Utah',
		'Ogden',
		'Provo-Orem-Lehi',
		'Salt Lake City-Murray',
		'Washington',
		'Seattle-Tacoma-Bellevue'
	]
	
	return l

# TF_QUERY_3 FIELDS #####################################################################
# Deal TF fields
def get_TF_deal_query_fields():
	lFields = ['Id',
			'AccountId__c',
			'AccountId__r.Name',
			'AccountId__r.FirstName',
			'AccountId__r.MiddleName__c',
			'AccountId__r.LastName',
			'AccountId__r.BillingAddress',
			'AccountId__r.BillingStreet',
			'AccountId__r.BillingCity',
			'AccountId__r.BillingState',
			'AccountId__r.BillingPostalCode',
			'AccountId__r.PersonMobilePhone',
			'AccountId__r.Phone',
			'AccountId__r.PersonEmail',
			'AccountId__r.Top100__c',
			'AccountId__r.Category__c',
			'AccountId__r.Description',
			'Acres__c',
			'Beneficiary__c',
			'Beneficiary__r.Id',
			'Beneficiary__r.Name',
			'Beneficiary_Contact__c',
			'Beneficiary_Contact__r.Name',
			'Buyer_Acting_As__c',
			'Case_Plan__c',
			'City__c',
			'City_Planner__c',
			'City_Planner__r.Name',
			'Classification__c',
			'County__c',
			'CreatedById',
			'CreatedDate',
			'Description__c',
			'Development_Status__c',
			'Digitized__c',
			'Encumbrance_Rating__c',
			'LastModifiedById',
			'LastModifiedDate',
			'Latitude__c',
			'Lead_Parcel__c',
			'Legal_Description__c',
			'List_Date__c',
			'Listing_Expiration_Date__c',
			'List_Price__c',
			'Loan_Amount__c',
			'Loan_Date__c',
			'Location__c',
			'Longitude__c',
			'Lot_Description__c',
			'Lot_Type__c',
			'Lot_Price_Rollup__c',
			'Lots__c',
			'Market__c',
			'Name',
			'OwnerId',
			'Owner_Entity__c',
			'Owner_Entity__r.Name',
			'Owner_Entity__r.BillingAddress',
			'Owner_Entity__r.BillingStreet',
			'Owner_Entity__r.BillingCity',
			'Owner_Entity__r.BillingState',
			'Owner_Entity__r.BillingPostalCode',
			'Owner_Entity__r.Phone',
			'Owner_Entity__r.Website',
			'OPR_Sent__c',
			'Parent_Opportunity__c',
			'Parent_Opportunity__r.PID__c',
			'Parent_Opportunity__r.Id',
			'Parent_Opportunity__r.Beneficiary__c',
			'Parent_Opportunity__r.Beneficiary_Contact__c',
			'Parent_Opportunity__r.Loan_Amount__c',
			'Parent_Opportunity__r.Loan_Date__c',
			'Parent_Opportunity__r.Encumbrance_Rating__c',
			'Parent_Opportunity__r.Credit_Bid_Amount__c',
			'Parcels__c',
			'PID__c',
			'Price_Per_Acre__c',
			'Price_Per_Lot__c',
			'P_Z_Description__c',
			'P_Z_Last_Event_Date__c',
			'Recorded_Instrument_Number__c',
			'RecordTypeId',
			'Report_Acres__c',
			'Research_Flag__c',
			'Sale_Date__c',
			'Sale_Price__c',
			'Source__c',
			'Source_ID__c',
			'StageName__c',
			'State__c',
			'Subdivision__c',
			'Submarket__c',
			'Type__c',
			'Verified__c',
			'Verified_By__c',
			'Verified_On__c',
			'Weighted_Avg_Price_Per_FF__c',
			'Zipcode__c',
			'Zoning__c',
			'Zoning_Applicant__c',
			'Zoning_Applicant__r.Name']

	fields_base = ', '.join(lFields)

	# Add other TF objects to fields
	fields_accounts_receivable = "(SELECT Id FROM Accounts_Receivable__r )"
	fields_commissions = "(SELECT Agent__c, Agent__r.Name, LAO_Agent__c, Agent__r.Phone, Agent__r.PersonEmail, Agent__r.Company__r.Id, Agent__r.Company__r.Name FROM Commissions__r)"
	fields_lot_details = "(Select Name, Acres__c, Lot_Count__c, Lot_Width__c, Lot_Depth__c, Price_per_parcel__c, Price_per_Front_Foot__c, Price_per_Lot__c From Lot_Details__r WHERE RecordTypeID != '012a0000001ZSieAAG')"
	fields_offer = "(SELECT Buyer__r.Id, Buyer__r.Name,  Buyer__r.Phone,  Buyer__r.PersonMobilePhone,  Buyer__r.PersonEmail,  Buyer__r.Top100__c, Buyer_Entity__r.Id, Buyer_Entity__r.Name, Buyer_Entity__r.Category__c, Offer_Status__c FROM Offers__r where Offer_Status__c = 'Accepted')"
	fields_notes = "(Select Id, LastModifiedDate, Title, Body From Notes)"
	fields_notes_and_attachments = "(Select Id, LastModifiedDate From NotesAndAttachments)"
	fields_package_information = "(Select Id, LastModifiedDate From Package_Information__r)"
	fields_requests = "(Select Id, CreatedDate, Map_Description__c, Record_Type_Name__c, Name From Request_Deal__r)"
	
	Fields = (f'{fields_base}, {fields_accounts_receivable}, {fields_commissions}, {fields_lot_details}, {fields_offer}, {fields_package_information}, {fields_requests}, {fields_notes}, {fields_notes_and_attachments}')

	return Fields

# Entity Account TF fields
def get_TF_entity_query_fields():
	lMain = [
			'Id',
			'Name',
			'BillingStreet',
			'BillingCity',
			'BillingState',
			'BillingPostalCode',
			'Category__c',
			'CreatedByID',
			'CreatedDate',
			'Description',
			'Fax',
			'Has_Address__c',
			'Has_Phone__c',
			'LastModifiedById',
			'LastModifiedDate',
			'LinkedIn_Url__c',
			'OwnerId',
			'ShippingStreet',
			'ShippingCity',
			'ShippingState',
			'ShippingPostalCode',
			'Phone',		
			'Website']
	lChild = [
			'ChildAccount__r.Id',
			'ChildAccount__r.FirstName',
			'ChildAccount__r.LastName',
			'ChildAccount__r.BillingStreet',
			'ChildAccount__r.MiddleName__c',
			'ChildAccount__r.Name',
			'ChildAccount__r.BillingCity',
			'ChildAccount__r.BillingState',
			'ChildAccount__r.Phone',
			'ChildAccount__r.PersonEmail']

	main = ', '.join(lMain)
	child = ', '.join(lChild)
	Fields = f'{main}, (Select {child} From AccountLinks__r)'

	return Fields

# Offer TF fields
def get_TF_offer_query_fields():
	lFields = [
		'Id,'
		'Buyer__r.Id',
		'Buyer__r.Name',
		'Buyer__r.FirstName',
		'Buyer__r.MiddleName__c',
		'Buyer__r.LastName',
		'Buyer__r.BillingStreet',
		'Buyer__r.BillingCity',
		'Buyer__r.BillingState',
		'Buyer__r.Phone',
		'Buyer__r.PersonEmail',
		'Buyer_Entity__r.Id',
		'Buyer_Entity__r.Name',
		'Buyer_Entity__r.BillingStreet',
		'Buyer_Entity__r.BillingCity',
		'Buyer_Entity__r.BillingState',
		'Buyer_Entity__r.BillingPostalCode',
		'Buyer_Entity__r.Phone',
		'Date_Accepted__c',
		'Offer_Price__c',
		'Offer_Status__c'
	]
	Fields = ', '.join(lFields)
	return Fields

def get_TF_contact_query_fields():
	lFields = [
	'Id',
	'Name',
	'FirstName',
	'MiddleName__c',
	'LastName',
	'BillingStreet',
	'BillingCity',
	'BillingState',
	'BillingPostalCode',
	'Category__c',
	'CC_ID__c',
	'Company__c',
	'Company__r.Name',
	'Company__r.BillingStreet',
	'Company__r.BillingCity',
	'Company__r.BillingState',
	'Company__r.BillingPostalCode',
	'Company__r.Category__c',
	'Company__r.Description',
	'Company__r.Phone',
	'Company__r.Website',
	'Company__r.LastModifiedDate',
	'CreatedByID',
	'CreatedDate',
	'Description',
	'Fax',
	'Has_Address__c',
	'Has_Email__c',
	'Has_Mobile__c',
	'Has_Phone__c',
	'Has_Phone_or_Email__c',
	'LastModifiedById',
	'LastModifiedDate',
	'LinkedIn_Url__c',
	'OwnerId',
	'PersonContactId',
	'PersonEmail',
	'PersonHasOptedOutOfEmail',
	'PersonHomePhone',
	'PersonMailingStreet',
	'PersonMailingCity',
	'PersonMailingState',
	'PersonMailingPostalCode',
	'Phone',
	'PersonMobilePhone',			
	'PersonTitle',
	'Top100__c',
	'Website'
	]
	Fields = ', '.join(lFields)
	return Fields

# Person Account TF fields
def get_TF_person_query_fields():
	lFields = [
		'Id',
		'Name',
		'FirstName',
		'MiddleName__c',
		'LastName',
		'BillingStreet',
		'BillingCity',
		'BillingState',
		'BillingPostalCode',
		'Category__c',
		'CC_ID__c',
		'Company__c',
		'Company__r.Name',
		'Company__r.BillingStreet',
		'Company__r.BillingCity',
		'Company__r.BillingState',
		'Company__r.BillingPostalCode',
		'Company__r.Category__c',
		'Company__r.Description',
		'Company__r.Phone',
		'Company__r.Website',
		'Company__r.LastModifiedDate',
		'CreatedByID',
		'CreatedDate',
		'Description',
		'Fax',
		'Has_Address__c',
		'Has_Email__c',
		'Has_Mobile__c',
		'Has_Phone__c',
		'Has_Phone_or_Email__c',
		'LastModifiedById',
		'LastModifiedDate',
		'LinkedIn_Url__c',
		'OwnerId',
		'PersonContactId',
		'PersonEmail',
		'PersonHasOptedOutOfEmail',
		'PersonHomePhone',
		'PersonMailingStreet',
		'PersonMailingCity',
		'PersonMailingState',
		'PersonMailingPostalCode',
		'Phone',
		'PersonMobilePhone',			
		'PersonTitle',
		'Top100__c',
		'Website'
	]
	Fields = ', '.join(lFields)
	return Fields

# Request TF fields
def get_TF_request_query_fields():
	lFields = [
		'Id',
		'Approver__c',
		'Approver__r.Name',
		'Assigned_Mapper__c',
		'Assigned_Mapper__r.Name',
		'CreatedById',
		'CreatedDate',
		'Deal_Lookup__c',
		'Deal_PID__c',
		'Description__c',
		'Map_Description__c',
		'LastModifiedById',
		'LastModifiedDate',
		'ModifyMapToDo__c',
		'Most_Recent_Activity__c',
		'Name',
		'Office__c',
		'OwnerId',
		'ParentId__c',
		'RecordTypeId',
		'Record_Type_Name__c',
		'Status__c',
		'Type__c'
	]
	Fields = ', '.join(lFields)
	return Fields

# Package TF fields
def get_TF_package_query_fields():
	lFields = [
		'Id',
		'DealID__c',
		'Field_Content__c',
		'Field_Name__c',
		'Field_Type__c'
	]
	Fields = ', '.join(lFields)
	return Fields

# Task TF fields
def get_TF_task_query_fields():
	lFields = [ 
		'Id',
		'WhoId',
		'Who.Name',
		'Who.Id',
		'Priority',
		'WhatId',
		'Subject',
		'Status',
		'Description'
	]
	Fields = ', '.join(lFields)
	return Fields

# Lot Detail TF fields
def get_TF_lot_detail_query_fields():
	lFields = [
		'Id',
		'Name',
		'Acres__c',
		'Lot_Count__c',
		'Lot_Depth__c',
		'Lot_Width__c',
		'Opportunity__c',
		'Price_per_Lot__c',
		'Price_per_parcel__c',
		'Price_per_Front_Foot__c',
		'Total_FF_Dollars__c',
		'RecordTypeID'
	]
	Fields = ', '.join(lFields)
	return Fields

